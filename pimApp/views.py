from django.shortcuts import render
from django.http import JsonResponse
from .global_service import DatabaseModel
from .models import *
from django.views.decorators.csrf import csrf_exempt
from .custom_middleware import get_current_client,get_current_user
import os
from django.conf import settings # type: ignore
from rest_framework.parsers import JSONParser # type: ignore
from openpyxl import Workbook # type: ignore
from django.http import HttpResponse # type: ignore
import pytz
from threading import Thread
us_timezone = pytz.timezone("US/Eastern")  # Change this based on your requirement
# def v1(request):
#     data = dict()
#     data['status'] = 'hello'
import phonenumbers
def is_valid_phone_number(phone):
    try:
        parsed_number = phonenumbers.parse(phone, None)
        return phonenumbers.is_valid_number(parsed_number)
    except phonenumbers.NumberParseException:
        return False
def obtainDashboardForClient(request):
    #  = request.GET.get('')
    data = dict()
    data['product_count'] = DatabaseModel.count_documents(product.objects)
    # data['brand_count'] = DatabaseModel.count_documents(Brand.objects,{'industry_type_id':industry_type_id})
    pipeline = [
        # {
        #     "$match":{'industry_type_id':ObjectId()}
        # },
    {
        '$lookup': {
            'from': 'product',
            'localField': '_id',
            'foreignField': 'brand_id',
            'as': 'product'
        }
    }, 
    {
            '$unwind': {
                'path': '$product',
                'preserveNullAndEmptyArrays': True
            }
    },
   {
        '$group': {
            "_id":"$_id",   
            "id": { '$first': '$_id'},   
            "name": { '$first': '$name'},   
            "product_list": { '$addToSet': '$product._id'},   
            }
        },{
            '$project':{
                '_id':0,
                'name':1,
                'id':{'$toString':'$id'},
                'product_count':{'$size':'$product_list'}
                
            }
        }
    ]
    Brand_result = list(brand.objects.aggregate(*pipeline))
    data['brand_count'] = len(Brand_result)
    data['brand_list'] =(Brand_result)
    pipeline = [
        {
        '$lookup': {
            'from': 'category',
            'localField': 'levels',
            'foreignField': '_id',
            'as': 'category'
        }
    }, 
    {
            '$unwind': {
                'path': '$category',
                'preserveNullAndEmptyArrays': True
            }
    },
      {
        '$lookup': {
            'from': 'product',
            'localField': '_id',
            'foreignField': 'category_id',
            'as': 'product'
        }
    }, 
    {
            '$unwind': {
                'path': '$product',
                'preserveNullAndEmptyArrays': True
            }
    },
    {
        '$group': {
            "_id":"$_id",   
            "category_id": { '$first': '$category_id'},   
            "id": { '$first': '$_id'},   
            "product_list": { '$addToSet': '$product._id'},   
            "category_list": { '$addToSet': '$category.name'},   
            }
            
        },
        {
        '$project': {
            '_id': 0,
            'str_category': {
                '$concat': [
                    { '$toString': '$category_id' }, 
                    ' - (', 
                    {
                        '$reduce': {
                            'input': { '$reverseArray': '$category_list' },
                            'initialValue': '',
                            'in': {
                                '$cond': {
                                    'if': { '$eq': [{ '$literal': '' }, '$$value'] },
                                    'then': '$$this',
                                    'else': { '$concat': ['$$value', ' > ', '$$this'] }
                                }
                            }
                        }
                    },
                    ')'
                ]
            },
            'category_id': 1,
            'id': { '$toString': '$id' },
            'product_count': { '$size': '$product_list' }
        }
    }
    ]
    result = list(category_config.objects.aggregate(*pipeline))
    data['category_count'] = len(result)
    data['category_list'] = result
    category_config_parent = DatabaseModel.list_documents(
    category_config.objects,
    {'levels': [0]}
)
    data['parent_category_list'] = dict()
    for i in category_config_parent:
        parent_level_dict = dict()
        parent_level_dict['name'] = i.name
        parent_level_dict['product_count'] = 0
        category_config_child = DatabaseModel.list_documents(
            category_config.objects,
            {'levels__in': [i.id]}
        )
        for jj in category_config_child:
            parent_level_dict['product_count'] +=DatabaseModel.count_documents(product.objects,{'category_id__in':[jj.id]})
        data['parent_category_list'].append(parent_level_dict)
    return JsonResponse(data,safe=False)



@csrf_exempt
def createBrand(request):
    # json_req = JSONParser().parse(request)
    # vendor_id = json_req.get('vendor_id')
    name = request.POST.get('name')
    country_of_origin = request.POST.get('country_of_origin')
    status = request.POST.get('status')
    website = request.POST.get('website')
    description = request.POST.get('description')
    warranty_details = request.POST.get('warranty_details')
    warranty_details_based = request.POST.get('warranty_details_based')
    name = str(name).title()
    logo = request.FILES.get('logo')
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    if logo:
        upload_result = cloudinary.uploader.upload(logo,folder="KM-DIGI/image")
        logo = upload_result.get("secure_url")
    else:
        logo = ""
    client_id = get_current_client()
    brand_obj = DatabaseModel.get_document(brand.objects,{'name':name,'client_id':client_id})
    data = dict()
    if brand_obj:
        data['is_created'] = False
        data['error'] = "This Brand Already Exists"
        return data
    else:
        # if vendor_id == "":
        brand_obj = DatabaseModel.save_documents(brand,{'name':name,'logo':str(logo),'country_of_origin':country_of_origin,'warranty_details':str(warranty_details),'warranty_details_based':warranty_details_based,'status':status,'website':website,'description':description})
        DatabaseModel.save_documents(brand_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Created','brand_id':brand_obj.id})

    data['is_created'] = True
    return data

@csrf_exempt
def brandUpdate(request):
    json_req = JSONParser().parse(request)
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    client_id = get_current_client()
    brand_id = json_req['update_obj']['id']
    # json_req['update_obj']['attribute_list'] = [ObjectId(i)   for i in json_req['update_obj']['attribute_list']]
    del json_req['update_obj']['id']
    # if json_req['update_obj']['logo']:
    #     upload_result = cloudinary.uploader.upload(json_req['update_obj']['logo'],folder="KM-DIGI/image")
    #     json_req['update_obj']['logo'] = upload_result.get("secure_url")
    brand_obj = DatabaseModel.update_documents(brand.objects,{'id':brand_id},json_req['update_obj'])
    DatabaseModel.save_documents(brand_log,{'user_id':user_login_id,'client_id':client_id,'action':'Updated','brand_id':brand_obj.id})

    data = dict()
    data['is_updated'] = True
    return data
import hashlib
import cloudinary.api
@csrf_exempt
def brandUpdateLogo(request):
    logo = request.FILES.get('logo')
    upload_check = cloudinary.uploader.upload(logo, folder="logo", resource_type="image", use_filename=True, unique_filename=False, overwrite=False)
    etag = upload_check.get("etag")  
    logo.seek(0)
    existing_images = cloudinary.api.resources(type="upload", max_results=100)
    for img in existing_images.get("resources", []):
        if img.get("etag") == etag:  
            data = dict()
            data['is_updated'] = True
            data['logo'] = upload_check["secure_url"]
            return data
    data = dict()
    data['is_updated'] = True
    data['logo'] = upload_check["secure_url"]
    return data

def obtainBrand(request):
    client_id = get_current_client()
    search_term = request.GET.get('search')
    data = dict()
    brand_id = request.GET.get('id')
    filter = request.GET.get('filter')
    if filter == "true" or filter == None:
        reverse_check = -1
    else:
        reverse_check = 1
    if brand_id:
        match = {
            '$match':{'client_id':ObjectId(client_id),'_id':ObjectId(brand_id)}
        }
    else:
        match = {
            '$match':{'client_id':ObjectId(client_id)}
        }
    pipeline = [
        match,
        
    {
            '$group': {
                "_id":'$_id',
                'brand_id':{'$first':'$brand_id'},
                'name':{'$first':'$name'},
                'country_of_origin':{'$first':'$country_of_origin'},
                'warranty_details':{'$first':'$warranty_details'},
                'warranty_details_based':{'$first':'$warranty_details_based'},
                'status':{'$first':'$status'},
                'website':{'$first':'$website'},
                'description':{'$first':'$description'},
                'logo':{'$first':'$logo'},
                'attribute_list':{'$first':'$attribute_list'},
                
        }
        },{
            '$project':{
                '_id':1,
                'brand_id':1,
                'name':1,
                'country_of_origin':1,
                'warranty_details':1,
                'warranty_details_based':1,
                'status':1,
                'website':1,
                'description':1,
                'logo':1,
                'attribute_list':1
            }
        }
    ,{
    '$match': {
    '$or': [
        { 'name': { '$regex': search_term, '$options': 'i' } }
]
    }
  },
    {
        '$sort': { '_id': reverse_check }
    },
    ]
    brand_list = list(brand.objects.aggregate(*pipeline))
    for i in brand_list:
        i['id'] = str (i['_id'])
        if brand_id:
            products_list = DatabaseModel.list_documents(product.objects,{'brand_id':i['id'],"client_id":client_id})
            i['product_count'] = len(products_list)
        i['sku_count'] = 0
        attribute_list = []
        if i['attribute_list'] == None:
            i['attribute_list'] = []
        for j in  i['attribute_list']:
            Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'id':j})
            if Attribute_obj:
                attribute_list.append({'id':str(Attribute_obj.id),'name':Attribute_obj.name,'values':Attribute_obj.values})
        i['attribute_list'] = attribute_list
        del i['_id']
    data['brand_list'] = brand_list
    data['brand_count'] = len(data['brand_list'])
    return data


@csrf_exempt
def createVendor(request):
    name = request.POST.get('name')
    client_id = get_current_client()
    contact_info_email = request.POST.get('contact_info_email')
    contact_info_phone = request.POST.get('contact_info_phone')
    business_type = request.POST.get('business_type')
    website = request.POST.get('website')
    description = request.POST.get('description')
    address = request.POST.get('address')
    city = request.POST.get('city')
    industry_info = request.POST.get('industry_info')
    tax_info = request.POST.get('tax_info')
    departments = request.POST.get('departments')
    country_code = request.POST.get('country_code')
    if country_code and contact_info_phone:
        contact_info_phone = country_code + " "+contact_info_phone
    # department_name = request.POST.get('department_name')
    # email = request.POST.get('email')
    # phone_number = request.POST.get('phone_number')
    logo = request.FILES.get('logo')
    if logo:
        upload_result = cloudinary.uploader.upload(logo,folder="KM-DIGI/image")
        logo = upload_result.get("secure_url")
    else:
        logo = ""
    name = name.title()
    Vendor_obj = DatabaseModel.get_document(Vendor.objects,{'name':name,'client_id':client_id})
    data = dict()
    if Vendor_obj:
        data['is_created'] = False
        data['error'] = "This Vendor Already Exists"
        return data
    else:
        if business_type =="" or business_type == None:
            Vendor_obj = DatabaseModel.save_documents(Vendor,{'name':name,'logo':str(logo),'address':address,'city':city,'contact_info_email':contact_info_email,'contact_info_phone':contact_info_phone,'description':description,'website':website,'tax_info':tax_info,'industry_info':industry_info,'client_id':ObjectId(client_id)})
        else:
            Vendor_obj = DatabaseModel.save_documents(Vendor,{'name':name,'logo':str(logo),'business_type':ObjectId(business_type),'address':address,'city':city,'contact_info_email':contact_info_email,'contact_info_phone':contact_info_phone,'description':description,'website':website,'tax_info':tax_info,'industry_info':industry_info,'client_id':ObjectId(client_id)})
        departments = json.loads(departments)
        for i in departments:
            if i['country_code'] and i['phone_number']:
                i['phone_number'] = i['country_code'] + " "+i['phone_number']
            Vendor_obj.departments.append(ContactInfo(department_name=str(i['department_name']), email=str(i['email']), phone_number=str(i['phone_number'])))
        Vendor_obj.save()
        user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
        DatabaseModel.save_documents(vendor_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Created','vendor_id':Vendor_obj.id})
    data['is_created'] = True
    return data


@csrf_exempt
def vendorUpdate(request):
    json_req = JSONParser().parse(request)
    vendor_id = json_req['update_obj']['id']
    if 'business_type' in json_req.get('update_obj', {}):
        if json_req['update_obj']['business_type'] == "" or json_req['update_obj']['business_type'] == None:  
            del json_req['update_obj']['business_type']  
        else:
            json_req['update_obj']['business_type'] = ObjectId(json_req['update_obj']['business_type'])
    del json_req['update_obj']['id']
    departments = json_req['update_obj']['departments']
    if json_req['update_obj']['logo']:
        upload_result = cloudinary.uploader.upload(json_req['update_obj']['logo'],folder="KM-DIGI/image")
        json_req['update_obj']['logo'] = upload_result.get("secure_url")
    json_req['update_obj']['departments'] = []
    if 'country_code'in json_req['update_obj']:
        json_req['update_obj']['contact_info_phone'] = json_req['update_obj']['country_code'] + " "+ json_req['update_obj']['contact_info_phone']
        del json_req['update_obj']['country_code']
    DatabaseModel.update_documents(Vendor.objects,{'id':vendor_id},json_req['update_obj'])
    Vendor_obj = DatabaseModel.get_document(Vendor.objects,{'id':vendor_id})
    data = dict()
    for i in departments:
        if 'country_code' in i and i['country_code'] and i['phone_number'] :
            i['phone_number'] = i['country_code'] + " "+i['phone_number']
        Vendor_obj.departments.append(ContactInfo(department_name=i['department_name'], email=str(i['email']), phone_number=str(i['phone_number'])))
    Vendor_obj.save()
    client_id = get_current_client()
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    DatabaseModel.save_documents(vendor_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Updated','vendor_id':Vendor_obj.id})
    data['is_updated'] = True
    return data


def obtainVendor(request):
    client_id = get_current_client()
    data = dict()
    search_term = request.GET.get('search')
    vendor_id = request.GET.get('id')
    filter = request.GET.get('filter')
    if filter == "true" or filter == None:
        reverse_check = -1
    else:
        reverse_check = 1
    if vendor_id:
        match = {
            '$match':{'client_id':ObjectId(client_id),'_id':ObjectId(vendor_id)}
        }
        group_obj = {
            '$group': {
                "_id":'$_id',
                'vendor_id':{'$first':'$vendor_id'},
                'name':{'$first':'$name'},
                'logo':{'$first':'$logo'},
                'business_type_id':{'$first':'$business_type._id'},
                'business_type_name':{'$first':'$business_type.name'},
                'city':{'$first':'$city'},
                'address':{'$first':'$address'},
                'contact_info_email':{'$first':'$contact_info_email'},
                'contact_info_phone':{'$first':'$contact_info_phone'},
                'description':{'$first':'$description'},
                'website':{'$first':'$website'},
                'tax_info':{'$first':'$tax_info'},
                'industry_info':{'$first':'$industry_info'},
                'departments':{'$first':'$departments'},
        }
        }
    else:
        match = {
            '$match':{'client_id':ObjectId(client_id)}
        }
        group_obj = {
            '$group': {
                "_id":'$_id',
                'vendor_id':{'$first':'$vendor_id'},
                'name':{'$first':'$name'},
                'logo':{'$first':'$logo'},
                'business_type_id':{'$first':'$business_type._id'},
                'business_type_name':{'$first':'$business_type.name'},
                'industry_info':{'$first':'$industry_info'},
                'website':{'$first':'$website'},
                'contact_info_email':{'$first':'$contact_info_email'},

        }
        }
    pipeline = [
        match,{
        '$lookup': {
            'from': 'business_type',
            'localField': 'business_type',
            'foreignField': '_id',
            'as': 'business_type'
        }
    }, 
    {
            '$unwind': {
                'path': '$business_type',
                'preserveNullAndEmptyArrays': True
            }
        },
   group_obj
       ,{
    '$match': {
    '$or': [
        { 'name': { '$regex': search_term, '$options': 'i' } }    
]
    }
  },
    {
        '$sort': { '_id': reverse_check }
    },
    ]
    vendor_list = list(Vendor.objects.aggregate(*pipeline))
    for i in vendor_list:
        i['id'] = str (i['_id'])
        if i['business_type_id']:
            i['business_type_id'] = str(i['business_type_id'])
        else:
            del i['business_type_id'] 
        del i['_id']
        if i.get('contact_info_phone'): 
            contact_info_phone = i['contact_info_phone'].split(" ", 1) 
            if len(contact_info_phone) == 2: 
                i['country_code'] = contact_info_phone[0]
                i['contact_info_phone'] = contact_info_phone[1]
            else:
                i['country_code'] = ""
        if 'departments' in i:
            for xx in i['departments']:
                if xx.get('phone_number'): 
                    contact_info_phone = xx['phone_number'].split(" ", 1) 
                    if len(contact_info_phone) == 2: 
                        xx['country_code'] = contact_info_phone[0]
                        xx['phone_number'] = contact_info_phone[1]
                    else:
                        xx['country_code'] = ""
    data['vendor_list'] = vendor_list
    data['vendor_count'] = len(data['vendor_list'])
    return data


@csrf_exempt
def createCategory(request):
    json_req = JSONParser().parse(request)
    category_name = json_req['name']   
    category_name = category_name.title()
    client_id = ObjectId(get_current_client())
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    category_config_id = json_req.get('category_config_id')
    data = dict()
    if category_config_id == "":
        category_obj = DatabaseModel.get_document(category.objects,{'name':category_name,'client_id':client_id})
        if category_obj == None:
            category_obj = DatabaseModel.save_documents(category,{'name':category_name,'client_id':client_id})
        category_config_obj = DatabaseModel.save_documents(category_config,{'name':category_name,'levels':[category_obj.id],'client_id':client_id,'end_level':True}) 
        DatabaseModel.save_documents(category_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Created','category_config_id':category_config_obj.id})
        data['is_created'] = True
        data['status'] = "Category Levels Created Successfully"
        return JsonResponse(data,safe=False)
    else:
        category_config_obj = DatabaseModel.get_document(category_config.objects,{'id':category_config_id,'client_id':client_id})
        category_obj = DatabaseModel.get_document(category.objects,{'name':category_name,'client_id':client_id})
        if category_obj == None:
            category_obj = DatabaseModel.save_documents(category,{'name':category_name,'client_id':client_id})
        category_level = list()
        if category_config_obj:
            category_level = category_config_obj.levels
            category_config_obj.end_level = False
            category_config_obj.save()
        category_level.append(category_obj.id)
        category_config_obj = DatabaseModel.save_documents(category_config,{'name':category_name,'levels':category_level,'client_id':client_id,'end_level':True})
        # if category_config_obj == None:
        #     category_obj = DatabaseModel.save_documents(category_config,{'name':category_name,'levels':category_level,'client_id':client_id})
        DatabaseModel.save_documents(category_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Created','category_config_id':category_config_obj.id})
    data['is_created'] = True
    data['status'] = "Category Levels Created Successfully"
    return JsonResponse(data,safe=False)
    return data

from django.http import JsonResponse
from bson import ObjectId
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def obtainCategory(request):
    data = dict()
    client_id = get_current_client()
    pipeline = [
        {
            '$match':{'client_id':ObjectId(client_id)}
        },
        {
            '$lookup': {
                'from': 'category',
                'localField': 'levels',
                'foreignField': '_id',
                'as': 'category'
            }
        },
        {
            '$unwind': {
                'path': '$category',
            }
        },
        {
            '$group': {
                "_id": '$_id',
                'levels': {'$push': {'id': '$category._id', "name": '$category.name'}}
            }
        },{
            '$sort':{'_id':-1}
        }
    ]

    category_config_list = list(category_config.objects.aggregate(pipeline))
    hierarchy = {}
    category_order = [str(entry["_id"]) for entry in category_config_list]
    for entry in category_config_list:
        levels = entry["levels"]
        current = hierarchy 
        path_tracker = [] 
        _ids_collected = set() 
        for level in levels:
            level_id = str(level["id"])
            level_name = level["name"]
            _ids_collected.add(str(entry["_id"])) 
            path_tracker.append(level_name) 
            unique_path = " > ".join(path_tracker) 
            if unique_path not in current:
                current[unique_path] = {
                    "id": level_id,
                    "name": level_name,
                    "_ids": set(),
                    "children": {}
                }
            current[unique_path]["_ids"].update(_ids_collected)
            current = current[unique_path]["children"] 

    def build_tree(node):
        sorted_children = sorted(
            [build_tree(child) for child in node["children"].values()],
            key=lambda x: ObjectId(x["id"]) if ObjectId.is_valid(x["id"]) else ObjectId()
        )
        ids_list = list(node["_ids"])
        # for i in ids_list:
        #     i = ObjectId(i)
        category_config_obj = DatabaseModel.list_documents(category_config.objects, {'name': node["name"],'id__in':ids_list})
        
        if not category_config_obj:
            return {
                "id": node["id"],
                "name": node["name"],
                "_ids": sorted(node["_ids"], key=lambda x: category_order.index(x) if x in category_order else float('inf')),
                "config_id": None,
                "levels_str": "",
                "product_count": 0,
                "sub_cat_product_count": 0,
                "children": sorted_children
            }
        category_config_obj = category_config_obj.order_by('-_id')
        config_obj = category_config_obj[0]
        levels_str = " > ".join(zz.name for zz in config_obj.levels)
        node["_ids"] = {str(id) for id in node["_ids"] if str(id) != str(config_obj.id)}
        parent_level_ids = []
        for zi in config_obj.levels:
            category_config_obj_name = DatabaseModel.get_document(category_config.objects,{'name':zi.name,'client_id':client_id})
            if config_obj.id != category_config_obj_name.id:
                parent_level_ids.append(str(category_config_obj_name.id))
        return {
            "id": node["id"],
            "name": node["name"],
            "_ids": sorted(node["_ids"], key=lambda x: category_order.index(x) if x in category_order else float('inf')),
            "config_id": str(config_obj.id),
            "levels_str": levels_str,
            'parent_level_ids':parent_level_ids,
            "product_count": DatabaseModel.count_documents(product.objects, {'category_id__in': [ObjectId(config_obj.id)]}),
            "sub_cat_product_count": DatabaseModel.count_documents(product.objects, {'category_id__in': list(node["_ids"])}),
            "children": sorted_children
            }
    hierarchy_list = sorted(
    [build_tree(node) for node in hierarchy.values()],
    key=lambda x: -int(x["id"]) if isinstance(x["id"], str) and x["id"].isdigit() else float('-inf')
)
    data['category_levels'] = hierarchy_list
    return data




import pandas as pd
import json
@csrf_exempt
def importXlFiles(request):
    pass
#     data = dict()
#     user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
#     # field_data = request.POST.get('field_data')
#     file_path = request.FILES.get('file')
#     data['status'] = False
#     client_id = get_current_client()
#     file = request.FILES['file']
#     try:
#         if file.name.endswith('.xlsx'):
#             df = pd.read_excel(file)
#         elif file.name.endswith('.csv') or file.name.endswith('.txt'):
#             df = pd.read_csv(file)
#         elif file.name.endswith('.ods'):
#             df = pd.read_excel(file, engine='odf')
#     except FileNotFoundError:
#         data['status'] = False
#         return data
#     for _, row in df.iterrows():
#         def clean_value(value, default=""):
#             return value if pd.notna(value) else default
#         def clean_value_float(value, default="0.0"):
#             return value if pd.notna(value) else default
#         breadcrumb = clean_value(row.get('Bread Crumb', "") or row.get('breadcrumb', ""))
#         vendor_name = clean_value(row.get('Vendor Name', "") or row.get('vendor name', ""))
#         name = clean_value(row.get('Product Name', "") or row.get('product name', ""))
#         sku = clean_value(row.get('sku', ""))
#         mpn = clean_value(row.get('mpn', "") or row.get('MPN', ""))
#         upc = clean_value(row.get('upc', ""))
#         end_level = clean_value(row.get('Category', "") or row.get('category', ""))
#         brand_name = clean_value(row.get('Brand', "") or row.get('brand', ""))
#         manufacture = clean_value(row.get('Manufacture', "") or row.get('Manufacture', ""))
#         ean = clean_value(row.get('ean', "") or row.get('ean', ""))
#         gtin = clean_value(row.get('gtin', "") or row.get('gtin', ""))
#         unspc = clean_value(row.get('unspc', "") or row.get('unspc', ""))
#         stock_quantity = int(stock_quantity) if pd.notna(stock_quantity := row.get('Stock Quantity', 0)) else ""
#         product_id = clean_value(row.get('product_id', ""))
#         model = clean_value(row.get('model', ""))
#         personalized_short_description = clean_value(row.get('personalized_short_description', ""))
#         long_description = clean_value(row.get('long_description', ""))
#         personalized_long_description = clean_value(row.get('personalized_long_description', ""))
#         Compliance = clean_value(row.get('Compliance', ""))
#         Prop65 = clean_value(row.get('Prop65', ""))
#         short_description = clean_value(row.get('short description', ""))
#         esg = clean_value(row.get('esg', ""))
#         Hazardous = clean_value(row.get('Hazardous', ""))
#         service_warranty = clean_value(row.get('service warrenty', ""))
#         product_warranty = clean_value(row.get('product warrenty', ""))
#         country_of_origin = clean_value(row.get('country of origin', ""))
#         currency = clean_value(row.get('currency', ""))
#         selling_price = clean_value(row.get('selling price', ""))
#         discount_price = clean_value(row.get('discount price', ""))
#         certifications = clean_value(row.get('certifications', ""))
#         application = clean_value(row.get('application', ""))
#         msrp = clean_value_float(row.get('msrp', "0.0"))
#         vendor_obj = DatabaseModel.get_document(Vendor.objects,{'name':vendor_name})
#         if vendor_obj == None:
#             vendor_obj = DatabaseModel.save_documents(Vendor,{'name':vendor_name})
#         brand_obj = DatabaseModel.get_document(brand.objects,{'name':brand_name})
#         if brand_obj == None:
#             brand_obj = DatabaseModel.save_documents(brand,{'name':brand_name})
#         manufacture_obj = DatabaseModel.get_document(Manufacture.objects,{'name':brand_name})
#         if manufacture_obj == None:
#             manufacture_obj = DatabaseModel.save_documents(Manufacture,{'name':manufacture})
#         levels = list()
#         for i in end_level:
#             category_obj = DatabaseModel.get_document(category.objects,{'name':i})
#             if category_obj:
#                 levels.append(category_obj.id)
#             else:
#                 category_obj = DatabaseModel.save_documents(category,{'name':i})
#                 levels.append(category_obj.id)
#         category_config_obj = DatabaseModel.get_document(category_config.objects,{'levels':levels})
#         if category_config_obj:
#             category_id = category_config_obj.id
#         else:
#             category_config_obj = DatabaseModel.save_documents(category_config,{'levels':levels})
#             category_id = category_config_obj.id
#         data['is_created'] = True
#         data['status'] = "Category Levels Created Successfully"
#         product_obj = DatabaseModel.save_documents(product,{
#             "product_id" : str(product_id),
#             "mpn" : str(mpn),
#             "sku" : str(sku),
#             "upc" : str(upc),
#             "ean" : str(ean),
#             "gtin" : str(gtin),
#             "unspc": str(unspc) ,
#             "model": str(model) ,
#             "vendor_id": vendor_obj.id ,
#             "brand_id": brand_obj.id ,
#             "manufacture_id": manufacture_obj.id ,
#             "category_id": [category_id] ,
#             "breadcrumb": str(breadcrumb) ,
#             "name": str(name),
#             "short_description": str(short_description) ,
#             "personalized_short_description" : str(personalized_short_description),
#             "long_description" : str(long_description),
#             "personalized_long_description" : str(personalized_long_description),
#             "feature_list" :[],
#             "attribute_list" :[],
#             "related_products" : [],
#             "application" : str(application),
#             "certifications" : str(certifications),
#             "Compliance" : str(Compliance),
#             "Prop65" : str(Prop65),
#             "esg" : str(esg),
#             "Hazardous" : str(Hazardous),
#             "service_warranty" : str(service_warranty),
#             "product_warranty" : str(product_warranty),
#             "country_of_origin" : str(country_of_origin),
#             "currency" : str(currency),
#             "msrp" : str(msrp),
#             "selling_price" : str(selling_price),
#             "discount_price" : str(discount_price),
#             "attachment_list" : [],
#             "image_list" : [],
#             "video_list": [],
#             })
#         process_row(row, product_obj)
#     data = dict()
#     data['status'] = True
#     return JsonResponse(data,safe=False)
# def process_row(row_dict, product_obj):
#     feature_number = 1
#     def clean_value(value, default=""):
#         return value if pd.notna(value) else default
#     while f'features {feature_number}' in row_dict:
#         feature = row_dict[f'features {feature_number}']
#         feature = clean_value(feature, default="")
#         if feature:
#             product_obj.feature_list.append(Feature(name=str(feature)))  # Use EmbeddedDocument instance
#         else:
#             break
#         feature_number += 1

#     attribute_number = 1
#     while f'attribute name {attribute_number}' in row_dict and f'attribute value {attribute_number}' in row_dict:
#         attribute_name = row_dict[f'attribute name {attribute_number}']
#         attribute_value = row_dict[f'attribute value {attribute_number}']
#         attribute_name = clean_value(attribute_name, default="")
#         attribute_value = clean_value(attribute_value, default="")
#         if attribute_name and attribute_value:
#             product_obj.attribute_list.append(Attribute__(name=attribute_name, value=str(attribute_value)))  # Use EmbeddedDocument instance
#         else:
#             break
#         attribute_number += 1

#     image_number = 1
#     while f'image name {image_number}' in row_dict and f'image url {image_number}' in row_dict:
#         image_name = row_dict[f'image name {image_number}']
#         image_url = row_dict[f'image url {image_number}']
#         image_url = clean_value(image_url, default="")
#         image_name = clean_value(image_name, default="")
#         if image_name and image_url:
#             product_obj.image_list.append(Image(name=image_name, url=image_url))  # Use EmbeddedDocument instance
#         else:   
#             break
#         image_number += 1

#     video = 1
#     while f'video name {video}' in row_dict and f'video url {video}' in row_dict:
#         video_name = row_dict[f'video name {video}']
#         video_url = row_dict[f'video url {video}']
#         video_name = clean_value(video_name, default="") 
#         video_url = clean_value(video_url, default="")
#         if video_name and video_url:
#             product_obj.video_list.append(Video(name=video_name, url=video_url))  # Use EmbeddedDocument instance
#         else:
#             break
#         video += 1

#     attachment = 1
#     while f'attachment name {attachment}' in row_dict and f'attachment url {attachment}' in row_dict:
#         attachment_name = row_dict[f'attachment name {attachment}']
#         attachment_url = row_dict[f'attachment url {attachment}']
#         attachment_name = clean_value(attachment_name, default="")
#         attachment_url = clean_value(attachment_url, default="")
#         if attachment_name and attachment_url:
#             product_obj.attachment_list.append(Attachment(name=attachment_name, url=str(attachment_url)))  # Use EmbeddedDocument instance
#         else:
#             break
#         attachment += 1

#     repro = 1
#     while f'related name {repro}' in row_dict and f'related url {repro}' in row_dict:
#         repro_name = row_dict[f'related name {repro}']
#         repro_url = row_dict[f'related url {repro}']
#         repro_url = clean_value(repro_url, default="")
#         repro_name = clean_value(repro_name, default="")
#         if repro_name and repro_url:
#             product_obj.related_products.append(RelatedProduct(name=repro_name, url=repro_url))  # Use EmbeddedDocument instance
#         else:
#             break
#         repro += 1
#     product_obj.save()
#     return 0
from django.core.files.base import ContentFile
import base64
@csrf_exempt
def createProduct(request):
    client_id = get_current_client()
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')

    category_config_obj = DatabaseModel.get_document(category_config.objects,{'id':request.POST.get('config_id')})
    category_level = ""
    if category_config_obj ==None:
        category_level = ""
    else:
        for j in category_config_obj.levels:
            category_level += j.name + ">"
        category_level = category_level.rstrip(">")
    feature_list = request.POST.get('feature_list')
    lst = feature_list.split("\n") 
    product_obj = DatabaseModel.get_document(product.objects,{'sku':str(request.POST.get('sku')),'client_id':client_id})
    if product_obj:
        data = dict()
        data['is_created'] = False
        data['error'] = "Product Already Exist"
        return data
    product_dict = {
            "product_id" : str(request.POST.get('product_id')),
            "mpn" : str(request.POST.get('mpn')),
            "sku" : str(request.POST.get('sku')),
            "upc" : str(request.POST.get('upc')),
            "ean" : str(request.POST.get('ean')),
            "gtin" : str(request.POST.get('gtin')),
            "unspc": str(request.POST.get('unspc')) ,
            "model": str(request.POST.get('model')) ,
            "client_id":ObjectId(client_id),
            "category_id":[category_config_obj.id] ,
            "breadcrumb": str(category_level) ,
            "name": str(request.POST.get('name')),
            "short_description": str(request.POST.get('short_description')) ,
            "personalized_short_description" : str(request.POST.get('personalized_short_description')),
            "long_description" : str(request.POST.get('long_description')),
            "personalized_long_description" : str(request.POST.get('personalized_long_description')),
            "feature_list" :[item for item in lst if item]  ,
            "attribute_list" : [],
            "related_products" : [],
            "application" : str(request.POST.get('application')),
            "certifications" : str(request.POST.get('certifications')),
            "Compliance" : str(request.POST.get('Compliance')),
            "Prop65" : str(request.POST.get('Prop65')),
            "esg" : str(request.POST.get('esg')),
            "Hazardous" : str(request.POST.get('Hazardous')),
            "service_warranty" : str(request.POST.get('service_warranty')),
            "product_warranty" : str(request.POST.get('product_warranty')),
            "country_of_origin" : str(request.POST.get('country_of_origin')),
            "currency" : str(request.POST.get('currency')),
            "msrp" : str(request.POST.get('msrp')),
            "selling_price" : str(request.POST.get('selling_price')),
            "discount_price" : str(request.POST.get('discount_price')),
            "attachment_list" : [],
            "image_list" : [],
            "video_list": [],
            }
    if request.POST.get('vendor_id'):
        product_dict['vendor_id'] = ObjectId(request.POST.get('vendor_id'))
    if request.POST.get('brand_id'):
        product_dict['brand_id'] = ObjectId(request.POST.get('brand_id'))
    if request.POST.get('manufacture_id'):
        Manufacture_obj = DatabaseModel.get_document(Manufacture.objects,{'name':request.POST.get('manufacture_id')})
        if Manufacture_obj== None:
            Manufacture_obj = DatabaseModel.save_documents(Manufacture,{'name':request.POST.get('manufacture_id')})
        product_dict['manufacture_id'] = Manufacture_obj.id
    product_obj = DatabaseModel.save_documents(product,product_dict)
    DatabaseModel.save_documents(product_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Created','product_id':product_obj.id})
    data_list = json.loads(request.POST.get('attribute_list', '[]')) 
    for i  in data_list:
        attribute_name = str(i['name'])
        attribute_value = i['values']
        if not attribute_name or not attribute_value:
            break
        Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'name':attribute_name,'module_name':'product','module_id__in':[str(product_obj.id)]})
        if Attribute_obj == None:
            Attribute_obj = DatabaseModel.save_documents(Attribute,{'name':attribute_name,'module_name':'product','module_id':[str(product_obj.id)],'values':attribute_value})
            product_obj.attribute_list.append(Attribute_obj.id)
        else:
            DatabaseModel.update_documents(Attribute.objects,{'name':attribute_name},{'add_to_set__values':attribute_value})
    # for i  in request.POST.get('image_list'):
    #     if i['type'] == 'base-64':
    #         format, imgstr = i['base64_image'].split(';base64,')
    #         ext = format.split('/')[-1] 
    #         image_data = ContentFile(base64.b64decode(imgstr), name=f"uploaded_image.{ext}")
    #         upload_result = cloudinary.uploader.upload(image_data, folder="KM-DIGI/image")
    #         image_url = upload_result.get("secure_url")
    #         image_name = i['name']
    #         image_url = str(image_url)
    #         product_obj.image_list.append(Image(name=image_name, url=image_url))
    #     elif i['type'] == 'base-64':
    #         pass
    # video_list = json.loads(request.POST.get('video_list', '[]')) 

    # for i  in video_list:
    #     video_name = i['name']
    #     video_url = i['url']
    #     product_obj.video_list.append(Video(name=video_name, url=video_url)) 
    # for i  in request.POST.get('attachment_list'):
    #     attachment_name = i['name']
    #     attachment_url = i['url']
    #     product_obj.attachment_list.append(Attachment(name=attachment_name, url=str(attachment_url)))
    # related_products = json.loads(request.POST.get('related_products', '[]')) 
    # for i  in related_products:
    #     repro_name = i['name']
    #     repro_url = i['url']
    #     product_obj.related_products.append(RelatedProduct(name=repro_name, url=repro_url))
    product_obj.save()
    product_id = product_obj.id
    data = dict()
    data['is_created'] = True
    data['product_id'] = str(product_obj.id)
    image_files = request.FILES.getlist('images')
    video_files = request.FILES.getlist('videos')
    document_files = request.FILES.getlist('documents') 
    client_id = get_current_client()
    product_obj = DatabaseModel.get_document(product.objects,{'id':product_id})
    image_count = 0
    video_count = 0
    doc_count = 0
    for  image_file in image_files:
        if product_obj == None:
            product_obj = DatabaseModel.get_document(product.objects,{'sku':str(image_file.file.name)})
        image_count +=1
        image_name = f'{product_obj.sku}-Image-{image_count}'
        image_file.file.name = image_name  
        upload_result = cloudinary.uploader.upload(image_file,folder="KM-DIGI/image",public_id=image_name )
        image_url = upload_result.get("secure_url")
        public_id = upload_result.get("public_id")
        product_obj.image_list.append(Image(name=image_name, url=str(image_url)))
        product_image = ProductImage(name=image_name, image_url=image_url,client_id = ObjectId(client_id),public_id=public_id)
        product_image.save()
    for  video_file in video_files:
        if product_obj == None:
            product_obj = DatabaseModel.get_document(product.objects,{'sku':str(image_file.file.name)})
        video_count +=1
        video_name = f'{product_obj.sku}-Video-{video_count}'
        video_file.file.name = video_name
        upload_result = cloudinary.uploader.upload(video_file, resource_type="video",folder="KM-DIGI/video",public_id=video_name )
        video_url = upload_result.get("secure_url")
        public_id = upload_result.get("public_id")
        product_obj.video_list.append(Video(name=video_name, url=str(video_url)))
        product_video = ProductVideo(name=video_name, video_url=video_url,client_id = ObjectId(client_id),public_id=public_id)
        product_video.save()
    for document_file in document_files:
        if product_obj == None:
            product_obj = DatabaseModel.get_document(product.objects,{'sku':str(image_file.file.name)})
        if product_obj:
            data['error_list'].append("sku name is mandatory for your image name")
        doc_count +=1
        document_name = f'{product_obj.sku}-doc-{doc_count}'
        document_file.file.name = document_name
        upload_result = cloudinary.uploader.upload(document_file, resource_type="raw",folder="KM-DIGI/doc",public_id=document_name )
        document_url = upload_result.get("secure_url")
        public_id = upload_result.get("public_id")
        product_obj.attachment_list.append(Attachment(name=document_name, url=str(document_url)))
        product_document = ProductDocument(name=document_name, document_url=document_url,client_id = ObjectId(client_id),public_id=public_id)
        product_document.save()
    product_obj.save()
    return data

def obtainBusinessType(request):
    pipeline = [
    {
            '$group': {
                "_id":'$_id',
                'name':{'$first':'$name'},
        }
        },
    {
        '$sort': { '_id': -1 }
    },
    ]
    brand_list = list(Business_type.objects.aggregate(*pipeline))
    for i in brand_list:
        i['id'] = str(i['_id'])
        del i['_id']
    return brand_list


def obtainCategoryList(request):
    search = request.GET.get('search')
    client_id = get_current_client()

    pipeline = [
        {
            '$match': {
                'name': { '$regex': search, '$options': 'i' } ,
                'client_id':ObjectId(client_id)
            }
        },{
        '$lookup': {
            'from': 'category',
            'localField': 'levels',
            'foreignField': '_id',
            'as': 'category'
        }
    }, 
    {
            '$unwind': {
                'path': '$category',
                'preserveNullAndEmptyArrays': True
            }
    },
        {
            '$group': {
                "_id": '$_id',
                'name': { '$first': '$name' },
                "levels":{ '$push': '$category.name'},
            }
        },
        {
            '$sort': { '_id': -1 }
        },
    ]

    category_list = list(category_config.objects.aggregate(*pipeline))
    for i in category_list:
        i['id'] = str(i['_id'])
        del i['_id']
        if 'levels' in i and i['levels']:
            levels = i['levels'][0].split(",") if isinstance(i['levels'], list) and len(i['levels']) == 1 else i['levels']
            i['level_str'] = " / ".join(levels)
        else:
            i['level_str'] = ""
        del i['levels']
    return category_list

@csrf_exempt
def findDuplicateCategory(request):
    search = request.GET.get('search')
    client_id = get_current_client()

    category_config_id = request.GET.get('category_config_id')
    match_1 = {}
    match = {}
    if category_config_id:
        match_1 = {
    '$expr': {'$eq': [{'$toLower': '$category.name'}, search.lower()]},'client_id':ObjectId(client_id)
}
        match = {
                '_id':ObjectId(category_config_id),
                'client_id':ObjectId(client_id)
            }
    else:
         match =  {
    '$expr': {'$eq': [{'$toLower': '$name'}, search.lower()],
              },'client_id':ObjectId(client_id)
}
    pipeline = [
        {
            '$match': match
        },{
        '$lookup': {
            'from': 'category',
            'localField': 'levels',
            'foreignField': '_id',
            'as': 'category'
        }
    }, 
    {
            '$unwind': {
                'path': '$category',
                # 'preserveNullAndEmptyArrays': True
            }
    },{
            '$match': match_1
        },
        {
            '$group': {
                "_id": '$_id',
                'name': { '$first': '$name' },
                'category_name': { '$first': '$category.name' },
                'levels': { '$first': '$levels' },
            }
        }
    ]

    category_obj = list(category_config.objects.aggregate(*pipeline))
    data = dict()
    data['error'] = False
    if category_obj:
        for i in category_obj:
            if len(i['levels']) > 0:
                if i['category_name'] == search :
                    data['error'] = True
                    return data
        return data
    return data

@csrf_exempt
def obtainAllProductList(request):
    # user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    client_id = get_current_client()
    json_req = JSONParser().parse(request)
    filter = json_req.get("filter")
    page = json_req.get("page")
    from_page = (page - 1) * 20
    to_page = (page * 20)
    search_term = json_req.get('search')
    if search_term == None:
        search_term = ""
    if filter == True or filter == None or filter == "":
        reverse_check = -1
    else:
        reverse_check = 1
    category_id = json_req.get("category_id",[])
    brand_id = json_req.get("brand_id",[])
    vendor_id = json_req.get("vendor_id",[])
    brand_obj = {}
    if len(brand_id) != 0:
        brand_id = [ObjectId(i) for i in brand_id]
        brand_obj = {"brand._id":{'$in':brand_id}}
    vendor_obj = {}
    if len(vendor_id) != 0:
        vendor_id = [ObjectId(i) for i in vendor_id]
        vendor_obj = {"vendor._id":{'$in':vendor_id}}
    category_obj = {}
    if len(category_id) != 0 :
        category_id = [ObjectId(i) for i in category_id]
        category_id_jj = list()
        category_obj_list = DatabaseModel.list_documents(category_config.objects,{'id__in':category_id,'client_id':client_id})
        for ii in category_obj_list:
            category_config_obj_jr = DatabaseModel.get_document(category.objects,{'name':ii.name,'client_id':client_id})
            category_config_obj_j = DatabaseModel.list_documents(category_config.objects,{'levels__in':[category_config_obj_jr.id],'client_id':client_id})
            for zzz in category_config_obj_j:
                category_id_jj.append(zzz.id)
        category_obj = {"category_config._id":{'$in':category_id_jj}}
    pipeline = [
    {
        "$match":{'client_id':ObjectId(client_id)}
    },
    {
        '$lookup': {
            'from': 'brand',
            'localField': 'brand_id',
            'foreignField': '_id',
            'as': 'brand'
        }
    }, 
    {
            '$unwind': {
                'path': '$brand',
                'preserveNullAndEmptyArrays': True
            }
        },{
            "$match":brand_obj
        },{
        '$lookup': {
            'from': 'vendor',
            'localField': 'vendor_id',
            'foreignField': '_id',
            'as': 'vendor'
        }
    }, 
    {
            '$unwind': {
                'path': '$vendor',
                'preserveNullAndEmptyArrays': True
            }
        },{
            "$match":vendor_obj
        },{
        '$lookup': {
            'from': 'Manufacture',
            'localField': 'manufacture_id',
            'foreignField': '_id',
            'as': 'Manufacture'
        }
    }, 
    {
            '$unwind': {
                'path': '$Manufacture',
                'preserveNullAndEmptyArrays': True
            }
        },
        {
        '$lookup': {
            'from': 'category_config',
            'localField': 'category_id',
            'foreignField': '_id',
            'as': 'category_config'
        }
    }, 
    {
            '$unwind': {
                'path': '$category_config',
                'preserveNullAndEmptyArrays': True
            }
        },{
            "$match":category_obj
        },
    #     {
    #     '$lookup': {
    #         'from': 'category',
    #         'localField': 'category_config.levels',
    #         'foreignField': '_id',
    #         'as': 'category'
    #     }
    # }, 
    # {
    #         '$unwind': {
    #             'path': '$category',
    #             'preserveNullAndEmptyArrays': True
    #         }
    #     },
    {
        '$group': {
            "_id":'$_id',
            "product_id" : {'$first': "$product_id"},
            "mpn" :{'$first': "$mpn"},
            "sku" : {'$first': "$sku"},
            "upc" : {'$first': "$upc"},
            "ean" : {'$first': "$ean"},
            "gtin" : {'$first': "$gtin"},
            "unspc":{'$first': "$unspc"},
            "model": {'$first': "$model"},
            "vendor_name": {'$first': "$vendor.name"},
            "brand_name": {'$first': "$brand.name"},
            "manufacture_name": {'$first': "$Manufacture.name"},
            "category_name": {'$push': "$category_config.name"},
            "name":{'$first': "$name"},
            "msrp" : {'$first': "$msrp"},
            "image_list" : {'$first': "$image_list"},
            "feature_list" : {'$first': "$feature_list"},
            "attribute_list" : {'$first': "$attribute_list"},
            "short_description" : {'$first': "$short_description"},
            "personalized_short_description" : {'$first': "$personalized_short_description"},
            "long_description" : {'$first': "$long_description"},
            "personalized_long_description" : {'$first': "$personalized_long_description"},
            
                }
            },{
    '$match': {
    '$or': [
        
        { 'product_id': { '$regex': search_term, '$options': 'i' } },  
        { 'sku': { '$regex': search_term, '$options': 'i' } },  
        { 'ean': { '$regex': search_term, '$options': 'i' } },  
        { 'mpn': { '$regex': search_term, '$options': 'i' } },  
        { 'name': { '$regex': search_term, '$options': 'i' } },
        { 'category_name': { '$regex': search_term, '$options': 'i' } },
        { 'brand_name': { '$regex': search_term, '$options': 'i' } }, 
        { 'vendor_name': { '$regex': search_term, '$options': 'i' } }, 
        { 'model': { '$regex': search_term, '$options': 'i' } },
        { 'feature_list': { '$regex': search_term, '$options': 'i' } },
        { 'msrp': { '$regex': search_term, '$options': 'i' } },
]
    }
  },
  {'$sort': {'_id':reverse_check}},
    ]
    result = list(product.objects.aggregate(*pipeline))
    data = dict()
    for j in result:
        j['id'] = str(j['_id'])
        del (j['_id'])
        j['product_id'] = str(j['product_id']) if 'product_id'in j else ""
        filled_fields = sum([ 
            bool(j['short_description'] or j['long_description'] or j['personalized_short_description'] or j['personalized_long_description']), 
            bool (len(j['feature_list'])>0), 
            bool(len(j['attribute_list']) >0), 
            bool(len(j['image_list']) > 0),
        ])
        del j['attribute_list']
        # del j['image_list']
        del j['feature_list']
        j['completeness_percentage'] = (filled_fields / 4) * 100 
    data['product_count'] = len(result)
    # data['product_list'] = result
    data['product_list'] = []
    if result:
        data['product_list'] = result[from_page:to_page]
    return data

@csrf_exempt
def obtainProductDetails(request):
    product_id = request.GET.get('id')
    client_id = get_current_client()

    pipeline = [
    {
        "$match":{'_id':ObjectId(product_id)}
    },
    {
        '$lookup': {
            'from': 'brand',
            'localField': 'brand_id',
            'foreignField': '_id',
            'as': 'brand'
        }
    }, 
    {
            '$unwind': {
                'path': '$brand',
                'preserveNullAndEmptyArrays': True
            }
        },{
        '$lookup': {
            'from': 'vendor',
            'localField': 'vendor_id',
            'foreignField': '_id',
            'as': 'vendor'
        }
    }, 
    {
            '$unwind': {
                'path': '$vendor',
                'preserveNullAndEmptyArrays': True
            }
        },{
        '$lookup': {
            'from': 'manufacture',
            'localField': 'manufacture_id',
            'foreignField': '_id',
            'as': 'manufacture'
        }
    }, 
    {
            '$unwind': {
                'path': '$manufacture',
                'preserveNullAndEmptyArrays': True
            }
        },{
        '$lookup': {
            'from': 'category_config',
            'localField': 'category_id',
            'foreignField': '_id',
            'as': 'category_config'
        }
    }, 
    {
            '$unwind': {
                'path': '$category_config',
                'preserveNullAndEmptyArrays': True
            }
        },{
        '$lookup': {
            'from': 'category',
            'localField': 'category_config.levels',
            'foreignField': '_id',
            'as': 'category'
        }
    }, 
    {
            '$unwind': {
                'path': '$category',
                'preserveNullAndEmptyArrays': True
            }
        },
    {
        '$group': {
            "_id":'$_id',
            "product_id" : {'$first': "$product_id"},
            "mpn" :{'$first': "$mpn"},
            "sku" : {'$first': "$sku"},
            "upc" : {'$first': "$upc"},
            "ean" : {'$first': "$ean"},
            "gtin" : {'$first': "$gtin"},
            "unspc":{'$first': "$unspc"},
            "model": {'$first': "$model"},
            "vendor_id": {'$first': "$vendor_id"},
            "vendor_name": {'$first': "$vendor.name"},
            "brand_id": {'$first': "$brand_id"},
            "brand_name": {'$first': "$brand.name"},
            "manufacturer_id":{'$first': "$manufacture.name"},
            "category_id":{'$push':"$category_config._id"},
            "breadcrumb":{'$first': "$breadcrumb"},
            "name":{'$first': "$name"},
            "short_description": {'$first': "$short_description"},
            "personalized_short_description" : {'$first': "$personalized_short_description"},
            "long_description" : {'$first': "$long_description"},
            "personalized_long_description" : {'$first': "$personalized_long_description"},
            "feature_list" : {'$first': "$feature_list"},
            "attribute_list" : {'$first': "$attribute_list"},
            "related_products" :{'$first': "$related_products"},
            "application" :{'$first': "$application"},
            "certifications" : {'$first': "$certifications"},
            "Compliance" : {'$first': "$Compliance"},
            "Prop65" : {'$first': "$Prop65"},
            "esg" :{'$first': "$esg"},
            "Hazardous" : {'$first': "$Hazardous"},
            "service_warranty" :{'$first': "$service_warranty"},
            "product_warranty" : {'$first': "$product_warranty"},
            "country_of_origin" : {'$first': "$country_of_origin"},
            "currency" : {'$first': "$currency"},
            "msrp" : {'$first': "$msrp"},
            "selling_price" :{'$first': "$selling_price"},
            "discount_price" :{'$first': "$discount_price"},
            "attachment_list" : {'$first': "$attachment_list"},
            "image_list" : {'$first': "$image_list"},
            "video_list": {'$first': "$video_list"},
            "category_group_list": {'$first': "$category_group_list"},
            "attribute_list": {'$first': "$attribute_list"},
            
                }
            }
    ]
    result = list(product.objects.aggregate(*pipeline))
    if len(result)>0:
        result = result[0]
        result['id'] = str(result['_id'])
        result['vendor_name'] = str(result['vendor_name']) if result['vendor_name'] != None  else ""
        result['vendor_id'] = str(result['vendor_id']) if result['vendor_id'] != None  else ""
        result['brand_name'] = str(result['brand_name']) if result['brand_name'] != None  else ""
        result['brand_id'] = str(result['brand_id']) if result['brand_id'] != None  else ""
        # result['manufacture_id'] = str(result['manufacture_id']) if result['manufacture_id'] != None  else ""
        # result['manufacture_name'] = str(result['manufacture_name']) if result['manufacture_name'] != None  else ""
        result['category_id'] = [str(i) for i in result['category_id']]
        result['product_id'] = str(result['product_id']) if 'product_id'in result else ""
        feature_str = "\n".join(f"{item}" for item in result['feature_list'])
        result['feature_list'] = feature_str
        attribute_list = []
        if result['attribute_list'] == None:
            result['attribute_list'] = []
        for j in  result['attribute_list']:
            Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'id':j})
            attribute_list.append({'id':str(Attribute_obj.id),'name':Attribute_obj.name,'values':Attribute_obj.values})
        result['brand_attribute_list'] =[]
        result['category_attribute_list'] =[]
        result['global_attribute_list'] =[]
        parent_level_list = []
        for x in result['category_id']:
            category_config_obj_l = DatabaseModel.get_document(category_config.objects,{'id':x})
            for zi in category_config_obj_l.levels:
                category_config_obj_name = DatabaseModel.get_document(category_config.objects,{'name':zi.name,'client_id':client_id})
                if category_config_obj_l.id != category_config_obj_name.id:
                    parent_level_list.append(str(category_config_obj_name.id))
        Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_id__in':[result['brand_id']]})
        for j in  Attribute_list_obj:
            result['brand_attribute_list'].append({'id':str(j.id),'name':j.name,'values':j.values})
        Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_id__in':parent_level_list})
        for j in  Attribute_list_obj:
            result['category_attribute_list'].append({'id':str(j.id),'name':j.name,'values':j.values})
        Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_name':"global"})
        for j in  Attribute_list_obj:
            result['global_attribute_list'].append({'id':str(j.id),'name':j.name,'values':j.values})
        result['attribute_list'] = attribute_list
        # if result['category_group_list']:
        #     for i in result['category_group_list']:
        #         i['b2c_company_id'] = str(i['b2c_company_id'])
        #         category_levels_list = list()
        #         for j in i['category_levels']:
        #             category_group_obj = DatabaseModel.get_document(category_group.objects,{'id':j})
        #             if category_group_obj:
        #                 category_levels_list.append({'id':str(j),'name':category_group_obj.name})
        #         i['category_levels'] = category_levels_list
        # else:
        #     result['category_group_list'] = []
        result['category_group_list'] = []
        channelCategory_obj = DatabaseModel.list_documents(channelCategory.objects,{'client_id':client_id,'category_config_id__in':result['category_id']})
        for ihj in channelCategory_obj:
            if ihj.channel_name:
                channelCategoryConfig_obj = DatabaseModel.get_document(
                    category_config.objects, {'id': ihj.category_config_id.id, 'client_id': client_id}
                )
                taxonomyconfig_level_str = []
                if channelCategoryConfig_obj:
                    for xzz in channelCategoryConfig_obj.levels:
                        taxonomyconfig_level_str.append(xzz.name)
                taxonomyconfig_level_str_str = " > ".join(taxonomyconfig_level_str).rstrip(">")
                taxonomy_level_str = " > ".join(ihj.taxonomy_level) if isinstance(ihj.taxonomy_level, list) else str(ihj.taxonomy_level)
                
                result['category_group_list'].append({
                    'channel_name': ihj.channel_name,
                    'taxonomy_level': taxonomy_level_str.rstrip(">"),
                    'category_taxonomy_level': taxonomyconfig_level_str_str
                })
                if 'category_group_list' in result and isinstance(result['category_group_list'], list):
                    result['category_group_list'] = sorted(
                        filter(None, result['category_group_list']),  # Removes None values
                        key=lambda x: x.get('channel_name', '')  # Sorts by channel_name
                    )
                else:
                    result['category_group_list'] = []
        del result['_id']
    return  result

from django.http import JsonResponse
import cloudinary.uploader
from .models import ProductImage
@csrf_exempt
def upload_image(request):
    if request.method == "POST":
        image_file = request.FILES.get('image')
        video_file = request.FILES.get('video') 
        document_file = request.FILES.get('documents') 
        if image_file:
            upload_result = cloudinary.uploader.upload(image_file,folder="KM-DIGI/image")
            image_url = upload_result.get("secure_url")
            public_id = upload_result.get("public_id")
            product_image = ProductImage(name=image_file.name, image_url=image_url,public_id = public_id)
            product_image.save()

            return JsonResponse({
                "message": "Image uploaded successfully",
                "image_url": image_url
            })
        elif video_file:
            upload_result = cloudinary.uploader.upload(video_file, resource_type="video",folder="KM-DIGI/video")
            video_url = upload_result.get("secure_url")
            public_id = upload_result.get("public_id")

            product_video = ProductVideo(name=video_file.name, video_url=video_url,public_id = public_id)
            product_video.save()

            return JsonResponse({
                "message": "Video uploaded successfully",
                "video_url": video_url
            })
        elif document_file:
            upload_result = cloudinary.uploader.upload(document_file, resource_type="raw",folder="KM-DIGI/doc")
            document_url = upload_result.get("secure_url")
            public_id = upload_result.get("public_id")

            product_document = ProductDocument(name=document_file.name, document_url=document_url,public_id = public_id)
            product_document.save()

            return JsonResponse({
                "message": "Document uploaded successfully",
                "document_url": document_url
            })

        return JsonResponse({"error": "No image/video file found"}, status=400)

    return JsonResponse({"error": "Invalid request method"}, status=400)


def list_images(request):
    images = ProductImage.objects.all()
    data = [{"name": img.name, "url": img.image_url} for img in images]
    return JsonResponse({"images": data})


@csrf_exempt
def v1(request):
    pipeline = [
        {
            '$lookup': {
                'from': 'category',
                'localField': 'levels',
                'foreignField': '_id',
                'as': 'category'
            }
        }, 
        {
            '$unwind': {
                'path': '$category',
            }
        },
        {
            '$group': {
                "_id": '$_id',
                'levels': {'$push': {'id': '$category._id', "name": '$category.name'}},
            }
        }
    ]

    category_config_list = list(category_config.objects.aggregate(pipeline))
    hierarchy = {}
    category_order = [str(entry["_id"]) for entry in category_config_list]  # Preserve order from database

    for entry in category_config_list:
        levels = entry["levels"]
        current = hierarchy
        _ids_collected = []

        for level in levels:
            level_id = str(level["id"])
            level_name = level["name"]
            _ids_collected.append(str(entry["_id"]))

            if level_id not in current:
                current[level_id] = {"id": level_id, "name": level_name, "_ids": set(), "children": {}}

            current[level_id]["_ids"].update(_ids_collected)
            current = current[level_id]["children"]

    def build_tree(node):
        sorted_children = sorted(
            [build_tree(child) for child in node["children"].values()],
            key=lambda x: (
                0 if x["name"] == "TABLE" else (1 if x["name"] == "CHAIRS" else category_order.index(x["_ids"][0]) if x["_ids"] else float('inf'))
            )
        )
        parent = ObjectId()
        category_config_list = DatabaseModel.list_documents(category_config.objects,{'id__in':node["_ids"]})
        for i in category_config_list:
            if len(i.levels) == 1:
                parent = i.id
        return {
            "id": node["id"],
            "name": node["name"],
            "_ids": sorted(node["_ids"], key=lambda x: category_order.index(x) if x in category_order else float('inf')),
            "product_count":DatabaseModel.count_documents(product.objects,{'category_id__in':node["_ids"]}),
            
            "sub_cat_product_count":DatabaseModel.count_documents(product.objects,{'category_id':ObjectId(parent)}),
            "children": sorted_children
                                                                       
        }

    hierarchy_list = sorted(
        [build_tree(node) for node in hierarchy.values()],
        key=lambda x: category_order.index(x["_ids"][0]) if x["_ids"] else float('inf')
    )

    return JsonResponse(hierarchy_list, safe=False)



@csrf_exempt
def productUpdate(request):
    json_req = JSONParser().parse(request)
    product_id = json_req['update_obj']['id']
    del json_req['update_obj']['id']
    del json_req['update_obj']['brand_attribute_list']
    del json_req['update_obj']['category_attribute_list']
    del json_req['update_obj']['global_attribute_list']
    if json_req['update_obj']['brand_id']:
        json_req['update_obj']['brand_id'] = ObjectId(json_req['update_obj']['brand_id'])
    else:
        json_req['update_obj']['brand_id'] = None
    if json_req['update_obj']['manufacturer_id']:
        Manufacture_obj = DatabaseModel.get_document(Manufacture.objects,{'name':json_req['update_obj']['manufacturer_id']})
        if Manufacture_obj == None:
            Manufacture_obj = DatabaseModel.save_documents(Manufacture,{'name':json_req['update_obj']['manufacturer_id']})
        json_req['update_obj']['manufacture_id'] = Manufacture_obj.id
    else:
        json_req['update_obj']['manufacture_id'] = None
    if json_req['update_obj']['category_id']:
        json_req['update_obj']['category_id'] = [ObjectId(i) for i in json_req['update_obj']['category_id']]
        category_config_obj = DatabaseModel.get_document(category_config.objects,{'id':json_req['update_obj']['category_id'][0]})
        category_level = ""
        if category_config_obj ==None:
            category_level = ""
        else:
            for j in category_config_obj.levels:
                category_level += j.name + ">"
            category_level = category_level.rstrip(">")
        json_req['update_obj']['breadcrumb'] = category_level
    # if 'category_id'  not in json_req['update_obj']:
    #     pass
    # elif json_req['update_obj']['category_id'] :
    #     json_req['update_obj']['category_id'] = []
    # else:
    #     json_req['update_obj']['category_id'] = None
    feature_list = json_req['update_obj']['feature_list'] 
    lst = feature_list.split("\n") 
    json_req['update_obj']['feature_list'] = [item for item in lst if item]  
    if json_req['update_obj']['vendor_id']:
        json_req['update_obj']['vendor_id'] = ObjectId(json_req['update_obj']['vendor_id'])
    else:
        json_req['update_obj']['vendor_id'] = None
    del json_req['update_obj']['vendor_name']
    del json_req['update_obj']['brand_name']
    del json_req['update_obj']['category_group_list']
    del json_req['update_obj']['manufacturer_id']
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    client_id = get_current_client()

    json_req['update_obj']['attribute_list'] = [ObjectId(i['id'])   for i in json_req['update_obj']['attribute_list']]
    DatabaseModel.update_documents(product.objects,{'id':product_id},json_req['update_obj'])
    DatabaseModel.save_documents(product_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Updated','product_id':product_id})
    # ProductLog(
    #         product_id= ObjectId(product_id),
    #         action= "updated",
    #         user= ObjectId(get_current_user()) ,
    #         timestamp=datetime.utcnow()
    #     ).save()
    data = dict()
    data['is_updated'] = True
    return data

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
import cloudinary.uploader
from bson import ObjectId
import re
@csrf_exempt
def updateFilesIntoProduct(request):
    data = dict()
    image_files = request.FILES.getlist('images')
    data['data'] = {'is_upload':True}
    data['data']['error_list'] = []
    product_id = request.POST.get('id')
    video_files = request.FILES.getlist('videos')
    document_files = request.FILES.getlist('documents') 
    client_id = get_current_client()
    if len(document_files)==0 and len(video_files)==0 and len(image_files) ==0:
        data['data']['is_upload'] = False
        data['data']['error'] = "Please Upload Atleast One File"
        return JsonResponse(data,safe=False)
    if product_id:
        product_obj = DatabaseModel.get_document(product.objects,{'id':product_id})
    else:
        product_obj = None
    for  image_file in image_files:
        if product_obj == None:
            file_name, _ = os.path.splitext(image_file.name)
            base_name = re.sub(r"\s*\(\d+\)$", "", file_name)
            product_obj = DatabaseModel.get_document(product.objects, {'sku__iexact': base_name,'client_id':client_id})
            if product_obj == None:
                data['data']['error_list'].append(f"file name {file_name} not support to any sku")
                # data['data']['is_upload'] = False
                continue
        image_count = len(product_obj.image_list)
        
        image_count +=1
        image_name = f'{product_obj.sku}-Image-{image_count}'
        image_file.file.name = image_name  
        upload_result = cloudinary.uploader.upload(image_file,folder="KM-DIGI/image",public_id=image_name )
        image_url = upload_result.get("secure_url")
        public_id = upload_result.get("public_id")
        DatabaseModel.update_documents(product.objects,{'id':product_obj.id},{'add_to_set__image_list':Image(name=image_name, url=str(image_url))})
        product_image = ProductImage(name=image_name, image_url=image_url,client_id = ObjectId(client_id),public_id=public_id)
        product_image.save()
        
    for  video_file in video_files:
        file_name, _ = os.path.splitext(video_file.name) 
        if product_obj == None:
            base_name = re.sub(r"\s*\(\d+\)$", "", file_name)
            product_obj = DatabaseModel.get_document(product.objects, {'sku__iexact': base_name})
            if product_obj == None:
                data['data']['error_list'].append("file name not support any sku")
                continue
        video_count = len(product_obj.video_list)
        
        video_count +=1
        video_name = f'{product_obj.sku}-Video-{video_count}'
        video_file.file.name = video_name
        upload_result = cloudinary.uploader.upload(video_file, resource_type="video",folder="KM-DIGI/video",public_id=video_name )
        video_url = upload_result.get("secure_url")
        public_id = upload_result.get("public_id")
        DatabaseModel.update_documents(product.objects,{'id':product_obj.id},{'add_to_set__video_list':Video(name=video_name, url=str(video_url))})

        product_video = ProductVideo(name=video_name, video_url=video_url,client_id = ObjectId(client_id),public_id=public_id)
        product_video.save()

    for document_file in document_files:
        file_name, _ = os.path.splitext(document_file.name) 
        if product_obj == None:
            base_name = re.sub(r"\s*\(\d+\)$", "", file_name)
            product_obj = DatabaseModel.get_document(product.objects, {'sku__iexact': base_name})
            if product_obj == None:
                data['data']['error_list'].append("file name not support any sku")
                continue
        doc_count = len(product_obj.attachment_list)

        doc_count +=1
        document_name = f'{product_obj.sku}-doc-{doc_count}'
        document_file.file.name = document_name
        upload_result = cloudinary.uploader.upload(document_file, resource_type="raw",folder="KM-DIGI/doc",public_id=document_name )
        document_url = upload_result.get("secure_url")
        public_id = upload_result.get("public_id")
        DatabaseModel.update_documents(product.objects,{'id':product_obj.id},{'add_to_set__attachment_list':Attachment(name=document_name, url=str(document_url))})
        product_document = ProductDocument(name=document_name, document_url=document_url,client_id = ObjectId(client_id),public_id=public_id)
        product_document.save()
        product_obj.save()
    
    return JsonResponse(data,safe=False)


def removeFiles(request):
    name = request.GET.get('name')
    id = request.GET.get('id')
    url = request.GET.get('url')
    action = request.GET.get('action')
    if action == 'image':
        domain_obj = ProductImage
        url_field = 'image_url'
        resource_type = "image" 
    elif action == 'video':
        domain_obj = ProductVideo
        url_field = 'video_url'
        resource_type = "video" 
    elif action == 'doc':
        domain_obj = ProductDocument
        url_field = 'document_url'
        resource_type = 'raw' 
    data = dict()
    product_obj = DatabaseModel.get_document(domain_obj.objects, { 'name': name, f'{url_field}': url })
    if product_obj:
        public_id = getattr(product_obj, "public_id", None)
        if public_id:
            if resource_type:
                delete_result = cloudinary.uploader.destroy(public_id, resource_type=resource_type)
            else:
                delete_result = {'result': 'Document deletion not supported via Cloudinary'}
        product_obj.delete()
        data['is_delete'] = True
        data['cloudinary_result'] = delete_result
    else:
        data['is_delete'] = False
    return JsonResponse(data, safe=False)

def removemedia(request):
    name = request.GET.get('name')
    id = request.GET.get('id')
    action = request.GET.get('action')
    product_obj = DatabaseModel.get_document(product.objects,{'id':id})
    data = dict()
    if action == 'image':
        for i in product_obj.image_list:
            if i.name == name:
                product_obj.image_list.remove(i)

                ProductImage_obj = DatabaseModel.get_document(ProductImage.objects,{ 'name': i.name, 'image_url': i.url })
                if ProductImage_obj:
                    public_id = ProductImage_obj.public_id
                    if public_id:
                        delete_result = cloudinary.uploader.destroy(public_id, resource_type="image")
                    ProductImage_obj.delete()

    elif action == 'video':
        for i in product_obj.video_list:
            if i.name == name:
                product_obj.video_list.remove(i)
                ProductVideo_obj = DatabaseModel.get_document(ProductVideo.objects, { 'name': i.name, 'video_url': i.url })
                if ProductVideo_obj:
                    public_id = getattr(ProductVideo_obj, "public_id", None)
                    if public_id:
                        delete_result = cloudinary.uploader.destroy(public_id, resource_type="video")
                    ProductVideo_obj.delete()

    elif action == 'document':
        for i in product_obj.attachment_list:
            if i.name == name:
                product_obj.attachment_list.remove(i)
                ProductDocument_obj = DatabaseModel.get_document(ProductDocument.objects, { 'name': i.name, 'document_url': i.url })
                if ProductDocument_obj:
                    public_id = getattr(ProductDocument_obj, "public_id", None)
                    if public_id:
                        delete_result = cloudinary.uploader.destroy(public_id, resource_type="raw")
                    ProductDocument_obj.delete()
    product_obj.save()
    data['is_delete'] = True
    return data

def obtainb2cCompany(request):
    b2cCompany_list = DatabaseModel.list_documents(b2c_company.objects)
    data = dict()
    data['b2cCompany_list'] = list()
    for i in b2cCompany_list:
        data['b2cCompany_list'].append({'id':str(i.id),'name':str(i.name)})
    return JsonResponse(data,safe=False)


def obtaincategoryGroupList(request):
    category_group_list = DatabaseModel.list_documents(category_group.objects)
    data = dict()
    data['category_group_list'] = list()
    for i in category_group_list:
        data['category_group_list'].append({'id':str(i.id),'name':str(i.name)})
    return JsonResponse(data,safe=False)


@csrf_exempt
def updateCategoryGroup(request):
    json_req = JSONParser().parse(request)
    b2c_company_id = json_req['b2c_company_id']
    product_id =  json_req['product_id']
    category_group_obj = DatabaseModel.get_document(category_group.objects,{'name':json_req['category_name']})
    if category_group_obj==None:
        category_group_obj = DatabaseModel.save_documents(category_group,{'name':json_req['category_name']})
    product_obj = DatabaseModel.get_document(product.objects,{'id':product_id})
    flag = 0
    for i in product_obj.category_group_list:
        if str(i.b2c_company_id.id) == b2c_company_id:
            i.category_levels.append(category_group_obj.id)
            flag = 1
    if flag == 0:
        product_obj.category_group_list.append(category_group_config(b2c_company_id=ObjectId(b2c_company_id), category_levels=[category_group_obj.id]))
    product_obj.save()
    data = dict()
    data['is_created'] = True
    return data


def obtainAttribute(request):
    module = request.GET.get('module')
    # module_id = request.GET.get('id')
    client_id = get_current_client()

    attribute_list = DatabaseModel.list_documents(Attribute.objects,{'module_name':module,'client_id':client_id})
    data = dict()
    data['attribute_list'] = list()
    for i in attribute_list:
        if i.module_name == module or (module == None and i.module_name == ""):
            module_name = []
            if i.module_name == 'product':
                product_obj = DatabaseModel.list_documents(product.objects,{'id__in':i.module_id,'client_id':client_id})
                for z in  product_obj:
                    module_name.append({'name':z.name,'id':str(z.id)})
            elif i.module_name == 'brand':
                brand_obj = DatabaseModel.list_documents(brand.objects,{'id__in':i.module_id,'client_id':client_id})
                for z in brand_obj:
                    module_name.append({'name':z.name,'id':str(z.id)})
            elif i.module_name == 'category':
                category_obj = DatabaseModel.list_documents(category_config.objects,{'id__in':i.module_id,'client_id':client_id})
                for z in category_obj:
                    module_name.append({'name':z.name,'id':str(z.id)})
            dict_ = {'id':str(i.id),'name':str(i.name),'code':i.code,'type':i.type,'values':i.values,'module_name':i.module_name,'module_name':module_name,'is_visible':i.is_visible}
            data['attribute_list'].append(dict_)
    data['attribute_list'] = sorted(data['attribute_list'], key=lambda x: ObjectId(x['id']),reverse=True)
    return data

@csrf_exempt
def createAttribute(request):
    json_req = JSONParser().parse(request)
    data = dict()
    module_name = json_req['module_name']
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    client_id = get_current_client()

    Attribute_obj_list = []
    if json_req['name']:
        json_req['name'].title()
    if module_name == 'brand' or module_name == 'category':
        Attribute_obj_list = DatabaseModel.list_documents(Attribute.objects,{'name':json_req['name'],'module_name':json_req['module_name'],'module_id__in':json_req['module_id'],'client_id':client_id})
        module_id = json_req['module_id']
    elif module_name == 'product':
        Attribute_obj_list = DatabaseModel.list_documents(Attribute.objects,{'name':json_req['name'],'module_name':json_req['module_name'],'module_id__in':json_req['module_id'],'client_id':client_id})
        module_id = json_req['module_id']
    else:
        Attribute_obj_list = DatabaseModel.list_documents(Attribute.objects,{'name':json_req['name'],'module_name':json_req['module_name'],'client_id':client_id})
        module_id = []
    for Attribute_obj in   Attribute_obj_list:
        if json_req.get('new'):
            json_req['new'] = str(json_req['new']).title()
            if json_req['new'] in Attribute_obj.values:
                data['is_created'] = False
                data['error'] = "Value Already Created"
                return data
            DatabaseModel.update_documents(Attribute.objects,{'id':Attribute_obj.id},{'add_to_set__values':json_req['new'],'module_id':module_id})
            DatabaseModel.save_documents(attribute_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Updated','attribute_id':Attribute_obj.id,'module_name':module_name})
            data['is_created'] = True
            Attribute_obj.save()
            return data
        elif json_req.get('values'):
            json_req['values'] =[str(i).title() for i in json_req['values']]
            DatabaseModel.update_documents(Attribute.objects,{'id':Attribute_obj.id},{'add_to_set__module_id':module_id,'add_to_set__values':json_req['values']})
            DatabaseModel.save_documents(attribute_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Updated','attribute_id':Attribute_obj.id,'module_name':module_name})
        else:
            Attribute_test_list = DatabaseModel.list_documents(Attribute.objects,{'name':json_req['name'],'module_name':json_req['module_name'],'module_id__in':json_req['module_id'],'id__ne':Attribute_obj.id,'client_id':client_id})
            Attribute_value_list = []
            for Attribute_test_ins in Attribute_test_list:
                Attribute_value_list.extend(Attribute_test_ins.values)
                DatabaseModel.delete_documents(Attribute.objects,{"id":Attribute_test_ins.id})
            DatabaseModel.update_documents(Attribute.objects,{'id':Attribute_obj.id},{'add_to_set__module_id':module_id,'add_to_set__values':Attribute_value_list})
            DatabaseModel.save_documents(attribute_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Created','attribute_id':Attribute_obj.id,'module_name':module_name})
            if len(Attribute_test_list)>0:
                break
        if module_name == 'product':
            DatabaseModel.update_documents(product.objects,{'id__in':Attribute_obj.module_id},{'add_to_set__attribute_list':[Attribute_obj.id]})
        elif module_name == 'brand':
            DatabaseModel.update_documents(brand.objects,{'id__in':Attribute_obj.module_id},{'add_to_set__attribute_list':[Attribute_obj.id]})
        elif module_name == 'category':
            DatabaseModel.update_documents(category_config.objects,{'id__in':Attribute_obj.module_id},{'add_to_set__attribute_list':[Attribute_obj.id]})
    if Attribute_obj_list:
        data['is_created'] = True
        return data
    if 'type' in json_req:
        json_req['values'] = [str(i).title() for i in json_req['values']]
        Attribute_obj = DatabaseModel.save_documents(Attribute,{'name':json_req['name'],'type':json_req['type'],'values':json_req['values'],'module_name':json_req['module_name'],'module_id':module_id,'client_id':ObjectId(client_id)})
        DatabaseModel.save_documents(attribute_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Created','attribute_id':Attribute_obj.id,'module_name':module_name})
        if module_name == 'product':
            DatabaseModel.update_documents(product.objects,{'id__in':json_req['module_id']},{'add_to_set__attribute_list':[Attribute_obj.id]})
        elif module_name == 'brand':
            DatabaseModel.update_documents(brand.objects,{'id__in':json_req['module_id']},{'add_to_set__attribute_list':[Attribute_obj.id]})
        elif module_name == 'category':
            DatabaseModel.update_documents(category_config.objects,{'id__in':json_req['module_id']},{'add_to_set__attribute_list':[Attribute_obj.id]})
    else:
        pass
    data['is_created'] = True
    return data

@csrf_exempt
def createAttributeGroup(request):
    json_req = JSONParser().parse(request)
    attributes = [ObjectId(i) for i in json_req['attributes']]
    DatabaseModel.save_documents(Attribute_group,{'name':json_req['name'],'code':json_req['code'],'attributes':attributes})
    data = dict()
    data['is_created'] = True
    return data


def obtainAttributeGroup(request):
    pipeline = [
        {
            '$lookup': {
                'from': 'attribute',
                'localField': 'attributes',
                'foreignField': '_id',
                'as': 'attribute'
            }
        }, 
        {
            '$unwind': {
                'path': '$attribute',
            }
        },
        {
            '$group': {
                "_id": '$_id',
                "name":{'$first':  '$name'},
                "code":{'$first':  '$code'},
                'Attribute_group_id':{'$first':  '$Attribute_group_id'},
                'attributes': {'$push': {'id': '$attribute._id', "name": '$attribute.name'}},
            }
        }
    ]
    attribute_list = list(Attribute_group.objects.aggregate(pipeline))
    data = dict()
    data['attribute_list'] = list()
    for i in attribute_list:
        i['id'] = str(i['_id'])
        del i['_id']
        i['Attribute_group_id'] = str(i['Attribute_group_id'])
        for j in i['attributes']:
            j['id'] = str(j['id'])
    data['attribute_list'] = attribute_list
    return data

@csrf_exempt
def updateAttributeGroup(request):
    json_req = JSONParser().parse(request)
    attributes = [ObjectId(i) for i in json_req['attributes']]
    DatabaseModel.update_documents(Attribute_group,{'id':json_req['id']},{'name':json_req['name'],'attributes':[attributes]})
    data = dict()
    data['is_updated'] = True
    return data

# @csrf_exempt
# def updateProductAttribute(request):
#     json_req = JSONParser().parse(request)
#     attribute_id = json_req['attribute_id']
#     value = json_req['value']
#     product_id = json_req['product_id']
#     product_obj = DatabaseModel.get_document(product.objects,{'id':product_id})
#     product_obj.attribute_list.append(ProductAttribute(attribute_id=ObjectId(attribute_id), value=str(value)))
#     product_obj.save()
#     data = dict()
#     data['is_updated'] = True
#     return data

# @csrf_exempt
# def updateProductAttributeGroup(request):
#     json_req = JSONParser().parse(request)
#     attribute_group_id = json_req['attribute_id']
#     product_id = json_req['product_id']
#     product_obj = DatabaseModel.get_document(product.objects,{'id':product_id})
#     product_obj.attribute_groups.append(ObjectId(attribute_group_id))
#     product_obj.save()
#     data = dict()
#     data['is_updated'] = True
#     return data


@csrf_exempt
def removeCategoryGroup(request):
    json_req = JSONParser().parse(request)
    b2c_company_id = json_req['b2c_company_id']
    product_id =  json_req['product_id']
    product_obj = DatabaseModel.get_document(product.objects,{'id':product_id})
    for i in product_obj.category_group_list:
        if str(i.b2c_company_id.id) == b2c_company_id:
           product_obj.category_group_list.remove(i)
    product_obj.save()
    data = dict()
    data['is_deleted'] = True
    return data


@csrf_exempt
def categoryUpdate(request):
    json_req = JSONParser().parse(request)
    category_id = json_req['update_obj']['id']
    config_id = json_req['update_obj']['config_id']
    json_req['update_obj']['attribute_list'] = [ObjectId(i)   for i in json_req['update_obj']['attribute_list']]
    del json_req['update_obj']['id']
    del json_req['update_obj']['config_id']
    DatabaseModel.update_documents(category_config.objects,{'id':config_id},json_req['update_obj'])
    DatabaseModel.update_documents(category.objects,{'id':category_id},{'name':json_req['update_obj']['name']})
    data = dict()
    data['is_updated'] = True
    return data

import pandas as pd
from io import BytesIO, StringIO

def clean_value(value, default=""):
    return value if pd.notna(value) else default


@csrf_exempt
def importCategory(request):
    data = {'status': False, 'error_list': [], 'is_error': False,'error_count':0,'added_count':0}
    client_id = ObjectId(get_current_client())
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    import_log_obj = DatabaseModel.save_documents(import_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'category'})

    if 'file' not in request.FILES:
        return JsonResponse(data, safe=False)

    file = request.FILES['file']
    if file.name.endswith('.xlsx'):
        df = pd.read_excel(file)
    elif file.name.endswith('.csv') or file.name.endswith('.txt'):
        df = pd.read_csv(file)
    elif file.name.endswith('.ods'):
        df = pd.read_excel(file, engine='odf')
    else:
        return JsonResponse(data, safe=False)

    if df.empty:
        data['is_error'] = True
        data['error'] = "Excel file should not be empty"
        return JsonResponse(data, safe=False)
    count = 0
    data_list = list(df.iterrows()) 
    import_log_obj.total_count = len(data_list)
    import_log_obj.save()
    for index, row in df.iterrows():
        count +=1
        import_log_obj.inprogress_count = import_log_obj.total_count - import_log_obj.completed_count 
        import_log_obj.save()
        dict_error = dict()
        dict_error['error-row'] = index + 2
        dict_error['error_list'] = []
        dict_error['is_error'] = False
        flag = False
        row_dict = row.to_dict()
        category_id_list = []
        categories = []
        for i in range(1, 9): 
            category_value = row_dict.get(f"Category {i}", "").strip() if pd.notna(row_dict.get(f"Category {i}")) else ""
            categories.append(category_value)
        if not categories[0]:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 1 is mandatory")
            dict_error['is_error'] = True
            flag = True
        else:
            DatabaseModel.update_documents(import_log.objects,{'id':import_log_obj.id},{'add_to_set__created_id_list':str(categories[0])})
        if  any(categories[2:8])  and not categories[1]:  
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 2 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[3:8]) and not categories[2]: 
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 3 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[4:8]) and not categories[3]:  
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 4 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[5:8]) and not categories[4]:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 5 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[6:8]) and not categories[5]:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 6 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  categories[7] and not categories[6]:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 7 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        category_id_list = []
        if flag == False:
            for i in range(8): 
                category_name = clean_value(categories[i], default="")
                if category_name:
                    category_name = category_name.title()
                    category_obj = DatabaseModel.get_document(
                        category.objects, {'name__iexact': category_name, 'client_id': client_id}
                    )
                    if category_obj is None:
                        category_obj = DatabaseModel.save_documents(
                            category, {'name': category_name, 'client_id': client_id}
                        )
                    category_config_obj_past = DatabaseModel.get_document(
                        category_config.objects, {'name__iexact': category_name, 'levels': category_id_list, 'client_id': client_id}
                    )
                    category_id_list.append(category_obj.id)
                    category_config_obj = DatabaseModel.get_document(
                        category_config.objects, {'name__iexact': category_name, 'levels': category_id_list, 'client_id': client_id}
                    )
                    if category_config_obj is None:
                        if category_config_obj_past:
                            category_config_obj_past.end_level = False
                            category_config_obj_past.save()
                        DatabaseModel.save_documents(
                            category_config, {'name': category_name, 'levels': category_id_list, 'client_id': client_id,'end_level':True}
                        )
                        # DatabaseModel.save_documents(category_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Created','category_config_id':category_obj.id})
                        import_log_obj.created_count +=1
                        import_log_obj.created_id_list.append(str(category_obj.id))
                        import_log_obj.save()
        if flag:
            data['error_list'].append(dict_error)
            data['error_count'] +=1
        else:
            data['added_count'] +=1
        import_log_obj.completed_count +=1
        import_log_obj.inprogress_count -=1
        import_log_obj.save()
    if len(data['error_list']) >0:
        data['is_error'] = True
    if len(data['error_list'])>0:
        data['status'] = False  
        data['is_error'] = True
    else:
        data['status'] = True 
    data['total_category'] = count
    import_log_obj.error_count = len(data['error_list'])
    import_log_obj.un_error_count = count - len(data['error_list'])
    import_log_obj.data = data
    import_log_obj.status = "completed"
    import_log_obj.save()
    return JsonResponse(data, safe=False)
from urllib.parse import urlparse
def is_valid_email(email):
    pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(pattern, email) is not None
def is_valid_url(url):
    parsed = urlparse(url)
    return all([parsed.scheme, parsed.netloc])
# def is_valid_phone_number(phone):
#     pattern = r"^\+?[1-9]\d{9,14}$"  
#     return bool(re.match(pattern, phone))

# def clean_phone_number(phone):
#     if phone is None or phone == "":  
#         return ""  
    
#     phone = str(phone).strip()  

#     if "." in phone:  
#         phone = phone.split(".")[0]  

#     return phone if phone.replace("+", "").isdigit() else ""
@csrf_exempt
def importVendor(request):
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    data = {'status': False,'error_list':[],'is_error':False,'error_count':0,'added_count':0}
    client_id = ObjectId(get_current_client())
    import_log_obj = DatabaseModel.save_documents(import_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'vendor'})
    if 'file' not in request.FILES:
        return JsonResponse(data,safe=False)
    file = request.FILES['file']
    if file.name.endswith('.xlsx'):
        df = pd.read_excel(file)
    elif file.name.endswith('.csv') or file.name.endswith('.txt'):
        df = pd.read_csv(file)
    elif file.name.endswith('.ods') :
        df = pd.read_excel(file, engine='odf')
    else:
        return JsonResponse(data,safe=False)
    count = 0
    data_list = list(df.iterrows()) 
    if len(data_list) == 0:
        data['is_error'] = True
        data['error'] = "Excel file should not be empty"
    import_log_obj.total_count = len(data_list)
    import_log_obj.save()
    for integer, row in df.iterrows():
        import_log_obj.inprogress_count = import_log_obj.total_count - import_log_obj.completed_count
        import_log_obj.save()
        count +=1
        dict_error = dict()
        dict_error['is_error'] =False
        dict_error['error_list']  = list()
        row_dict = row.to_dict()
        name = row_dict.get('name') or row_dict.get('Name') or row_dict.get('NAME')
        name = clean_value(name, default="")
        contact_info_email = row_dict.get('contact info Email') or row_dict.get('Contact Info Email')
        contact_info_email = str(clean_value(contact_info_email, default=""))
        contact_info_phone = row_dict.get('contact info Phone') or row_dict.get('Contact Info Phone')
        if type(contact_info_phone) == float:
            contact_info_phone = ""
        # contact_info_phone = clean_phone_number(
        #     row_dict.get('contact info Phone') or row_dict.get('Contact Info Phone')
        # )
        if is_valid_phone_number(str(contact_info_phone)):
            contact_info_phone = contact_info_phone
        else: 
            if contact_info_phone != "":
                contact_info_phone = ""
                dict_error['error-row'] = integer + 2
                dict_error['error_list'].append("Invalid  Phone Number")
        website = (row_dict.get('Website',""))
        website = str(clean_value(website, default=""))
        # if contact_info_phone:
            # if is_valid_phone_number(contact_info_phone) == False:
            #     contact_info_phone = ""
            #     dict_error['error-row'] = integer + 2
            #     dict_error['error_list'].append("Invalid Phone Number format")
        
        
        if name == "":
            dict_error['error-row'] = integer + 2
            dict_error['error_list'].append("Name is Mandatory")
            if contact_info_email:
                if is_valid_email(contact_info_email) == False:
                    contact_info_email = ""
                    dict_error['error-row'] = integer + 2
                    dict_error['error_list'].append("Invalid Email format")
            dict_error['is_error'] = True
            if website:
                website = str(clean_value(website, default=""))
                if is_valid_url(website):
                    website = website
                else:
                    dict_error['error-row'] = integer + 2
                    dict_error['error_list'].append("Invalid Website format")
        else:
            if contact_info_email:
                if is_valid_email(contact_info_email) == False:
                    contact_info_email = ""
                    dict_error['error-row'] = integer + 2
                    dict_error['error_list'].append("Invalid Email format")
            if website:
                website = str(clean_value(website, default=""))
                if is_valid_url(website):
                    website = website
                else:
                    dict_error['error-row'] = integer + 2
                    dict_error['error_list'].append("Invalid Website format")
            business_type = row_dict.get('Business Type')
            business_type = clean_value(business_type, default="")
            business_type_id = ""
            business_type_obj = DatabaseModel.get_document(Business_type.objects,{'name':business_type})
            if business_type_obj:
                business_type_id = business_type_obj.id
            else:
                if business_type != "":
                    business_type_id =DatabaseModel.save_documents(Business_type,{'name':business_type}).id
            logo = row_dict.get('Logo')
            logo = str(clean_value(logo, default=""))
            if logo:
                if logo:
                    try:
                        upload_result = cloudinary.uploader.upload(logo,folder="KM-DIGI/image")
                        logo = upload_result.get("secure_url")
                    except:
                        dict_error['error-row'] = integer + 2
                        dict_error['error_list'].append("Public URL required")
            description = row_dict.get('Description')
            description = str(clean_value(description, default=""))
            address = row_dict.get('Address')
            address = str(clean_value(address, default=""))
            city = row_dict.get('City')
            city = str(clean_value(city, default=""))
            industry_info = row_dict.get('Industry Info')
            industry_info = str(clean_value(industry_info, default=""))
            tax_info = row_dict.get('Tax Info')
            tax_info = str(clean_value(tax_info, default=""))
            name = str(name).title()
            Vendor_obj = DatabaseModel.get_document(Vendor.objects,{'name__iexact':name,'client_id':client_id})
            if Vendor_obj == None:
                if business_type == "":
                    Vendor_obj = DatabaseModel.save_documents(Vendor,{'name':name,'logo':str(logo),'address':address,'city':city,'contact_info_email':str(contact_info_email),'contact_info_phone':str(contact_info_phone),'description':description,'website':website,'tax_info':tax_info,'industry_info':industry_info,'client_id':ObjectId(client_id)})
                else:
                    Vendor_obj = DatabaseModel.save_documents(Vendor,{'name':name,'logo':str(logo),'business_type':ObjectId(business_type_id),'address':address,'city':city,'contact_info_email':str(contact_info_email),'contact_info_phone':str(contact_info_phone),'description':description,'website':website,'tax_info':tax_info,'industry_info':industry_info,'client_id':ObjectId(client_id)})
                vendor_number = 1
                while f'Department {vendor_number} Name' in row_dict and f'Department {vendor_number} Email' in row_dict    and f'Department {vendor_number} Phone Number' in row_dict:
                    department_name = row_dict[f'Department {vendor_number} Name']
                    email = row_dict[f'Department {vendor_number} Email']
                    email = clean_value(email, default="")
                    if email:
                        if is_valid_email(email) == False:
                            email = ""
                            dict_error['error-row'] = integer + 2
                            dict_error['error_list'].append("email format should be  contact info Department Email header")
                    phone_number = row_dict[f'Department {vendor_number} Phone Number']
                    phone_number = clean_value(phone_number, default="")
                    if type(phone_number) == float:
                        phone_number = ""
                    department_name = clean_value(department_name, default="")
                    if is_valid_phone_number(str(phone_number)):
                        phone_number = phone_number
                    else: 
                        if phone_number != "":
                            phone_number = ""
                            dict_error['error-row'] = integer + 2
                            dict_error['error_list'].append("Invalid Department Phone Number")
                    if department_name:
                        Vendor_obj.departments.append(ContactInfo(department_name=department_name, email=str(email), phone_number=str(phone_number)))
                    vendor_number +=1
                Vendor_obj.save()
                import_log_obj.created_count +=1
                import_log_obj.created_id_list.append(str(Vendor_obj.id))
                import_log_obj.save()


                # DatabaseModel.save_documents(vendor_log,{'user_id':user_login_id,'client_id':client_id,'action':'Created (Import)','vendor_id':Vendor_obj.id})

            else:
                if business_type_id != "":
                    Vendor_obj = DatabaseModel.update_documents(Vendor.objects,{'id':Vendor_obj.id},{'logo':str(logo),'business_type':ObjectId(business_type_id),'address':address,'city':city,'contact_info_email':str(contact_info_email),'contact_info_phone':str(contact_info_phone),'description':description,'website':website,'tax_info':tax_info,'industry_info':industry_info})
                else:
                    Vendor_obj = DatabaseModel.update_documents(Vendor.objects,{'id':Vendor_obj.id},{'logo':str(logo),'address':address,'city':city,'contact_info_email':str(contact_info_email),'contact_info_phone':str(contact_info_phone),'description':description,'website':website,'tax_info':tax_info,'industry_info':industry_info})
                vendor_number = 1
                while f'Department {vendor_number} Name' in row_dict and f'Department {vendor_number} Email' in row_dict    and f'Department {vendor_number} Phone Number' in row_dict:
                    department_name = row_dict[f'Department {vendor_number} Name']
                    email = row_dict[f'Department {vendor_number} Email']
                    email = clean_value(email, default="")
                    if email:
                        if is_valid_email(email) == False:
                            email = ""
                            dict_error['error-row'] = integer + 2
                            dict_error['error_list'].append("email format should be  contact info Department Email header")
                    phone_number = row_dict[f'Department {vendor_number} Phone Number']
                    phone_number = clean_value(phone_number, default="")
                    if type(phone_number) == float:
                        phone_number = ""
                    department_name = clean_value(department_name, default="")
                    if department_name:
                        Vendor_obj.departments.append(ContactInfo(department_name=department_name, email=str(email), phone_number=str(phone_number)))
                    vendor_number +=1
                Vendor_obj.save()
                # DatabaseModel.save_documents(vendor_log,{'user_id':user_login_id,'client_id':client_id,'action':'Updated (Import)','vendor_id':Vendor_obj.id})
                import_log_obj.updated_count +=1
                import_log_obj.updated_id_list.append(str(Vendor_obj.id))
                import_log_obj.save()
            data['added_count'] +=1
        if len(dict_error['error_list'])>0:
            data['error_list'].append(dict_error)
            if dict_error['is_error']:
                data['is_error'] = True
                data['error_count'] +=1
        import_log_obj.completed_count +=1
        import_log_obj.inprogress_count -=1
        import_log_obj.save()
    if len(data['error_list'])>0:
        data['status'] = False 
        data['is_error'] = True 
    else:
        data['status'] = True  
    data['total_vendor'] = count
    import_log_obj.error_count = len(data['error_list'])
    import_log_obj.un_error_count = count - len(data['error_list'])
    import_log_obj.data = data
    import_log_obj.status = "completed"
    import_log_obj.save()
    return JsonResponse(data,safe=False)


@csrf_exempt
def importBrand(request):
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    data = {'status': False,'error_list':[],'is_error':False,'error_count':0,'added_count':0}
    client_id =ObjectId(get_current_client())
    import_log_obj = DatabaseModel.save_documents(import_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'brand','total_count':100000,'completed_count':0})
    if 'file' not in request.FILES:
        return JsonResponse(data,safe=False)
    file = request.FILES['file']
    if file.name.endswith('.xlsx'):
        df = pd.read_excel(file)
    elif file.name.endswith('.csv') or file.name.endswith('.txt'):
        df = pd.read_csv(file)
    elif file.name.endswith('.ods'):
        df = pd.read_excel(file, engine='odf')
    else:
        return JsonResponse(data,safe=False)  
    count=0
    data_list = list(df.iterrows()) 
    import_log_obj.total_count = len(data_list)
    import_log_obj.save()
    if len(data_list) == 0:
        data['is_error'] = True
        data['error'] = "Excel file should not be empty"
    for integer, row in df.iterrows():
        count +=1
        import_log_obj.inprogress_count = import_log_obj.total_count - import_log_obj.completed_count 
        import_log_obj.save()
        dict_error = dict()
        dict_error['is_error'] =False
        dict_error['error_list']  = list()
        row_dict = row.to_dict()
        name = row_dict.get('name') or row_dict.get('Name') or row_dict.get('NAME')
        name = clean_value(name, default="")
        website = row_dict.get('Website')
        website = clean_value(website, default="")
        
        if website:
            website = str(clean_value(website, default=""))
            if is_valid_url(website):
                website = website
            else:
                dict_error['error-row'] = integer + 2
                dict_error['error_list'].append("Invalid Website format")
        if name == "":
            dict_error['error-row'] = integer + 2
            dict_error['error_list'].append("Name is Mandatory")
            dict_error['is_error'] = True
        else:
            country_of_origin = row_dict.get('Country Of Origin')
            country_of_origin = clean_value(country_of_origin, default="")
            status = row_dict.get('Status')
            status = clean_value(status, default="")
            description = row_dict.get('Description')
            description = clean_value(description, default="")
            warranty_details = row_dict.get('Warranty Details')
            warranty_details_based = clean_value(row_dict.get('Warranty Details Option'), default="")
            warranty_details_based = warranty_details_based.lower()
            if warranty_details_based in ['years','months','weeks','days'] or warranty_details_based == "":
                warranty_details = clean_value(warranty_details, default="")
                warranty_details_based = clean_value(warranty_details_based, default="")
            else:
                dict_error['error-row'] = integer + 2
                dict_error['error_list'].append("Invalid warrenty Details Options")
                dict_error['is_error'] = True
            logo = row_dict.get('Logo')
            logo = str(clean_value(logo, default=""))
            if logo:
                if logo:
                    try:
                        upload_result = cloudinary.uploader.upload(logo,folder="KM-DIGI/image")
                        logo = upload_result.get("secure_url")
                    except:
                        dict_error['error-row'] = integer + 2
                        dict_error['error_list'].append("Public URL required")
            client_id = get_current_client()
            name = str(name).title()
            brand_obj = DatabaseModel.get_document(brand.objects,{'name__iexact':name,'client_id':client_id})
            if brand_obj == None:
                brand_obj = DatabaseModel.save_documents(brand,{'name':name,'logo':str(logo),'country_of_origin':country_of_origin,'warranty_details':warranty_details,'warranty_details_based':warranty_details_based,'status':status,'website':website,'description':description,'client_id':ObjectId(client_id)})
                # DatabaseModel.save_documents(brand_log,{'user_id':user_login_id,'client_id':client_id,'action':'Create (Import)','brand_id':brand_obj.id})
                import_log_obj.created_count +=1
                import_log_obj.created_id_list.append(str(brand_obj.id))
                import_log_obj.save()


            else:
                DatabaseModel.update_documents(brand.objects,{'name__iexact':name},{'logo':str(logo),'country_of_origin':country_of_origin,'warranty_details':warranty_details,'status':status,'warranty_details_based':warranty_details_based,'website':website,'description':description})
                # DatabaseModel.save_documents(brand_log,{'user_id':user_login_id,'client_id':client_id,'action':'Import-Updated','brand_id':brand_obj.id})
                import_log_obj.updated_count +=1
                import_log_obj.updated_id_list.append(str(brand_obj.id))
                import_log_obj.save()

            data['added_count'] +=1
        if len(dict_error['error_list'])>0:
            data['error_list'].append(dict_error)
            if dict_error['is_error']:
                data['is_error'] = True
                data['error_count'] +=1
        import_log_obj.completed_count +=1
        import_log_obj.inprogress_count -=1
        import_log_obj.save()
    if len(data['error_list'])>0:
        data['status'] = False  
    else:
        data['status'] = True   
    import_log_obj.error_count = len(data['error_list'])
    import_log_obj.un_error_count = count - len(data['error_list'])
    import_log_obj.data = data
    import_log_obj.status = "completed"
    import_log_obj.save()
    data['total_brand'] = count
    return JsonResponse(data,safe=False)

import difflib
@csrf_exempt
def importAttribute(request):
    client_id = get_current_client()
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')

    data = {'status': False,'error_list':[],'is_error':False,'error_count':0,'added_count':0}
    import_log_obj = DatabaseModel.save_documents(import_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'attribute'})

    if 'file' not in request.FILES:
        return JsonResponse(data,safe=False)
    file = request.FILES['file']
    if file.name.endswith('.xlsx'):
        df = pd.read_excel(file)
    elif file.name.endswith('.csv') or file.name.endswith('.txt'):
        df = pd.read_csv(file)
    elif file.name.endswith('.ods'):
        df = pd.read_excel(file, engine='odf')
    else:
        return JsonResponse(data,safe=False)  
   
    data_list = list(df.iterrows()) 
    if len(data_list) == 0:
        data['is_error'] = True
        data['error'] = "Excel file should not be empty"
    count = 0
    import_log_obj.total_count = len(data_list)
    import_log_obj.save()
    for integer, row in df.iterrows():
        count +=1
        import_log_obj.inprogress_count = import_log_obj.total_count - import_log_obj.completed_count 
        import_log_obj.save()
        dict_error = dict()
        dict_error['is_error'] =False
        dict_error['error_list']  = list()
        row_dict = row.to_dict()
        name = row_dict.get('Attribute Name') or row_dict.get('Attribute Name') or row_dict.get('Attribute Name')
        name = clean_value(name, default="")
        if str(name) == "":
            dict_error['error-row'] = integer + 2
            dict_error['error_list'].append("Name is Mandatory")
            dict_error['is_error'] = True
        else:
            name = name.title()
            module_name = row_dict.get('Module Based Name')
            module_name = clean_value(module_name, default="")
            type = row_dict.get('Type')
            type = clean_value(type, default="")
            module_name = difflib.get_close_matches(module_name.lower(), ['brand','category','product','global'], n=1)
            if len(module_name) == 0:
                module_name = ""
                dict_error['error-row'] = integer + 2
                dict_error['error_list'].append("Module based name is mandatory")
                dict_error['is_error'] = True
            else:
                module_name = module_name[0]
            module_id_list = list()
            if module_name == 'brand':
                module_id_name = str(row_dict.get('Brand'))
                module_id_name = clean_value(module_id_name, default="")

                if module_id_name == "":
                    dict_error['error-row'] = integer + 2
                    dict_error['error_list'].append("Brand value is mandatory")
                    dict_error['is_error'] = True
                brand_list = DatabaseModel.list_documents(brand.objects,{'name__in':[module_id_name]})
                for i in brand_list:
                    module_id_list.append(str(i.id))
            elif module_name == 'category':
                module_id_name = row_dict.get('Category Names')
                module_id_name = clean_value(module_id_name, default="")

                if module_id_name == "":
                    dict_error['error-row'] = integer + 2
                    dict_error['error_list'].append("Category Value is mandatory")
                    dict_error['is_error'] = True
                    
                category_config_list = DatabaseModel.list_documents(category_config.objects,{'name__in':[module_id_name]})
                for i in category_config_list:
                    module_id_list.append(str(i.id))
            
            Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'name__iexact':name,'type':type,'module_name':module_name,'client_id':client_id})
            if Attribute_obj == None:
                Attribute_obj = DatabaseModel.save_documents(Attribute,{'name':name,'type':type,'values':[],'module_name':module_name,'module_id':module_id_list,'client_id':ObjectId(client_id)})
                # DatabaseModel.save_documents(attribute_log,{'user_id':ObjectId(user_login_id),'client_id':ObjectId(client_id),'action':'Create (Import)','attribute_id':Attribute_obj.id})
                import_log_obj.created_count +=1
                import_log_obj.module_name = module_name

                import_log_obj.created_id_list.append(str(Attribute_obj.id))
                import_log_obj.save()
            module_id_list_id = "nsouaso"
            if len(module_id_list) > 0:
                module_id_list_id = module_id_list[0]
            elif module_id_list_id not in Attribute_obj.module_id:
                import_log_obj.updated_count +=1
                import_log_obj.module_name = module_name
                Attribute_obj.module_id.extend(module_id_list)
                Attribute_obj.save()
                import_log_obj.created_id_list.append(str(Attribute_obj.id))
                import_log_obj.save()
            else:
                import_log_obj.updated_count +=1
                import_log_obj.module_name = module_name
                import_log_obj.created_id_list.append(str(Attribute_obj.id))
                import_log_obj.save()
            num_attribute = 1
            check_value_1 = str(clean_value(f'Value {num_attribute}' in row_dict, default=""))
            if check_value_1:
                while f'Value {num_attribute}' in row_dict :
                    value = row_dict[f'Value {num_attribute}']
                    value = clean_value(value, default="")
                    if value:
                        value = str(value).title()
                        DatabaseModel.update_documents(Attribute.objects,{'id':Attribute_obj.id},{'add_to_set__values':str(value)})
                    num_attribute +=1
            else:
                dict_error['error-row'] = integer + 2
                dict_error['error_list'].append("Value 1 is mandatory")
                dict_error['is_error'] = True
        if data['is_error'] == False:
            data['added_count'] +=1
        if len(dict_error['error_list'])>0:
            data['error_list'].append(dict_error)
            if dict_error['is_error']:
                data['is_error'] = True
                data['error_count'] +=1
        import_log_obj.completed_count +=1
        import_log_obj.inprogress_count -=1
        import_log_obj.save()
    if len(data['error_list']) >0:
        data['status'] = False  
    else:
        data['status'] = True 
    import_log_obj.error_count = len(data['error_list'])
    import_log_obj.un_error_count = count - len(data['error_list'])
    import_log_obj.data = data
    import_log_obj.status = "completed"
    import_log_obj.save()
    data['total_attribute'] = count
    return JsonResponse(data,safe=False)


@csrf_exempt
def exportVendor(request):
    client_id = get_current_client()
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')

    export_log_obj = DatabaseModel.save_documents(export_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'vendor'})

    pipeline = [
        {
            '$match':{'client_id':ObjectId(client_id)}
        },
        {
            '$lookup': {
                "from": 'business_type',
                "localField": 'business_type',
                "foreignField": "_id",
                "as": "business_type"
            }
        },
         {
            '$unwind': {
                'path': '$business_type',
                'preserveNullAndEmptyArrays': True
            }
    },
        {
            "$group": {
                "_id": "$_id",
                "name":{ "$first":"$name"},
                "description":{ "$first":"$description"},
                "business_type_name":{ "$first":"$business_type.name"},
                "address":{ "$first":"$address"},
                "city":{ "$first":"$city"},
                "contact_info_email":{ "$first":"$contact_info_email"},
                "contact_info_phone":{ "$first":"$contact_info_phone"}, 
                "logo":{ "$first":"$logo"},
                "website":{ "$first":"$website"},
                "tax_info":{ "$first":"$tax_info"},
                "industry_info":{ "$first":"$industry_info"}, 
                "departments":{ "$first":"$departments"}, 
        }
    },{
        '$sort':{'_id':-1}
    }
    ]
    result = list(Vendor.objects.aggregate(*pipeline))
    max_departments = 0
    for i in result:
        if i['departments'] != None:
            if max_departments < len(i['departments']):
                max_departments = len(i['departments'])
    export_log_obj.total_count = len(result)
    export_log_obj.save()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Vendor"
    headers = ["S.No","Name","Contact Info Email","Contact Info Phone","Business Type","Website","Description","Address","City","Industry Info","Tax Info","Logo"]
    departments_headers = []
    for i in range(1, max_departments+1):
        departments_headers.append(f"department {i} name")
        departments_headers.append(f"department {i} email")
        departments_headers.append(f"department {i} phone number")
    headers.extend(departments_headers)
    worksheet.append(headers)

    for i, item in enumerate(result):
        business_type_name = item.get("business_type_name", "")
        row = [
            i + 1,
            item.get("name", ""),
            item.get("contact_info_email", ""),
            item.get("contact_info_phone", ""),
            business_type_name,
            item.get("website", ""),
            item.get("description", ""),
            item.get("address", ""),
            item.get("city", ""),
            item.get("industry_info", ""),
            item.get("tax_info", ""),
            item.get("logo", ""),
        ]
        departments = item.get("departments", [])
        for j in departments:
            row.append(j.get('department_name', ''))
            row.append(j.get('email', ''))
            row.append(j.get('phone_number',''))
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0) 
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="vendor.xlsx"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@csrf_exempt
def exportBrand(request):
    client_id = get_current_client()
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')

    export_log_obj = DatabaseModel.save_documents(export_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'brand'})

    pipeline = [
        {
            '$match':{'client_id':ObjectId(client_id)}
        },
        {
            "$group": {
                "_id": "$_id",
                "name":{ "$first":"$name"},
                "brand_id":{ "$first":"$brand_id"},
                "description":{ "$first":"$description"},
                "logo":{ "$first":"$logo"},
                "country_of_origin":{ "$first":"$country_of_origin"},
                "warranty_details":{ "$first":"$warranty_details"},
                "warranty_details_based":{ "$first":"$warranty_details_based"},
                
                "status":{ "$first":"$status"},
                "website":{ "$first":"$website"},
        }
    },{
        '$sort':{'_id':-1}
    }
    ]
    result = list(brand.objects.aggregate(*pipeline))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Brand"
    headers = ["S.No","Name","Status","Description","logo","Country Of Origin","Website","Warranty Details","Warranty Details Option",]
    worksheet.append(headers)
    for i, item in enumerate(result):
        business_type_name = item.get("business_type_name", "")
        if len (business_type_name) >0:
            business_type_name = business_type_name[0]
        else:
            business_type_name = ""
        if item.get("warranty_details_based", ""):
            warranty_details_based = item.get("warranty_details_based", "")
        else:
            warranty_details_based = ""
        row = [
            i + 1,
            item.get("name", ""),
            item.get("status", ""),
            item.get("description", ""),
            item.get("logo", ""),
            item.get("country_of_origin", ""),
            item.get("website", ""),
             item.get("warranty_details", "") ,
            warranty_details_based,
        ]
        worksheet.append(row)
    export_log_obj.total_count = len(result)
    export_log_obj.save()
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0) 
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="Brand.xlsx"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

@csrf_exempt
def exportCategory(request):
    client_id = get_current_client()
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    export_log_obj = DatabaseModel.save_documents(export_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'category'})
    pipeline = [
         {
            '$match':{'client_id':ObjectId(client_id)}
        },
         {
            '$lookup': {
                "from": 'category',
                "localField": 'levels',
                "foreignField": "_id",
                "as": "category"
            }
        },
         {
            '$unwind': {
                'path': '$category',
                'preserveNullAndEmptyArrays': True
            }
    },
        {
            "$group": {
                "_id": "$_id",
                "name":{ "$first":"$name"},
                "category_levels":{ "$push":"$category.name"},
        }
    },
    {
        '$sort':{'_id':-1}
    }
    ]
    result = list(category_config.objects.aggregate(*pipeline))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Category"
    headers = ["S.No","Name","Product Count","category_levels","Category 1","Category 2","Category 3","Category 4","Category 5","Category 6","Category 7","Category 8",]
    worksheet.append(headers)
    for i, item in enumerate(result):
        product_count = DatabaseModel.count_documents(product.objects,{'category_id':item.get("_id")})
        category_level = ""

        list_1 = list()
        inti = 1
        l_1 = ""
        for j in item.get("category_levels", []):
            category_level += j + ">"
            list_1.append(j)
        
        category_level = category_level.rstrip(">")
        row = [
            i + 1,
            item.get("name", ""),
            product_count,
            category_level,
        ]
        for iz in list_1:
            row.append(iz)
            inti +=1
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0) 
    export_log_obj.total_count = len(result)
    export_log_obj.save()
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="Category.xlsx"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response

@csrf_exempt
def exportProduct(request):
    client_id = get_current_client()
    json_req = JSONParser().parse(request)
    category_id = json_req.get("category_id",[])
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    # category_id = []
    brand_id = json_req.get("brand_id",[])
    # brand_id = None
    vendor_id = json_req.get("vendor_id",[])
    # vendor_id = None
    brand_obj = {}
    if len(brand_id) != 0:
        brand_id = [ObjectId(i) for i in brand_id]
        brand_obj = {"brand._id":{'$in':brand_id}}
    vendor_obj = {}
    if len(vendor_id) != 0:
        vendor_id = [ObjectId(i) for i in vendor_id]
        vendor_obj = {"vendor._id":{'$in':vendor_id}}
    category_obj = {}
    if len(category_id) != 0 :
        category_id = [ObjectId(i) for i in category_id]
        category_id_jj = list()
        category_obj_list = DatabaseModel.list_documents(category_config.objects,{'id__in':category_id,'client_id':client_id})
        for ii in category_obj_list:
            category_config_obj_jr = DatabaseModel.get_document(category.objects,{'name':ii.name,'client_id':client_id})
            category_config_obj_j = DatabaseModel.list_documents(category_config.objects,{'levels__in':[category_config_obj_jr.id],'client_id':client_id})
            for zzz in category_config_obj_j:
                category_id_jj.append(zzz.id)
        category_obj = {"category_id":{'$in':category_id_jj}}
    pipeline = [
         {
            '$match':{'client_id':ObjectId(client_id)}
        },{
'$match':category_obj
        },
         {
            '$lookup': {
                "from": 'category_config',
                "localField": 'category_id',
                "foreignField": "_id",
                "as": "category_config"
            }
        },
         {
            '$unwind': {
                'path': '$category_config',
                'preserveNullAndEmptyArrays': True
            }
    },
    # {
    #      '$match':category_obj
    # },
     {
            '$lookup': {
                "from": 'vendor',
                "localField": 'vendor_id',
                "foreignField": "_id",
                "as": "vendor"
            }
        },
         {
            '$unwind': {
                'path': '$vendor',
                'preserveNullAndEmptyArrays': True
            }
    },{
         '$match':vendor_obj
    }, {
            '$lookup': {
                "from": 'brand',
                "localField": 'brand_id',
                "foreignField": "_id",
                "as": "brand"
            }
        },
         {
            '$unwind': {
                'path': '$brand',
                'preserveNullAndEmptyArrays': True
            }
    },{
         '$match':brand_obj
    },{
            '$lookup': {
                "from": 'manufacture',
                "localField": 'manufacture_id',
                "foreignField": "_id",
                "as": "manufacture"
            }
        },
         {
            '$unwind': {
                'path': '$manufacture',
                'preserveNullAndEmptyArrays': True
            }
    },
    {
    "$group": {
        "_id": "$_id",
        "name": { "$first": "$name" },
        "product_id": { "$first": "$product_id" },
        "mpn": { "$first": "$mpn" },
        "sku": { "$first": "$sku" },
        "upc": { "$first": "$upc" },
        "ean": { "$first": "$ean" },
        "gtin": { "$first": "$gtin" },
        "unspc": { "$first": "$unspc" },
        "model": { "$first": "$model" },
        "vendor_name": { "$first": "$vendor.name" },
        "brand_name": { "$first": "$brand.name" },
        "brand_id": { "$first": "$brand._id" },
        "manufacture_name": { "$first": "$manufacture.name" },
        "category_name": { "$first": "$category_config.name" },
        "category_id": { "$first": "$category_id" },
        "breadcrumb": { "$first": "$breadcrumb" },
        "short_description": { "$first": "$short_description" },
        "personalized_short_description": { "$first": "$personalized_short_description" },
        "long_description": { "$first": "$long_description" },
        "personalized_long_description": { "$first": "$personalized_long_description" },
        "feature_list": { "$first": "$feature_list" },
        "attribute_list": { "$first": "$attribute_list" },
        "related_products": { "$first": "$related_products" },
        "category_group_list": { "$first": "$category_group_list" },
        "application": { "$first": "$application" },
        "certifications": { "$first": "$certifications" },
        "Compliance": { "$first": "$Compliance" },
        "Prop65": { "$first": "$Prop65" },
        "esg": { "$first": "$esg" },
        "Hazardous": { "$first": "$Hazardous" },
        "service_warranty": { "$first": "$service_warranty" },
        "product_warranty": { "$first": "$product_warranty" },
        "country_of_origin": { "$first": "$country_of_origin" },
        "currency": { "$first": "$currency" },
        "msrp": { "$first": "$msrp" },
        "selling_price": { "$first": "$selling_price" },
        "discount_price": { "$first": "$discount_price" },
        "attachment_list": { "$first": "$attachment_list" },
        "image_list": { "$first": "$image_list" },
        "video_list": { "$first": "$video_list" }
    }
    }
    ,
    {
        '$sort':{'_id':-1}
    }
    ]
    result = list(product.objects.aggregate(*pipeline))

    workbook = Workbook()
    worksheet = workbook.active
    action = json_req.get('action')
    if action == "amazon":
        DatabaseModel.save_documents(export_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'product-amazon','total_count':len(result)})
        worksheet.title = "Product"
        headers = ['feed_product_type']
        category_count = 0
        attribute_count = 0
        for i in result:
            i['attribute_list'] = i.get("attribute_list",[])
            parent_level_list = []
            for x in i['category_id']:
                category_config_obj_l = DatabaseModel.get_document(category_config.objects,{'id':x})
                if category_config_obj_l:
                    for zi in category_config_obj_l.levels:
                        category_config_obj_name = DatabaseModel.get_document(category_config.objects,{'name':zi.name,'client_id':client_id})
                        if category_config_obj_name:
                            if category_config_obj_l.id != category_config_obj_name.id:
                                parent_level_list.append(str(category_config_obj_name.id))
            Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_id__in':[str(i['brand_id'])]})
            for Attribute_obj in  Attribute_list_obj:
                i['attribute_list'].append(Attribute_obj.id)
            Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_name':'category','module_id__in':parent_level_list})
            for Attribute_obj in  Attribute_list_obj:
                i['attribute_list'].append(Attribute_obj.id)
            Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_name':"global"})
            for Attribute_obj in  Attribute_list_obj:
                i['attribute_list'].append(Attribute_obj.id)
            if attribute_count < len(i['attribute_list']):
                attribute_count = len(i['attribute_list'])
        for i in result:
            if category_count < len(i['category_id']):
                category_count = len(i['category_id'])
        for i in range(1, category_count + 1):
            headers.append(f'Additional Taxonomy{i}')
        headers.extend(['item_sku','update_delete','brand_name','external_product_id','external_product_id_type','model','product_description','part_number','manufacturer','item_name','gtin_exemption_reason','language_value1','language_value2','language_value3','model_name','edition','recommended_browse_nodes','swatch_image_url','relationship_type','variation_theme','parent_sku','parent_child',"bullet_point1","bullet_point2","bullet_point3","bullet_point4","bullet_point5","generic_keywords","recommended_uses_for_product","cost_price","cost_price_currency",
'main_image_url'])
        
        image_count = 0
        for i in result:
            if image_count < len(i['image_list']):
                image_count = len(i['image_list'])
        for i in range(1, image_count + 1):
            headers.append(f'other_image_url{i}')
        # Feature_count = 0
        # for i in result:
        #     if Feature_count < len(i['feature_list']):
        #         Feature_count = len(i['feature_list'])
        # for i in range(1, Feature_count + 1):
        headers.append(f'special_features')
        for i in range(1, attribute_count + 1):
            headers.append(f"Attribute{i}")
        worksheet.append(headers)
        for i, item in enumerate(result):
            channel_category_level = []
            channelCategory_obj = DatabaseModel.list_documents(channelCategory.objects,{'channel_name':"amazon",'client_id':client_id,'category_config_id__in':item['category_id']})
            for iii in  channelCategory_obj:
                category_level = " > ".join(iii.taxonomy_level) if isinstance(iii.taxonomy_level, list) else str(iii.taxonomy_level)
                channel_category_level.append(category_level)
            feature_list = item.get("feature_list",[])
            feature_list += [""] * (5 - len(feature_list))
            bullet_point1, bullet_point2, bullet_point3, bullet_point4, bullet_point5 = feature_list[:5]
            channel_category_level_1 = ""
            if len(channel_category_level)>0:
                channel_category_level_1 = channel_category_level[0]
            row = [channel_category_level_1]
            for i in range(1, category_count + 1):
                if len(channel_category_level) >i:
                    row.append(channel_category_level[i])
                else:
                    row.append("")
            row.extend([
        item.get("sku", ""),
        '',
        item.get("brand_name", ""), 
        item.get("product_id", ""), 
        item.get("upc", ""),
        item.get("model", ""), 
        item.get("short_description", "")+"\n"+item.get("personalized_short_description", "")+"\n"+(item.get("long_description", ""))+"\n"+(item.get("personalized_long_description", "")),
        item.get("mpn", ""),
        item.get("manufacture_name", ""),
        item.get("name", ""),
        '', #gtin_exemption_reason
        '', #language_value1
        '', #language_value2
        '', #language_value3
        '', #model_name
        '', #edition
        '', #recommended_browse_nodes
        '', #swatch_image_url
        '', #relationship_type
        '', #variation_theme
        '', #parent_sku
        '', #parent_child
            bullet_point1,
            bullet_point2,
            bullet_point3,
            bullet_point4,
            bullet_point5,
            "", #generic_keywords
            "", #recommended_uses_for_product
            item.get("msrp", ""), #cost_price
            item.get("currency", ""), #cost_price_currency
            ])             
            image_list = item.get("image_list",[])
            for j in image_list:
                row.append(j.get('url', ''))
            for i in range(image_count - len(image_list)):
                row.append("")
            feature_list = item.get("feature_list",[])
            # for j in feature_list:
            #     row.append(j)
            # for i in range(Feature_count - len(feature_list)):
            #     row.append("")
            feature_str = "\n".join(f"• {item}" for item in feature_list)
            row.append(feature_str)
            attribute_list_str = list()
            for j in item['attribute_list']:
                Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'id':j})
                if Attribute_obj:
                    value_str = ""
                    for j in Attribute_obj.values:
                        value_str += j + ","
                    value_str = value_str.rstrip(",")
                    attribute_list_str.append(f"{Attribute_obj.name}:{value_str}")
            row.extend(attribute_list_str)
            worksheet.append(row)            
    elif action == "shopify":
        DatabaseModel.save_documents(export_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'product-shopify','total_count':len(result)})
        worksheet.title = "Product"
        headers = ['Handle',
        'Title','Body (HTML)','Vendor','Product category']
        category_count = 0
        for i in result:
            if category_count < len(i['category_id']):
                category_count = len(i['category_id'])
        for i in range(1, category_count + 1):
            headers.append(f'Additional Taxonomy{i}')
        headers.extend(['Type','Tags','Published on online store','Status','SKU','Barcode','Option1 name','Option1 value','Option2 name',	'Option2 value'
        ])
        # attribute_count = 0
        # for i in result:
        #     if attribute_count < len(i['attribute_list']):
        #         attribute_count = len(i['attribute_list'])
        # for i in range(1, attribute_count + 1):
        #     headers.append(f'Option{i} name')
        #     headers.append(f'Option{i} value')
        headers.extend(['Price','Price / International','Compare-at price','Compare-at price / International','Cost per item','Charge tax','Tax code','Inventory policy','Inventory tracker','Inventory quantity','Continue selling when out of stock','Weight value (grams)','Weight unit for display','Requires shipping','Fulfillment service','Product image URL','Image position','Image alt text','Variant image URL','Gift card','SEO title','SEO description','Google Shopping / Google product category','Google Shopping / Gender','Google Shopping / Age group','Google Shopping / MPN','Google Shopping / AdWords Grouping','Google Shopping / AdWords labels','Google Shopping / Condition','Google Shopping / Custom product','Compare At Price / United States','Price / United States','Key Features','Specifications'])
        
        #     if feature_count < len(i['feature_list']):
        #         feature_count = len(i['feature_list'])
        # for i in range(1, feature_count + 1):
        #     headers.append(f"Feature {i}")
        worksheet.append(headers)
        for i, item in enumerate(result):
            category_level = ""
            end_level_id = None
            category_config_obj = DatabaseModel.get_document(category_config.objects,{'id__in':item['category_id']})
            if category_config_obj ==None:
                category_level = ""
            else:
                for j in category_config_obj.levels:
                    category_level += j.name + ">"
                list_level = category_config_obj.levels
                if len(list_level) > 0:
                    end_level_id = list_level[-1].name
                category_level = category_level.rstrip(">")
            formatted_description = (
    item.get("short_description", "") +"\n"+
    item.get("long_description", "") +"\n"+ item.get("personalized_short_description", "") +"\n"+ item.get("personalized_long_description", "")
)    
            channel_category_level = []
            channelCategory_obj = DatabaseModel.list_documents(channelCategory.objects,{'channel_name':"shopify",'client_id':client_id,'category_config_id__in':item['category_id']})
            for iii in  channelCategory_obj:
                category_level = " > ".join(iii.taxonomy_level) if isinstance(iii.taxonomy_level, list) else str(iii.taxonomy_level)
                channel_category_level.append(category_level)
            feature_list = item.get("feature_list",[])
            feature_list += [""] * (5 - len(feature_list))
            bullet_point1, bullet_point2, bullet_point3, bullet_point4, bullet_point5 = feature_list[:5]
            channel_category_level_1 = ""
            if len(channel_category_level)>0:
                channel_category_level_1 = channel_category_level[0]
            # row = [channel_category_level_1]
            
            row = [
                item.get("name", "").replace(" ", "-").lower(),
                item.get("name", ""),
                formatted_description,
                item.get("vendor_name", ""),channel_category_level_1,]
            for i in range(1, category_count + 1):
                if len(channel_category_level) >i:
                    row.append(channel_category_level[i])
                else:
                    row.append("")
            row.extend([
                end_level_id,#Type 
                item.get("name", "").replace(" ", ",").lower(),#Tags
                "",#Published on online store
                "active",#Status
                item.get("sku", ""),
                "",#Barcode
                "",
                "",
                "",
                "",
            ])
            # attribute_list = item.get("attribute_list",[])
            # for j in attribute_list:
            #     Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'id':j})
            #     if Attribute_obj:
            #         row.append(Attribute_obj.name)
            #         value_str = ""
            #         for j in Attribute_obj.values:
            #             value_str += j + ","
            #         value_str = value_str.rstrip(",")
            #         row.append(value_str)
            #     else:
            #         row.append("")
            #         row.append("")
            # for i in range(attribute_count - len(attribute_list)):
            #     row.append("")
            #     row.append("")
            extra_list = [
                item.get("msrp", ""),
                '',#Price / International
                '',#Compare-at price
                '',#Compare-at price / International
                '',#Cost per item
                '',#Charge tax
                '',#Tax code
                'continue',
                'shopify',#Inventory tracker
                '',#Inventory quantity
                '',#Continue selling when out of stock
                '',#Weight value (grams)
                '',#Weight unit for display','Requires shipping
                '',#Fulfillment service
                'manual',#Product image URL
                '',#Image position
                '',#Image alt text
                '',#Variant image URL
                '',#Gift card
                '',#SEO title
                '',#SEO description
                '',#Google Shopping / Google product category
                '',#Google Shopping / Gender
                '',#Google Shopping / Age group
                '',#Google Shopping / MPN
                '',#Google Shopping / AdWords Grouping
                '',#Google Shopping / AdWords labels
                '',#Google Shopping / Condition
                '',#Google Shopping / Custom product
                "",
                item.get("msrp", ""),
                item.get("selling_price", ""),
                ]
            feature_list = item.get("feature_list",[])
            row.extend(extra_list)
            feature_str = "\n".join(f"• {item}" for item in feature_list)
            row.append(feature_str)
            row.append(feature_str)
            worksheet.append(row)
    elif action == "bigcommerce":
        worksheet.title = "Product"
        DatabaseModel.save_documents(export_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'product-bigcommerce','total_count':len(result)})
        headers = [
    "Item",
    "ID",
    "Name",
    "Type",
    "SKU",
    "Options",
    "Inventory Tracking",
    "Current Stock",
    "Low Stock",
    "Price",
    "Cost Price",
    "Retail Price",
    "Sale Price",
    "Brand ID",
    "Channels",
    "Categories",
    "Description",
    "Custom Fields",
    "Page Title",
    "Product URL",
    "Meta Description",
    "Search Keywords",
    "Meta Keywords",
    "Bin Picking Number",
    "UPC",
    "EAN",
    "Global Trade Number",
    "Manufacturer Part Number",
    "Free Shipping",
    "Fixed Shipping Cost",
    "Shipping Groups",
    "Origin Locations",
    "Dimensional Rules",
    "Weight",
    "Width",
    "Height",
    "Depth",
    "Is Visible",
    "Is Featured",
    "Warranty",
    "Tax Class",
    "Product Condition",
    "Show Product Condition",
    "Sort Order",
    "Variant Image URL",
    "Internal Image URL (Export)",
    "Image URL (Import)",
    "Image Description",
    "Image is Thumbnail",
    "Image Sort Order",
    "YouTube ID",
    "Video Title",
    "Video Description",
    "Video Sort Order"
]           
        category_count = 0
        for i in result:
            if category_count < len(i['category_id']):
                category_count = len(i['category_id'])
        for i in range(1, category_count + 1):
            headers.append(f'Additional Taxonomy{i}')
        worksheet.append(headers)
        for i, item in enumerate(result):
            category_level = ""
            formatted_description = (
                item.get("short_description", "") +"\n"+
                item.get("long_description", "") +"\n"+ item.get("personalized_short_description", "") +"\n"+ item.get("personalized_long_description", "")
            )  
            category_config_obj = DatabaseModel.get_document(category_config.objects,{'id__in':item['category_id']})
            if category_config_obj ==None:
                category_level = ""
            else:
                for j in category_config_obj.levels:
                    category_level += j.name + ">"
                category_level = category_level.rstrip(">")
            attribute_list = item.get("attribute_list",[])
            feature_list = item.get("feature_list",[])
            image_list = item.get("image_list",[])
            video_list = item.get("video_list",[])
            video_str =""
            video_url =""
            image_str  = ""
            if len(image_list) > 0:
                image_str = image_list[0]['url']
            if len(video_list)>0:
                video_str = video_list[0]['name']
                video_url = video_list[0]['url']
            feature_str = "|".join(f"{item}" for item in feature_list)
            # image_str = "|".join(f"{item.url}" for item in image_list)
            attribute_list_str = list()
            for j in attribute_list:
                Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'id':j})
                if Attribute_obj:
                    value_str = ""
                    for j in Attribute_obj.values:
                        value_str += j + ","
                    value_str = value_str.rstrip(",")
                    attribute_list_str.append(f"{Attribute_obj.name}={value_str}")
            parent_level_list = []
            for x in item['category_id']:
                category_config_obj_l = DatabaseModel.get_document(category_config.objects,{'id':x})
                if category_config_obj_l:
                    for zi in category_config_obj_l.levels:
                        category_config_obj_name = DatabaseModel.get_document(category_config.objects,{'name':zi.name,'client_id':client_id})
                        if category_config_obj_name:
                            if category_config_obj_l.id != category_config_obj_name.id:
                                parent_level_list.append(str(category_config_obj_name.id))
            Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_id__in':[item['brand_id']]})
            for Attribute_obj in  Attribute_list_obj:
                if Attribute_obj:
                    value_str = ""
                    for j in Attribute_obj.values:
                        value_str += j + ","
                    value_str = value_str.rstrip(",")
                    attribute_list_str.append(f"{Attribute_obj.name}={value_str}")
            Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_id__in':parent_level_list})
            for Attribute_obj in  Attribute_list_obj:
                if Attribute_obj:
                    value_str = ""
                    for j in Attribute_obj.values:
                        value_str += j + ","
                    value_str = value_str.rstrip(",")
                    attribute_list_str.append(f"{Attribute_obj.name}={value_str}")
            Attribute_list_obj = DatabaseModel.list_documents(Attribute.objects,{'client_id':client_id,'module_name':"global"})
            for j in  Attribute_list_obj:
                if Attribute_obj:
                    value_str = ""
                    for j in Attribute_obj.values:
                        value_str += j + ","
                    value_str = value_str.rstrip(",")
                    attribute_list_str.append(f"{Attribute_obj.name}={value_str}")
            
            attribute_str =  "\n".join(f"• {item}" for item in attribute_list_str)
            row = [
                "product",
                item.get("product_id", ""), #ID
                item.get("name", ""), #Name
                "",#Type
                item.get("sku", ""), #sku
                attribute_str,#Options
                "",#Inventory Tracking
                "",#Current Stock
                "",#Low Stock
                item.get("selling_price", ""),
                "",#Cost Price
                item.get("msrp", ""),
                item.get("discount_price", ""),
                item.get("brand_name", ""),
                "",#Channels
                category_level,#Categories
                formatted_description, #Description
                feature_str,#Custom Fields
                "",#Page Title
                "",#Product URL
                "",#Meta Description
                "",#Search Keywords
                "",#Meta Keywords
                "",#Bin Picking Number
                item.get("upc", ""),#UPC/EAN
                item.get("ean", ""),#UPC/EAN
                item.get("gtin", ""),#Global Trade Number
                item.get("mpn", ""),#Manufacturer Part Number
                "",#Free Shipping
                "",#Fixed Shipping Cost
                "",#Shipping Groups
                item.get("country_of_origin", ""),#Origin Locations
                "",#Dimensional Rules
                "",#Weight
                "",#Width
                "",#Weight
                "",#Depth
                "",#Is Visible
                "",#Is Featured
                item.get("product_warranty", ""),#Warranty
                "",#Tax Class
                "",#Product Condition
                "",#Show Product Condition
                "",#Sort Order
                "",#Variant Image URL
                "",#Internal Image URL (Export)
                
                image_str,#Image URL (Import)
                "",#Image Description
                "",#Image is Thumbnail
                "",#Image Sort Order
                video_url,#YouTube ID
                video_str,#Video Title
                "",#Video Description
                "",#Video Sort Order
            ]
            channel_category_level = []
            channelCategory_obj = DatabaseModel.list_documents(channelCategory.objects,{'channel_name':"bigcommerce",'client_id':client_id,'category_config_id__in':item['category_id']})
            for iii in  channelCategory_obj:
                category_level = " > ".join(iii.taxonomy_level) if isinstance(iii.taxonomy_level, list) else str(iii.taxonomy_level)
                channel_category_level.append(category_level)
            feature_list = item.get("feature_list",[])
            feature_list += [""] * (5 - len(feature_list))
            bullet_point1, bullet_point2, bullet_point3, bullet_point4, bullet_point5 = feature_list[:5]
            if len(channel_category_level)>0:
                channel_category_level_1 = channel_category_level[0]
            for i in range(1, category_count + 1):
                if len(channel_category_level) >i:
                    row.append(channel_category_level[i])
                else:
                    row.append("")
            worksheet.append(row)
    else:
        DatabaseModel.save_documents(export_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'product','total_count':len(result)})
        worksheet.title = "Product"
        headers = [
        "S.No", "Name", "Product ID", "MPN", "SKU", "UPC", "EAN", "GTIN", "UNSPC", "Model",
        "Vendor Name", "Brand Name", "Manufacture Name", "Category 1 Name","Category 2 Name","Category 3 Name","Category 4 Name","Category 5 Name","Category 6 Name","Category 7 Name","Category 8 Name", "Category Levels", "Breadcrumb",
        "Short Description", "Personalized Short Description", "Long Description", "Personalized Long Description", "Application",
        "Certifications", "Compliance", "Prop65", "ESG", "Hazardous",
        "Service Warranty", "Product Warranty", "Country of Origin", "Currency", 
        "MSRP", "Selling Price", "Discount Price"
        ]

        feature_count = 0
        attribute_count = 0
        image_count = 0
        video_count = 0
        attachment_count = 0
        related_product_count = 0
        for i in result:
            i['attribute_list'] = i.get("attribute_list",[])
            parent_level_list = []

            # Aggregation pipeline to retrieve the category configuration by ids from i['category_id']
            aggregation_pipeline_category_config = [
                {
                    "$match": {
                        "id": {"$in": i['category_id']}  # Match categories where the id is in i['category_id']
                    }
                },
                {
                    "$lookup": {
                        "from": "category_config",  # Lookup related category config data
                        "localField": "levels.name",  # Match level name from the 'levels' array
                        "foreignField": "name",  # Match 'name' field in category_config
                        "as": "parent_category_config"  # Put the resulting documents into 'parent_category_config'
                    }
                }
            ]

            category_config_objs = category_config.objects.aggregate(*aggregation_pipeline_category_config)

            for category_config_obj in category_config_objs:
                # Iterate over the 'parent_category_config' that was populated in the previous step
                for zi in category_config_obj.get("parent_category_config", []):
                    # Check if the category config name matches and if they are from the same client_id
                    if zi.get("client_id") == client_id:
                        if category_config_obj['id'] != zi['id']:
                            parent_level_list.append(str(zi['id']))
            aggregation_pipeline_brand = [
                {
                    "$match": {
                        "client_id": client_id,
                        "module_id": {"$in": [str(i['brand_id'])]}
                    }
                }
            ]
            Attribute_list_obj_brand = Attribute.objects.aggregate(*aggregation_pipeline_brand)
            for Attribute_obj in Attribute_list_obj_brand:
                i['attribute_list'].append(Attribute_obj['_id'])

            # Aggregation pipeline for category-level attributes based on parent_level_list
            aggregation_pipeline_category = [
                {
                    "$match": {
                        "client_id": client_id,
                        "module_name": "category",
                        "module_id": {"$in": parent_level_list}
                    }
                }
            ]
            Attribute_list_obj_category = Attribute.objects.aggregate(*aggregation_pipeline_category)
            for Attribute_obj in Attribute_list_obj_category:
                i['attribute_list'].append(Attribute_obj['_id'])
            # Aggregation pipeline for global attributes
            aggregation_pipeline_global = [
                {
                    "$match": {
                        "client_id": client_id,
                        "module_name": "global"
                    }
                }
            ]
            Attribute_list_obj_global = Attribute.objects.aggregate(*aggregation_pipeline_global)
            for Attribute_obj in Attribute_list_obj_global:
                i['attribute_list'].append(Attribute_obj['_id'])
            
            if feature_count < len(i['feature_list']):
                feature_count = len(i['feature_list'])
            if attribute_count < len(i['attribute_list']):
                attribute_count = len(i['attribute_list'])
            if image_count < len(i['image_list']):
                image_count = len(i['image_list'])
            if video_count < len(i['video_list']):
                video_count = len(i['video_list'])
            if attachment_count < len(i['attachment_list']):
                attachment_count = len(i['attachment_list'])
            if related_product_count < len(i['related_products']):
                related_product_count = len(i['related_products'])
        category_count = 0

        for i in result:
            if category_count < len(i['category_id']):
                category_count = len(i['category_id'])
        for i in range(1, category_count + 1):
            headers.append(f'Additional Taxonomy{i}')
        for i in range(1, feature_count + 1):
            headers.append(f"Feature {i} Name")
        for i in range(1, attribute_count + 1):
            headers.append(f"Attribute {i} Name")
            headers.append(f"Attribute {i} Value")
        for i in range(1, related_product_count + 1):
            headers.append(f"Related Product {i} Name")
            headers.append(f"Related Product {i} URL")
        for i in range(1, image_count + 1):
            headers.append(f"Image {i} Name")
            headers.append(f"Image {i} URL")
        for i in range(1, video_count + 1):
            headers.append(f"Video {i} Name")
            headers.append(f"Video {i} URL")
        for i in range(1, attachment_count + 1):
            headers.append(f"Attachment {i} Name")
            headers.append(f"Attachment {i} URL")
        worksheet.append(headers)

        for i, item in enumerate(result):
            category_level = ""
            category_config_obj = DatabaseModel.get_document(category_config.objects,{'id__in':item['category_id']})
            category_config_obj_list = DatabaseModel.list_documents(category_config.objects,{'id__in':item['category_id']})
            if category_config_obj ==None:
                category_level = ""
            else:
                co = 1
                cat_1 = ""
                cat_2 = ""
                cat_3 = ""
                cat_4 = ""
                cat_5 = ""
                cat_6 = ""
                cat_7 = ""
                cat_8 = ""
                list_c = list()
                for j in category_config_obj.levels:
                    list_c.append(j.name)
                    category_level += j.name + ">"
                additional_category_list = []
                flag = 0
                for ice in range(category_count):
                    category_level_str = ""
                    r = ice+1
                    if len(category_config_obj_list) > r:
                        category_config_obj = category_config_obj_list[r]
                        for jce in category_config_obj.levels:
                            category_level_str += jce.name + ">"
                        additional_category_list.append(category_level_str)
                category_level = category_level.rstrip(">")
                for z_ in list_c:
                    if co ==1:
                        cat_1 = z_
                    elif co==2:
                        cat_2 = z_
                    elif co==3:
                        cat_3 = z_
                    elif co==4:
                        cat_4 = z_
                    elif co==5:
                        cat_5 = z_
                    elif co==6:
                        cat_6 = z_
                    elif co==7:
                        cat_7 = z_
                    elif co==8:
                        cat_8 = z_
                    co +=1
            row = [
                i + 1,
                item.get("name", ""),
                item.get("product_id", ""),
                item.get("mpn", ""),
                item.get("sku", ""),
                item.get("upc", ""),
                item.get("ean", ""),
                item.get("gtin", ""),
                item.get("unspc", ""),
                item.get("model", ""),
                item.get("vendor_name", ""),
                item.get("brand_name", ""),
                item.get("manufacture_name", ""),
                cat_1,
                cat_2,
                cat_3,
                cat_4,
                cat_5,
                cat_6,
                cat_7,
                cat_8,
                category_level,
                item.get("breadcrumb", ""),
                item.get("short_description", ""),
                item.get("personalized_short_description", ""),
                item.get("long_description", ""),
                item.get("personalized_long_description", ""),
                # item.get("category_group_list", ""),
                # "",
                item.get("application", ""),
                item.get("certifications", ""),
                item.get("Compliance", ""),
                item.get("Prop65", ""),
                item.get("esg", ""),
                item.get("Hazardous", ""),
                item.get("service_warranty", ""),
                item.get("product_warranty", ""),
                item.get("country_of_origin", ""),
                item.get("currency", ""),
                item.get("msrp", ""),
                item.get("selling_price", ""),
                item.get("discount_price", "")
            ]
            for j in additional_category_list:
                row.append(j)
            for i in range(category_count - len(additional_category_list)):
                row.append("")
            feature_list = item.get("feature_list",[])
            attribute_list = item.get("attribute_list",[])
            image_list = item.get("image_list",[])
            video_list = item.get("video_list",[])
            attachment_list = item.get("attachment_list",[])
            related_products = item.get("related_products",[])
            for j in feature_list:
                row.append(j)
            for i in range(feature_count - len(feature_list)):
                row.append("")
            for j in attribute_list:
                Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'id':j})
                if Attribute_obj:
                    row.append(Attribute_obj.name)
                    value_str = ""
                    for j in Attribute_obj.values:
                        value_str += j + ","
                    value_str = value_str.rstrip(",")
                    row.append(value_str)
                else:
                    row.append("")
                    row.append("")
            for i in range(attribute_count - len(attribute_list)):
                row.append("")
                row.append("")
            for j in related_products:
                row.append(j.get('name', ''))
                row.append(j.get('url', ''))
            for i in range(related_product_count - len(related_products)):
                row.append("")
                row.append("")
            for j in image_list:
                row.append(j.get('name', ''))
                row.append(j.get('url', ''))
            for i in range(image_count - len(image_list)):
                row.append("")
                row.append("")
            for j in video_list:
                row.append(j.get('name', ''))
                row.append(j.get('url', ''))
            for i in range(video_count - len(video_list)):
                row.append("")
                row.append("")
            for j in attachment_list:
                row.append(j.get('name', ''))
                row.append(j.get('url', ''))
            for i in range(attachment_count - len(attachment_list)):
                row.append("")
                row.append("")
            
            worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0) 
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="Product.xlsx"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@csrf_exempt
def exportAttribute(request):
    client_id = get_current_client()
    is_visible = request.GET.get('is_active')
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')

    export_log_obj = DatabaseModel.save_documents(export_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'brand'})

    if is_visible == None:
        visible_obj = {'client_id':ObjectId(client_id),'module_name':{'$ne':'product'}}
    elif is_visible == "true":
        visible_obj = {'client_id':ObjectId(client_id),'is_visible':True,'module_name':{'$ne':'product'}}
    elif is_visible == "false":
        visible_obj = {'client_id':ObjectId(client_id),'is_visible':False,'module_name':{'$ne':'product'}}
    pipeline = [
       {
            '$match': visible_obj
        },
         {
            '$match':{'client_id':ObjectId(client_id)}
        },
        {
            "$group": {
                "_id": "$_id",
                "name":{ "$first":"$name"},
                "type":{ "$first":"$type"},
                "code":{ "$first":"$code"},
                "values":{ "$first":"$values"},
                "module_name":{ "$first":"$module_name"},
                "module_id":{ "$first":"$module_id"},
        }
    },
    {
        '$sort':{'_id':-1}
    }
    ]
    result = list(Attribute.objects.aggregate(*pipeline))
    workbook = Workbook()
    export_log_obj.total_count = len(result)
    export_log_obj.save()
    worksheet = workbook.active
    worksheet.title = "Attribute"
    headers = ["S.No","Attribute Name","Type","Module Based Name","Brand","Category Names","Value 1","Value 2","Value 3","Value 4","Value 5","Value 6","Value 7","Value 8","Value 9","Value 10"]
    worksheet.append(headers)
    for i, item in enumerate(result):
        # y = ",".join(item.get("values", []))
        module_name = item.get("module_name")
        brand_module_id_name = list()
        category_module_id_name = list()
        if module_name == 'brand':
            brand_list = DatabaseModel.list_documents(brand.objects,{'id__in':item.get("module_id")})
            for t__ in brand_list:
                brand_module_id_name.append(str(t__.name)) 
        elif module_name == 'category':
            category_config_list = DatabaseModel.list_documents(category_config.objects,{'id__in':item.get("module_id")})
            for t__ in category_config_list:
                category_module_id_name.append(str(t__.name))
        brand_module_id_name = "".join(brand_module_id_name)
        category_module_id_name = "".join(category_module_id_name)

        row = [
            i + 1,
            item.get("name", ""),
            item.get("type", ""),
            # item.get("code", ""),
            module_name,
            brand_module_id_name,
            category_module_id_name
        ]
        row.extend(item.get("values", []))
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0) 
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="Attribute.xlsx"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response
import pycountry
import math
def process_features(product_obj, get_field):
    i = 1
    while True:
        feature_name = get_field(f"Feature {i} Name")
        if not feature_name:
            break
        product_obj.feature_list.append(feature_name)
        # product_obj.save()
        i += 1

def process_attributes(product_obj, get_field, client_id):
    i = 1
    while True:
        attribute_name = get_field(f"Attribute {i} Name")
        attribute_value = get_field(f"Attribute {i} Value")
        if not attribute_name or not attribute_value:
            break
        Attribute_obj = DatabaseModel.get_document(
            Attribute.objects,
            {'name': attribute_name, 'module_id__in': [str(product_obj.id)], 'client_id': client_id}
        )
        if Attribute_obj is None:
            Attribute_obj = DatabaseModel.save_documents(
                Attribute,
                {'name': attribute_name, 'module_name': 'product', 'module_id': [str(product_obj.id)], 'values': [attribute_value]}
            )
            product_obj.attribute_list.append(Attribute_obj.id)
        else:
            DatabaseModel.update_documents(
                Attribute.objects,
                {'id': Attribute_obj.id},
                {'add_to_set__values': attribute_value}
            )
            product_obj.attribute_list.append(Attribute_obj.id)
        # product_obj.save()
        i += 1

def process_related_products(product_obj, get_field):
    i = 1
    while True:
        related_name = get_field(f"Related Product {i} Name")
        related_url = get_field(f"Related Product {i} URL")
        if not related_name or not related_url:
            break
        product_obj.related_products.append(RelatedProduct(name=related_name, url=related_url))
        # product_obj.save()

        i += 1

def process_category_groups(product_obj, get_field):
    i = 1
    while True:
        group_name = get_field(f"Category Group {i} Name")
        group_taxonomy = get_field(f"Category Group {i} Taxonomy")
        if not group_name or not group_taxonomy:
            break
        b2c_company_obj = DatabaseModel.get_document(b2c_company.objects, {'name': group_name})
        if b2c_company_obj is None:
            b2c_company_obj = DatabaseModel.save_documents(b2c_company, {'name': group_name})
        group_taxonomy = [category.strip() for category in group_taxonomy.split(">")]
        category_levels = []
        for z1 in group_taxonomy:
            category_group_obj = DatabaseModel.get_document(category_group.objects, {'name': z1})
            if category_group_obj is None:
                category_group_obj = DatabaseModel.save_documents(category_group, {'name': z1})
            category_levels.append(category_group_obj.id)
        product_obj.category_group_list.append(
            category_group_config(b2c_company_id=b2c_company_obj.id, category_levels=category_levels)
        )
        # product_obj.save()

        i += 1

def process_attachments(product_obj, row_dict, client_id, num, dict_error, name, get_field):
    i = 1
    while True:
        attachment_name = row_dict.get(f"Attachment {i} Name")
        attachment_url = row_dict.get(f"Attachment {i} URL")
        if not attachment_name or not attachment_url:
            break
        try:
            ProductDocument_obj = DatabaseModel.get_document(ProductDocument.objects, {
                'original_url__in': [attachment_url], 'client_id': client_id
            })

            if ProductDocument_obj is None:
                upload_result = cloudinary.uploader.upload(
                    attachment_url, resource_type="raw", folder="KM-DIGI/doc"
                )
                document_url = upload_result.get("secure_url")
                public_id = upload_result.get("public_id")

                product_document = ProductDocument(
                    name=attachment_name, document_url=document_url,
                    public_id=public_id, client_id=client_id,
                    original_url=[attachment_url, document_url]
                )
                product_document.save()
                product_obj.attachment_list.append(Attachment(name=attachment_name, url=document_url))
            else:
                product_obj.attachment_list.append(Attachment(
                    name=ProductDocument_obj.name,
                    url=ProductDocument_obj.document_url
                ))
        except Exception:
            dict_error['error-row'] = num + 2
            dict_error['error_list'].append(f"Public URL required Attachment {i} URL")
            dict_error['product_sku'] = str(get_field("SKU"))
            dict_error['product_name'] = name
        
        # product_obj.save()
        
        i += 1


def process_images(product_obj, row_dict, client_id, num, dict_error, name, get_field):
    i = 1
    while True:
        image_name = row_dict.get(f"Image {i} Name")
        image_url = row_dict.get(f"Image {i} URL")
        if not image_name or not image_url:
            break
        try:
            ProductImage_obj = DatabaseModel.get_document(ProductImage.objects, {
                'original_url__in': [image_url], 'client_id': client_id
            })

            if ProductImage_obj is None:
                upload_result = cloudinary.uploader.upload(image_url, folder="KM-DIGI/image")
                image_url_ = upload_result.get("secure_url")
                public_id = upload_result.get("public_id")

                product_image = ProductImage(
                    name=image_name, image_url=image_url_,
                    public_id=public_id, client_id=client_id,
                    original_url=[image_url, image_url_]
                )
                product_image.save()
                product_obj.image_list.append(Image(name=image_name, url=image_url_))
            else:
                product_obj.image_list.append(Image(
                    name=ProductImage_obj.name,
                    url=ProductImage_obj.image_url
                ))
        except Exception:
            dict_error['error-row'] = num + 2
            dict_error['error_list'].append(f"Public URL required Image {i} URL")
            dict_error['product_sku'] = str(get_field("SKU"))
            dict_error['product_name'] = name
        # product_obj.save()
        
        i += 1


def process_videos(product_obj, row_dict, client_id, num, dict_error, name, get_field):
    i = 1
    while True:
        video_name = row_dict.get(f"Video {i} Name")
        video_url = row_dict.get(f"Video {i} URL")
        if not video_name or not video_url:
            break
        try:
            ProductVideo_obj = DatabaseModel.get_document(ProductVideo.objects, {
                'original_url__in': [video_url], 'client_id': client_id
            })

            if ProductVideo_obj is None:
                upload_result = cloudinary.uploader.upload(
                    video_url, resource_type="video", folder="KM-DIGI/video"
                )
                video_url_ = upload_result.get("secure_url")
                public_id = upload_result.get("public_id")

                product_video = ProductVideo(
                    name=video_name, video_url=video_url_,
                    public_id=public_id, client_id=client_id,
                    original_url=[video_url, video_url_]
                )
                product_video.save()
                product_obj.video_list.append(Video(name=video_name, url=video_url_))
            else:
                product_obj.video_list.append(Video(
                    name=ProductVideo_obj.name,
                    url=ProductVideo_obj.video_url
                ))
        except Exception:
            dict_error['error-row'] = num + 2
            dict_error['error_list'].append(f"Public URL required Video {i} URL")
            dict_error['product_sku'] = str(get_field("SKU"))
            dict_error['product_name'] = name
        # product_obj.save()
        
        i += 1
        
@csrf_exempt
def importProduct(request):
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')

    client_id =ObjectId(get_current_client())
    data = {'status': False,'error_list':[],'is_error':False,'error_count':0,'added_count':0,'field_error':0}
    import_log_obj = DatabaseModel.save_documents(import_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'product'})

    if 'file' not in request.FILES:
        return JsonResponse(data,safe=False)
    file = request.FILES['file']
    if file.name.endswith('.xlsx'):
        df = pd.read_excel(file)
    elif file.name.endswith('.csv') or file.name.endswith('.txt'):
        df = pd.read_csv(file)
    elif file.name.endswith('.ods') :
        df = pd.read_excel(file, engine='odf')
    else:
        return JsonResponse(data,safe=False)  
    count = 0
    data_list = list(df.iterrows()) 
    import_log_obj.total_count = len(data_list)
    import_log_obj.save()
    if len(data_list) == 0:
        data['is_error'] = True
        data['error'] = "Excel file should not be empty"
    for num, row in df.iterrows():
        dict_error = dict()
        import_log_obj.inprogress_count = import_log_obj.total_count - import_log_obj.completed_count 
        import_log_obj.save()
        dict_error['is_error'] =False
        dict_error['error_list']  = list()
        flag = False
        count +=1
        row_dict = row.to_dict()
        for key, value in row_dict.items():
            if value is None or (isinstance(value, float) and math.isnan(value)):
                row_dict[key] = None
        normalized_dict = {key.lower(): value for key, value in row_dict.items()}
        def get_field(field_name, default=""):
            ans = normalized_dict.get(field_name.lower(), default)
            if ans == None:
                ans = ""
            else:
                ans = str(ans)
            return ans
        category_number = 1
        category_id_list = list()
        category_name = ""
        last_config_id = None
        while f'Category {category_number} Name' in row_dict :
            category_name = row_dict[f'Category {category_number} Name']
            category_name = clean_value(category_name, default="")
            
            if category_name:
                category_name = str(category_name).title()
                category_obj = DatabaseModel.get_document(category.objects,{'name__iexact':category_name,'client_id':client_id})
                if category_obj == None:
                    category_obj = DatabaseModel.save_documents(category,{'name':category_name,'client_id':client_id})
                category_config_obj_past = DatabaseModel.get_document(category_config.objects,{'levels':category_id_list,'client_id':client_id})
                category_id_list.append(category_obj.id)
                category_config_obj = DatabaseModel.get_document(category_config.objects,{'name__iexact':category_name,'levels':category_id_list})
                if category_config_obj == None and category_name != "":
                    if category_config_obj_past:
                        category_config_obj_past.end_level = False
                        category_config_obj_past.save()
                    category_config_obj = DatabaseModel.save_documents(category_config,{'name':category_name,'levels':category_id_list,'end_level':True})
                    # DatabaseModel.save_documents(category_log,{'user_id':user_login_id,'client_id':client_id,'action':'Create (Import)','category_id':category_config_obj.id})
                last_config_id = category_config_obj.id
            category_number +=1
        if f'Category 1 Name' in row_dict:
            category_name = row_dict[f'Category 1 Name']
            category_name = clean_value(category_name, default="")
            if category_name == "":
                dict_error['error-row'] = num + 2
                dict_error['error_list'].append("Category 1 Name is Mandatory")
                dict_error['is_error'] = True
                flag = True
        vendor_name = get_field("Vendor Name")
        brand_name = get_field("Brand Name")
        manufacture_name = get_field("Manufacture Name")
        Vendor_id = None
        if vendor_name:
            Vendor_obj = DatabaseModel.get_document(Vendor.objects,{'name__iexact':vendor_name,'client_id':client_id})
            if Vendor_obj == None:
                Vendor_obj = DatabaseModel.save_documents(Vendor,{'name':vendor_name,'client_id':client_id})
                # DatabaseModel.save_documents(vendor_log,{'user_id':user_login_id,'client_id':client_id,'action':'Create (Import)','vendor_id':Vendor_obj.id})

            Vendor_id = Vendor_obj.id
        brand_id = None
        if brand_name:
            brand_obj = DatabaseModel.get_document(brand.objects,{'name__iexact':brand_name,'client_id':client_id})
            if brand_obj == None:
                brand_obj = DatabaseModel.save_documents(brand,{'name':brand_name,'client_id':client_id})
                # DatabaseModel.save_documents(brand_log,{'user_id':user_login_id,'client_id':client_id,'action':'Create (Import)','brand_id':brand_obj.id})
            brand_id = brand_obj.id
        else:
            dict_error['error-row'] = num + 2
            dict_error['error_list'].append("Brand is Mandatory")
            dict_error['is_error'] = True
            flag = True
        Manufacture_id = None
        if manufacture_name:
            Manufacture_obj = DatabaseModel.get_document(Manufacture.objects,{'name__iexact':manufacture_name,'client_id':client_id})
            if Manufacture_obj== None:
                Manufacture_obj = DatabaseModel.save_documents(Manufacture,{'name':manufacture_name,'client_id':client_id})
            Manufacture_id = Manufacture_obj.id
        if str(get_field("Name")) == "":
            dict_error['error-row'] = num + 2
            dict_error['error_list'].append("Name is Mandatory")
            dict_error['is_error'] = True
            flag = True
        if str(get_field("SKU")) == "":
            dict_error['error-row'] = num + 2
            dict_error['error_list'].append("SKU is Mandatory")
            dict_error['is_error'] = True
            flag = True
        # if str(get_field("MPN")) =="":
        #     dict_error['error-row'] = num + 2
        #     dict_error['error_list'].append("MPN should be Mandatory")
        #     dict_error['is_error'] = True
        #     flag = True
        COUNTRY_ALIASES = {
            "america": "United States",
            "usa": "United States",
            "uk": "United Kingdom",
            "uae": "United Arab Emirates",
            "south korea": "Korea, Republic of",
            "north korea": "Korea, Democratic People's Republic of",
            "ivory coast": "Côte d'Ivoire",
        }
        def get_country_code(country_name):
            if not country_name:
                return None
            country_name = country_name.strip().lower()
            normalized_name = COUNTRY_ALIASES.get(country_name, country_name)
            try:
                country = pycountry.countries.lookup(normalized_name)
                return country.alpha_2
            except LookupError:
                return None 
        country_name = get_field("Country of Origin") 
        country_code = get_country_code(country_name)
        if flag == False:
            field_error = False
            name = str(get_field("Name")).title()
            product_obj = DatabaseModel.get_document(product.objects,{'sku':str(get_field("SKU")),'client_id':client_id})
            product_dict = {
                    "product_id": str(get_field("Product ID")),
                    "mpn": str(get_field("MPN")),
                    "sku": str(get_field("SKU")),
                    "upc": str(get_field("UPC")),
                    "ean": str(get_field("EAN")),
                    "gtin": str(get_field("GTIN")),
                    "unspc": str(get_field("UNPSC")),
                    "model":str(get_field("Model")),
                    "name": name,
                    "client_id": client_id, 
                    "vendor_id": Vendor_id,  
                    "brand_id": brand_id, 
                    "category_id": [last_config_id],
                    "manufacture_id": Manufacture_id,
                    "breadcrumb":str( get_field("Breadcrumb")),
                    "short_description": str(get_field("Short Description")),
                    "personalized_short_description": str(get_field("Personalized Short Description")),
                    "long_description": str(get_field("Long Description")),
                    "personalized_long_description": str(get_field("Personalized Long Description")),
                    "feature_list": [], 
                    "attribute_list": [],
                    "related_products": [], 
                    "application":str(get_field("Application")),
                    "certifications": str(get_field("Certifications")),
                    "Compliance":( get_field("Compliance")),
                    "Prop65": (get_field("Prop65")),
                    "esg": (get_field("ESG")),
                    "Hazardous": (get_field("Hazardous")),
                    "service_warranty": (get_field("Service Warranty")),
                    "product_warranty": (get_field("Product Warranty")),
                    "country_of_origin": country_code,
                    "currency": (get_field("Currency")),
                    "msrp":( get_field("MSRP")),
                    "selling_price": (get_field("Selling Price")),
                    "discount_price": (get_field("Discount Price")),
                    "attachment_list": [], 
                    "image_list": [],  
                    "video_list": [], 
                    "category_group_list":[],
            }
            if product_obj == None:
               
                product_obj = DatabaseModel.save_documents(product,product_dict)
                # DatabaseModel.save_documents(product_log,{'user_id':user_login_id,'client_id':client_id,'action':'Create (Import)','product_id':product_obj.id})
                import_log_obj.created_count +=1
                import_log_obj.created_id_list.append(str(product_obj.id))
                import_log_obj.save()

            else:
                product_obj = DatabaseModel.update_documents(product.objects,{ "id": product_obj.id},product_dict)
                # DatabaseModel.save_documents(product_log,{'user_id':user_login_id,'client_id':client_id,'action':'Import-Updated','product_id':product_obj.id})
                import_log_obj.updated_count +=1
                if product_obj:
                    import_log_obj.updated_id_list.append(str(product_obj.id))
                    import_log_obj.save()



            # i = 1
            # while True:
            #     feature_name = get_field(f"Feature {i} Name")
            #     # feature_value = get_field(f"Feature {i} Value")
            #     if not feature_name:
            #         break
            #     product_obj.feature_list.append(feature_name)
            #     i += 1
            # i = 1
            # while True:
            #     attribute_name = get_field(f"Attribute {i} Name")
            #     attribute_value = get_field(f"Attribute {i} Value")
            #     if not attribute_name or not attribute_value:
            #         break
            #     Attribute_obj = DatabaseModel.get_document(Attribute.objects,{'name':attribute_name,'module_id__in':[str(product_obj.id)],'client_id':client_id})
            #     if Attribute_obj == None:
            #         Attribute_obj = DatabaseModel.save_documents(Attribute,{'name':attribute_name,'module_name':'product','module_id':[str(product_obj.id)],'values':[attribute_value]})
            #         product_obj.attribute_list.append(Attribute_obj.id)
            #     else:
            #         DatabaseModel.update_documents(Attribute.objects,{'id':Attribute_obj.id},{'add_to_set__values':attribute_value})
            #         product_obj.attribute_list.append(Attribute_obj.id)

            #     i += 1
            # i = 1
            # while True:
            #     related_name = get_field(f"Related Product {i} Name")
            #     related_url = get_field(f"Related Product {i} URL")
            #     if not related_name or not related_url:
            #         break
            #     product_obj.related_products.append(RelatedProduct(name=related_name, url=related_url))
            #     i += 1
            # i = 1
            # while True:
            #     group_name = get_field(f"Category Group {i} Name")
            #     group_taxonomy = get_field(f"Category Group {i} Taxonomy")
            #     if not group_name or not group_taxonomy:
            #         break
            #     b2c_company_obj = DatabaseModel.get_document(b2c_company.objects,{'name':group_name})
            #     if b2c_company_obj == None:
            #         b2c_company_obj = DatabaseModel.save_documents(b2c_company,{'name':group_name})
            #     group_taxonomy = [category.strip() for category in group_taxonomy.split(">")]
            #     category_levels = list()
            #     for z1 in group_taxonomy:
            #         category_group_obj = DatabaseModel.get_document(category_group.objects,{'name':z1})
            #         if category_group_obj == None:
            #             category_group_obj = DatabaseModel.save_documents(category_group,{'name':z1})
            #         category_levels.append(category_group_obj.id)
            #     product_obj.category_group_list.append(category_group_config(b2c_company_id=b2c_company_obj.id, category_levels=category_levels))
            #     i += 1
            threads = [
                Thread(target=process_features, args=(product_obj, get_field)),
                Thread(target=process_attributes, args=(product_obj, get_field, client_id)),
                Thread(target=process_related_products, args=(product_obj, get_field)),
                Thread(target=process_category_groups, args=(product_obj, get_field)),
                Thread(target=process_attachments, args=(product_obj, row_dict, client_id, num, dict_error, name, get_field)),
                Thread(target=process_images, args=(product_obj, row_dict, client_id, num, dict_error, name, get_field)),
                Thread(target=process_videos, args=(product_obj, row_dict, client_id, num, dict_error, name, get_field)),
            ]

            # Start all threads
            for thread in threads:
                thread.start()

            # Wait (join) all threads
            for thread in threads:
                thread.join()

            # Now all processing is finished → safe to save
            product_obj.save()
                        # i = 1
            # while True:
            #     attachment_name = row_dict.get(f"Attachment {i} Name")
            #     attachment_url = row_dict.get(f"Attachment {i} URL")
            #     if not attachment_name or not attachment_url:
            #         break
            #     try:
            #         ProductDocument_obj = DatabaseModel.get_document(ProductDocument.objects, {'original_url__in': [attachment_url],'client_id':client_id})

            #         if ProductDocument_obj is None:
            #             upload_result = cloudinary.uploader.upload(attachment_url, resource_type="raw", folder="KM-DIGI/doc")
            #             document_url = upload_result.get("secure_url")
            #             public_id = upload_result.get("public_id")

            #             product_document = ProductDocument(name=attachment_name, document_url=document_url, public_id=public_id, client_id=client_id,original_url = [attachment_url,document_url])
            #             product_document.save()
            #             product_obj.attachment_list.append(Attachment(name=attachment_name, url=document_url))

            #         else:
            #             product_obj.attachment_list.append(Attachment(name=ProductDocument_obj.name, url=ProductDocument_obj.document_url))
            #     except Exception as e:
            #         dict_error['error-row'] = num + 2 
            #         dict_error['error_list'].append(f"Public URL required Attachment {i} URL")
            #         dict_error['product_sku'] = str(get_field("SKU"))
            #         dict_error['product_name'] = name
            #     i += 1

            # # Process Images
            # i = 1
            # while True:
            #     image_name = row_dict.get(f"Image {i} Name")
            #     image_url = row_dict.get(f"Image {i} URL")
            #     if not image_name or not image_url:
            #         break
            #     try:
            #         ProductImage_obj = DatabaseModel.get_document(ProductImage.objects, {'original_url__in': [image_url],'client_id':client_id})
            #         if ProductImage_obj is None:
            #             upload_result = cloudinary.uploader.upload(image_url, folder="KM-DIGI/image")
            #             image_url_ = upload_result.get("secure_url")
            #             public_id = upload_result.get("public_id")

            #             product_image = ProductImage(name=image_name, image_url=image_url_, public_id=public_id, client_id=client_id,original_url =[image_url,image_url_] )
            #             product_image.save()
            #             product_obj.image_list.append(Image(name=image_name, url=image_url_))

            #         else:
            #             product_obj.image_list.append(Image(name=ProductImage_obj.name, url=ProductImage_obj.image_url))
            #     except Exception as e:
            #         dict_error['error-row'] = num + 2
            #         dict_error['error_list'].append(f"Public URL required Image {i} URL")
            #         dict_error['product_sku'] = str(get_field("SKU"))
            #         dict_error['product_name'] = name
            #     i += 1

            # i = 1
            # while True:
            #     video_name = row_dict.get(f"Video {i} Name")
            #     video_url = row_dict.get(f"Video {i} URL")
            #     if not video_name or not video_url:
            #         break
            #     try:
            #         ProductVideo_obj = DatabaseModel.get_document(ProductVideo.objects, {'original_url__in': [video_url],'client_id':client_id})

            #         if ProductVideo_obj is None:
            #             upload_result = cloudinary.uploader.upload(video_url, resource_type="video", folder="KM-DIGI/video")
            #             video_url_ = upload_result.get("secure_url")
            #             public_id = upload_result.get("public_id")

            #             product_video = ProductVideo(name=video_name, video_url=video_url_, public_id=public_id, client_id=client_id,original_url = [video_url,video_url_])
            #             product_video.save()
            #             product_obj.video_list.append(Video(name=video_name, url=video_url_))

            #         else:
            #             product_obj.video_list.append(Video(name=ProductVideo_obj.name, url=ProductVideo_obj.video_url))
            #     except Exception as e:
            #         dict_error['error-row'] = num + 2
            #         dict_error['error_list'].append(f"Public URL required Video {i} URL")
            #         dict_error['product_sku'] = str(get_field("SKU"))
            #         dict_error['product_name'] = name
            #     i += 1
            product_obj.save()
            if flag == False:
                data['added_count'] +=1 
            if field_error:
                data['field_error'] +=1 
        # else:
        if len(dict_error['error_list'])>0:
            data['error_list'].append(dict_error)
            if dict_error['is_error']:
                data['is_error'] = True
                data['error_count'] +=1
        import_log_obj.completed_count +=1
        import_log_obj.inprogress_count -=1
        import_log_obj.save()
    if len(data['error_list'])>0:
        data['status'] = False  
    else:
        data['status'] = True   
    import_log_obj.error_count = len(data['error_list'])
    import_log_obj.un_error_count = count - len(data['error_list'])
    import_log_obj.data = data
    import_log_obj.status = "completed"
    import_log_obj.save() 
    data['total_product'] = count
    return JsonResponse(data,safe=False)

import io
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

def sample_ots_file(request):
    headers1 = ["S.No", "Attribute Name", "Type", "Module Based Name", "Brand", "Category Names", "Value 1", "Value 2"]
    sample_data1 = []
    df_attributes = pd.DataFrame(sample_data1, columns=headers1)

    # Sheet2 Data
    headers2 = ["Field Name", "Short Description", "Examples", "Mandatory"]
    sample_data2 = [
        ["Attributes Name", "Name of the Attributes", "Color", "Yes"],
        ["Type", "The text type of the attributes value", "Text", ""],
        ["Module based Name", "Specifies if the attribute is created for a Category, Brand, or Global", "Select either brand/ category/Global", "Yes"],
        ["Brand", "Brand associated with the attribute", "Lenovo", "Yes"],
        ["Category Name", "Category under which the attribute is created.", "Laptops", "Yes"],
        ["Value 1", "First value for the attribute", "Black", "Yes"],
        ["Value 2", "Second value for the attribute", "Grey", ""]
    ]
    df_sheet2 = pd.DataFrame(sample_data2, columns=headers2)

    # Fetching brand and category data from MongoDB
    client_id = get_current_client()
    
    pipeline = [
        {"$match": {"_id": ObjectId(client_id)}},
        {"$lookup": {"from": "brand", "localField": "_id", "foreignField": "client_id", "as": "brand"}},
        {"$lookup": {"from": "category_config", "localField": "_id", "foreignField": "client_id", "as": "category_config"}},
        {"$project": {"brand_list": "$brand.name", "category_list": "$category_config.name"}}
    ]

    objects = list(client.objects.aggregate(pipeline))
    brand_list = objects[0].get("brand_list", []) if objects else []
    category_list = objects[0].get("category_list", []) if objects else []

    output = io.BytesIO()
    file_format = request.GET.get("file_format", "xlsx")

    if file_format == "xlsx":
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_attributes.to_excel(writer, sheet_name="Attribute", index=False)
            df_sheet2.to_excel(writer, sheet_name="Field Descriptions", index=False)

            wb = writer.book
            ws = wb["Attribute"]

            # Define dropdowns
            module_based_dropdown = DataValidation(type="list", formula1='"text,integer,decimal,boolean,multiselect"', allow_blank=True)
            module_based_dropdown_bc = DataValidation(type="list", formula1='"brand,category,global"', allow_blank=True)

            def format_dropdown_options(options):
                """Formats a list of options for Excel's DataValidation formula1."""
                if not options:
                    return '"Option1,Option2"'  # Default options

                formatted_options = ','.join(options).replace('"', '""')  # Escape quotes
                return f'"{formatted_options[:250]}"'  # Ensure within 255-character limit

            brand_options = format_dropdown_options(brand_list)
            category_options = format_dropdown_options(category_list)

            brand_dropdown = DataValidation(type="list", formula1=brand_options, allow_blank=True)
            category_dropdown = DataValidation(type="list", formula1=category_options, allow_blank=True)

            # Apply validation for multiple rows
            for row in range(2, 101): 
                ws[f"C{row}"].font = Font(bold=True)  # Type
                ws[f"D{row}"].font = Font(bold=True)  # Module Based Name
                ws[f"E{row}"].font = Font(bold=True)  # Brand
                ws[f"F{row}"].font = Font(bold=True)  # Category

                module_based_dropdown.add(ws[f"C{row}"])
                module_based_dropdown_bc.add(ws[f"D{row}"])
                brand_dropdown.add(ws[f"E{row}"])
                category_dropdown.add(ws[f"F{row}"])

            ws.add_data_validation(module_based_dropdown)
            ws.add_data_validation(module_based_dropdown_bc)
            ws.add_data_validation(brand_dropdown)
            ws.add_data_validation(category_dropdown)

        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = "sampleAttribute.xlsx"

    output.seek(0)
    response = HttpResponse(output, content_type=content_type)
    response["Content-Disposition"] = f'attachment; filename={filename}'
    return response





@csrf_exempt
def updateVisibilityForattribute(request):
    json_req = JSONParser().parse(request)
    
    DatabaseModel.update_documents(Attribute.objects,{'id':json_req['id']},{'is_visible':json_req['is_visible']})
    
    data = dict()
    data['is_update'] = True
    return data


def obtainManufacture(request):
    client_id = get_current_client()
    data = dict()
    manufacture_id = request.GET.get('id')
    if manufacture_id:
        match = {
            '$match':{'client_id':ObjectId(client_id),'_id':ObjectId(manufacture_id)}
        }
    else:
        match = {
            '$match':{'client_id':ObjectId(client_id)}
        }

    pipeline = [
        match,
    {
            '$group': {
                "_id":'$_id',
                'name':{'$first':'$name'},
        }
        },
    {
        '$sort': { '_id': -1 }
    },
    ]
    manufacture_list = list(Manufacture.objects.aggregate(*pipeline))
    for i in manufacture_list:
        i['id'] = str (i['_id'])
        del i['_id']
    data['manufacture_list'] = manufacture_list
    data['manufacture_count'] = len(data['manufacture_list'])
    return data


import requests
from django.conf import settings
from rest_framework.response import Response
from rest_framework.decorators import api_view

SHOPIFY_API_URL = "https://your-store.myshopify.com/admin/api/2024-01"
HEADERS = {
    "X-Shopify-Access-Token": settings.SHOPIFY_ACCESS_TOKEN
}

@api_view(["GET"])
def get_categories(request):

    custom_collections = requests.get(f"{SHOPIFY_API_URL}/custom_collections.json", headers=HEADERS).json()


    smart_collections = requests.get(f"{SHOPIFY_API_URL}/smart_collections.json", headers=HEADERS).json()


    products = requests.get(f"{SHOPIFY_API_URL}/products.json?fields=product_type", headers=HEADERS).json()
    product_types = list(set(p["product_type"] for p in products.get("products", []) if p["product_type"]))


    categories = {
        "custom_categories": custom_collections.get("custom_collections", []),
        "shopify_default_categories": smart_collections.get("smart_collections", []),
        "product_types": product_types
    }

    return Response(categories)

def obtainClientAttachment(request):
    client_id = get_current_client()
    pipeline = [
        {
            '$match': {'_id':ObjectId(client_id)}
        },
        {
        '$lookup': {
            'from': 'product_image',
            'localField': '_id',
            'foreignField': 'client_id',
            'as': 'product_image'
        }
    }, 
    {
            '$unwind': {
                'path': '$product_image',
                'preserveNullAndEmptyArrays': True
            }
    },
    {
        '$lookup': {
            'from': 'product_video',
            'localField': '_id',
            'foreignField': 'client_id',
            'as': 'product_video'
        }
    }, 
    {
            '$unwind': {
                'path': '$product_video',
                'preserveNullAndEmptyArrays': True
            }
    },{
        '$lookup': {
            'from': 'product_document',
            'localField': '_id',
            'foreignField': 'client_id',
            'as': 'product_document'
        }
    }, 
    {
            '$unwind': {
                'path': '$product_document',
                'preserveNullAndEmptyArrays': True
            }
    },
        {
            '$group': {
                "_id": "$_id",
                'image': { '$addToSet': {'name':'$product_image.name','url':"$product_image.image_url"} },
                'video': { '$addToSet': {'name':'$product_video.name','url':"$product_video.video_url"} },
                'document': { '$addToSet': {'name':'$product_document.name','url':"$product_document.document_url"} },
            }
        },
          {
        '$addFields': {
            'image': { '$cond': { 'if': { '$eq': ['$image', [{}]] }, 'then': '$$REMOVE', 'else': '$image' } },
            'video': { '$cond': { 'if': { '$eq': ['$video', [{}]] }, 'then': '$$REMOVE', 'else': '$video' } },
            'document': { '$cond': { 'if': { '$eq': ['$document', [{}]] }, 'then': '$$REMOVE', 'else': '$document' } },
        }
    }
    ]

    client_obj = list(client.objects.aggregate(*pipeline))
    if client_obj:
        client_obj = client_obj[0]
        del client_obj['_id']
    data = dict()
    # data['client_obj'] = client_obj
    return data


@csrf_exempt
def importDAM(request):
    data = {'status': False,'is_error':False,'error_count':0,'added_count':0}
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')

    client_id =ObjectId(get_current_client())
    import_log_obj = DatabaseModel.save_documents(import_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'dam'})

    if 'file' not in request.FILES:
        return JsonResponse(data,safe=False)
    file = request.FILES['file']
    if file.name.endswith('.xlsx'):
        df = pd.read_excel(file)
    elif file.name.endswith('.csv') or file.name.endswith('.txt'):
        df = pd.read_csv(file)
    elif file.name.endswith('.ods'):
        df = pd.read_excel(file, engine='odf')
    else:
        return JsonResponse(data,safe=False)  
    data['error_list'] = []
    count = 0
    data_list = list(df.iterrows()) 
    if len(data_list) == 0:
        data['is_error'] = True
        data['error'] = "Excel file should not be empty"
    import_log_obj.total_count = len(data_list)
    import_log_obj.save()
    for integer, row in df.iterrows():
        count +=1
        import_log_obj.inprogress_count = import_log_obj.total_count - import_log_obj.completed_count 
        import_log_obj.save()
        dict_error = dict()
        dict_error['is_error'] =False
        dict_error['error_list']  = list()
        row_dict = row.to_dict()
        for key, value in row_dict.items():
            if value is None or (isinstance(value, float) and math.isnan(value)):
                row_dict[key] = None
        sku = row_dict.get('sku') or row_dict.get('SKU') or row_dict.get('sku number')
        sku = clean_value(sku, default="")
        sku = str(sku)
        if sku == "":
            dict_error['error-row'] = integer + 2
            dict_error['error_list'].append("SKU is Mandatory")
            dict_error['is_error'] = True

        else:
            product_obj = DatabaseModel.get_document(product.objects,{'sku':(sku),'client_id':client_id})
            if product_obj:
                import_log_obj.created_id_list.append(str(product_obj.id))
                import_log_obj.save()
                url_error = False
                sku = product_obj.sku
                image_count = len(product_obj.image_list)
                video_count = len(product_obj.video_list)
                doc_count = len(product_obj.attachment_list)
                i = 1
                while True:
                    doc_count += 1
                    attachment_name = f"{sku}-Doc-{doc_count}"
                    attachment_url = row_dict.get(f"Attachment {i} URL", "")
                    if not attachment_url:
                        break
                    try:
                        ProductDocument_obj = DatabaseModel.get_document(ProductDocument.objects, {'original_url__in': [attachment_url],'client_id':client_id})
                        if ProductDocument_obj is None:
                            upload_result = cloudinary.uploader.upload(attachment_url, resource_type="raw", folder="KM-DIGI/doc")
                            document_url = upload_result.get("secure_url")
                            public_id = upload_result.get("public_id")
                            product_document = ProductDocument(name=attachment_name, document_url=document_url, public_id=public_id, client_id=client_id,original_url = [attachment_url,document_url])
                            product_document.save()
                            product_obj.attachment_list.append(Attachment(name=attachment_name, url=document_url))
                        else:
                            product_obj.attachment_list.append(Attachment(name=ProductDocument_obj.name, url=ProductDocument_obj.document_url))
                        import_log_obj.document_count +=1
                        import_log_obj.save()
                    except Exception as e:
                        dict_error['error-row'] = integer + 2 
                        dict_error['error_list'].append(f"Public URL required Attachment {i} URL")
                        dict_error['product_sku'] = product_obj.sku
                        dict_error['product_name'] = product_obj.name
                    i += 1
                # Process Images (Avoid Duplicates)
                i = 1
                while True:
                    image_count += 1
                    image_name = f"{sku}-Image-{image_count}"
                    image_url = row_dict.get(f"Image {i} URL", "")
                    if not image_url:
                        break
                    # try:
                    ProductImage_obj = DatabaseModel.get_document(ProductImage.objects,{'original_url__in': [str(image_url)],'client_id':client_id})
                    if ProductImage_obj is None:
                        upload_result = cloudinary.uploader.upload(image_url, folder="KM-DIGI/image")
                        image_url_ = upload_result.get("secure_url")
                        public_id = upload_result.get("public_id")
                        product_image = ProductImage(name=image_name, image_url=image_url_, public_id=public_id, client_id=client_id,original_url =[image_url,image_url_] )
                        product_image.save()
                        product_obj.image_list.append(Image(name=image_name, url=image_url_))
                        
                    else:
                        product_obj.image_list.append(Image(name=ProductImage_obj.name, url=ProductImage_obj.image_url))
                    import_log_obj.image_count +=1
                    import_log_obj.save()
                    product_obj.save()
                    # except Exception as e:
                    #     dict_error['error-row'] = integer + 2
                    #     dict_error['error_list'].append(f"Public URL required Image {i} URL")
                    #     dict_error['product_sku'] = product_obj.sku
                    #     dict_error['product_name'] = product_obj.name
                    i += 1

                # Process Videos (Avoid Duplicates)
                i = 1
                while True:
                    video_count += 1
                    video_name = f"{sku}-Video-{video_count}"
                    video_url = row_dict.get(f"Video {i} URL", "")

                    if not video_url:
                        break
                    try:
                        ProductVideo_obj = DatabaseModel.get_document(ProductVideo.objects, {'original_url__in': [video_url],'client_id':client_id})
                        if ProductVideo_obj is None:
                            upload_result = cloudinary.uploader.upload(video_url, resource_type="video", folder="KM-DIGI/video")
                            video_url_ = upload_result.get("secure_url")
                            public_id = upload_result.get("public_id")

                            product_video = ProductVideo(name=video_name, video_url=video_url_, public_id=public_id, client_id=client_id,original_url = [video_url,video_url_])
                            product_video.save()
                            product_obj.video_list.append(Video(name=video_name, url=video_url_))
                        else:
                            product_obj.video_list.append(Video(name=ProductVideo_obj.name, url=ProductVideo_obj.video_url))
                        import_log_obj.video_count +=1
                        import_log_obj.save()
                    except Exception as e:
                        dict_error['error-row'] = integer + 2
                        dict_error['error_list'].append(f"Public URL required Video {i} URL")
                        dict_error['product_sku'] = product_obj.sku
                        dict_error['product_name'] = product_obj.name

                    i += 1
                product_obj.save()
                if url_error:
                    dict_error['is_error'] = True
                else:
                    data['added_count'] +=1

            else:
                dict_error['error-row'] = integer + 2
                dict_error['error_list'].append("Invalid SKU")
                dict_error['is_error'] = True
        if len(dict_error['error_list'])>0:
            data['error_list'].append(dict_error)
            if dict_error['is_error']:
                data['is_error'] = True
                data['error_count'] +=1
        import_log_obj.completed_count +=1
        import_log_obj.inprogress_count -=1
        import_log_obj.save()
    if len(dict_error['error_list'])>0:
        data['status'] = False 
    else:
        data['status'] = True 
    import_log_obj.error_count = len(data['error_list'])
    import_log_obj.un_error_count = count - len(data['error_list'])
    import_log_obj.data = data
    import_log_obj.status = "completed"
    import_log_obj.save()
    data['total_dam'] = count

    return JsonResponse(data,safe=False)


from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

def sampleVendorImportXLSX(request):
    wb = Workbook()
    
    # Sheet 1: Vendors
    ws1 = wb.active
    ws1.title = "Vendors"
    
    headers = [
        "Name", "contact info Email", "contact info Phone", "Business Type", "Website", "Description",
        "Address", "City", "Industry Info", "Tax Info", "Logo",
        "Department 1 Name", "Department 1 Email", "Department 1 Phone Number",
        "Department 2 Name", "Department 2 Email", "Department 2 Phone Number"
    ]
    ws1.append(headers)  
    
    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num, value=header)
        if header == "Name":  
            cell.font = bold_font 
        else:
            cell.font = normal_font 
    
    # Data validation for Business Type
    pipeline = [
        {
            '$group': {
                "_id": None,
                'name': {'$addToSet': '$name'},
            }
        }
    ]
    
    brand_list = list(Business_type.objects.aggregate(*pipeline))
    business_type_options = []
    if brand_list:
        business_type_options = brand_list[0]['name']
    
    business_type_dropdown = DataValidation(
        type="list", formula1=f'"{",".join(business_type_options)}"', allow_blank=True
    )
    
    for row in range(2, 101):
        business_type_dropdown.add(ws1[f"D{row}"])
    
    ws1.add_data_validation(business_type_dropdown)  
    
    # Sheet 2: Field Descriptions
    ws2 = wb.create_sheet(title="Field Descriptions")
    field_data = [
        ["Field Name", "Short Description", "Examples", "Mandatory"],
        ["Name", "The official name of the vendor/supplier", "Lenovo Group Ltd", "Yes"],
        ["Contact info email", "Email of the vendor", "manufacturing@lenovo.com", ""],
        ["contact info phone", "Phone Number of the vendor", "9845678095", ""],
        ["Business Type", "The category of business (Manufacturer, Distributor, Retailer, wholesaler)", "Manufacturer", ""],
        ["Website", "The vendor’s official website URL", "www.lenovo.com", ""],
        ["Description", "A brief summary of the vendor’s business, products, or services.", "A leading global manufacturer of PCs, tablets, and smart devices", ""],
        ["Address", "The physical address of the vendor's business", "1009 Think Place, Morrisville, NC 27560, USA", ""],
        ["City", "The city where the vendor's business is located", "Morrisville", ""],
        ["Industry Information", "A brief summary of the vendor’s industry information", "A leading global manufacturer of PCs, tablets, and smart devices", ""],
        ["Tax Information", "Vendor’s tax details", "65-5432178", ""],
        
        ["Logo", "The vendor’s brand logo (image URL)", "https://i.pinimg.com/736x/70/79/45/707945642fb9ddd80b41fef535ee30f3.jpg", ""],
        ["Department 1 Name", "	The department associated with the vendor", "Sales Team", ""],
        ["Department 1 Email", "	The vendor’s department official email address.", "manufacturing@lenovo.com", ""],
        ["Department 1 Phone Number", "TThe vendor’s department contact number, including country code if required.", "9845678095", ""],
    ]
    
    for row in field_data:
        ws2.append(row)
    
    # Set response type and save workbook
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sample_vendor_importXLSX.xlsx"'
    
    wb.save(response)
    return response

import csv
from django.http import HttpResponse

def sampleVendorImportCSV(request):
    headers = [
        "Name", "Contact Info Email", "Contact Info Phone", "Business Type", "Website"
    ]
    pipeline = [
        {
            '$group': {
                "_id": None,
                'name': {'$addToSet': '$name'},
            }
        }
    ]
    brand_list = list(Business_type.objects.aggregate(*pipeline))
    business_type_options = brand_list[0]['name'] if brand_list else []

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename=sample_vendor_import.csv'
    response.write('\ufeff') 
    writer = csv.writer(response)
    writer.writerow(headers)
    dropdown_row = ["", "", "", "/".join(business_type_options), ""]
    writer.writerow(dropdown_row)

    return response

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.datavalidation import DataValidation
from django.http import HttpResponse

def sampleProductImportExcel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sample Product Import"

    # Headers for Sheet1
    headers = [
        "Name", "Product ID", "MPN", "SKU", "UPC", "EAN", "GTIN", "UNPSC", "Model",
        "Vendor Name", "Brand Name", "Manufacture Name", "Category 1 Name", "Category 2 Name",
        "Category 3 Name", "Breadcrumb", "MSRP", "Selling Price", "Discount Price", "Currency",
        "Country of Origin", "Service Warranty", "Product Warranty", "Application", "Certifications",
        "Compliance", "Prop65", "ESG", "Hazardous", "Short Description", "Personalized Short Description",
        "Long Description", "Personalized Long Description", "Feature 1 Name", "Feature 2 Name",
        "Attribute 1 Name", "Attribute 1 Value", "Attribute 2 Name", "Attribute 2 Value",
        "Related Product 1 URL", "Related Product 1 Name", "Related Product 2 URL", "Related Product 2 Name",
        "Attachment 1 Name", "Attachment 1 URL", "Attachment 2 Name", "Attachment 2 URL",
        "Image 1 Name", "Image 1 URL", "Image 2 Name", "Image 2 URL",
        "Video 1 Name", "Video 1 URL", "Video 2 Name", "Video 2 URL"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        if header in ["Name", "SKU",  "Category 1 Name", "Brand Name"]:
            cell.font = Font(bold=True)
    ws.row_dimensions[1].height = 30
    # Create Sheet2 - "Product Fields"
    ws2 = wb.create_sheet("Product Fields")

    sheet2_headers = ["Field Name", "Short Description", "Examples", "Mandatory"]
    sheet2_data = [
        ["Product Name", "The official name of the product", "Lenovo ThinkPad X1 Carbon Gen 11", "Yes"],
        ["MPN", "Manufacturer part number assigned by the manufacturer", "11HM000BUS", "Yes"],
        ["SKU", "Stock keeping unit, product identifier for inventory tracking", "LEN-X1J11-001", "Yes"],
        ["UPC", "Universal product code used for retail identification", "770813987654", ""],
        ["Product ID", "A unique identifier for the product in the system", "THINKPAD-X1-GEN11", ""],
        ["EAN", "European article number used in global trade", "98713987654", ""],
        ["UNSPC Model", "United Nations Standard Products and Services Code", "43211503 (Laptops)", ""],
        ["Vendor", "The supplier or distributor of the product", "Lenovo", ""],
        ["Brand", "The brand name under which the product is sold", "ThinkPad", "Yes"],
        ["Category 1 Name", "Parent category", "Electronics", "Yes"],
        ["Category 2 Name", "Subcategory of Parent", "Laptops", "Yes"],
        ["Category 3 Name", "Subcategory of Subcategory", "Premium laptops", "Yes"],
        ["Bread Crumb", "Category structure", "Electronics > Laptops > Premium laptops", ""],
        ["Manufacturer", "The company that produces the product", "Lenovo Group Ltd.", ""],
        ["Short Description", "Brief product summary", "Ultralight 14'' business laptop", ""],
        ["Personalized Short Description", "Customized short description", "ThinkPad X1 Carbon Gen 11", ""],
        ["Long Description", "Detailed product description", "The Lenovo ThinkPad X1 Carbon Gen 11...", ""],
        ["Personalized Long Description", "Custom long description", "Power up your productivity...", ""],
        ["Feature 1 Name", "Key product feature", "Fingerprint Reader", ""],
        ["Feature 2 Name", "Key product feature", "Backlit Keyboard", ""],
        ["Attributes 1 Name", "Specific product characteristic", "Color", ""],
        ["Attributes 1 Value", "Value of the attribute", "Black", ""],
        ["Attributes 2 Name", "Specific product characteristic", "Material", ""],
        ["Attributes 2 Value", "Value of the attribute", "Aluminum", ""],
        ["Attachment 1 Name", "Document related to the product", "Lenovo ThinkPad X1 Datasheet", ""],
        ["Attachment 1 URL", "Link to attachment", "https://pdfobject.com/pdf/sample.pdf", ""],
        ["Image 1 Name", "Product image file name", "ThinkPad_X1_image", ""],
        ["Image 1 URL", "Product image URL", "https://p2-ofp.static.pub/fes/cms/2022/09/26/i6zlcap44kafmcywlh54d9rd1wieh1215035.png", ""],
        ["Video 1 Name", "Product video file name", "ThinkPad X1 Carbon Gen 11 Overview", ""],
        ["Video 1 URL", "Product video URL", "https://videos.pexels.com/video-files/3195394/3195394-uhd_2560_1440_25fps.mp4", ""],
        ["Selling Price", "Price at which the product is sold", "2000", ""],
        ["Discount Price", "Price after discount", "1600", ""],
        ["MSRP", "Manufacturer's suggested retail price", "1700", ""],
        ["Currency", "Currency used", "USD", ""],
        ["Service Warranty", "Warranty period for services", "1 Year On-Site Support", ""],
        ["Product Warranty", "Warranty period for product", "3 Years Manufacturer Warranty", ""],
        ["Certifications", "Industry or safety certifications", "Energy Star, EPEAT Gold", ""],
        ["Application", "Intended use", "Business & Professional Use", ""],
        ["GTIN", "Global identifier for the product", "195713987654", ""],
        ["ESG", "Sustainability compliance", "Environmentally friendly packaging", ""],
        ["Hazardous", "Contains hazardous materials", "No", ""],
        ["Compliance", "Regulatory standards met", "RoHS, REACH, FCC, CE", ""],
        ["Prop65", "Contains chemicals listed under Prop 65", "No Prop65 listed substances", ""],
        ["Country Of Origin", "Where the product is manufactured", "China", ""],
    ]

    # Adding headers to Sheet2
    for col_num, header in enumerate(sheet2_headers, start=1):
        cell = ws2.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        ws2.column_dimensions[cell.column_letter].width = 50  # Adjust column width

    # Adding data to Sheet2
    for row_num, row_data in enumerate(sheet2_data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            ws2.cell(row=row_num, column=col_num, value=value)
            
    ws2.row_dimensions[1].height = 30
    # Return the Excel file
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="sample_product_import.xlsx"'
    wb.save(response)

    return response


def obtainCountForUserWiseFiles(request):
    client_id = get_current_client()

    # pipeline = [
    #     {"$match": {"_id": ObjectId(client_id)}},
    #     {"$lookup": {
    #         "from": "product_image",
    #         "localField": "_id",
    #         "foreignField": "client_id",
    #         "as": "product_image"
    #     }},
    #      {
    #         '$unwind': {
    #             'path': '$product_image',
    #             'preserveNullAndEmptyArrays': True
    #         }
    # },
    #     {"$lookup": {
    #         "from": "product_video",
    #         "localField": "_id",
    #         "foreignField": "client_id",
    #         "as": "product_video"
    #     }}, {
    #         '$unwind': {
    #             'path': '$product_video',
    #             'preserveNullAndEmptyArrays': True
    #         }
    # },
    #     {"$lookup": {
    #         "from": "product_document",
    #         "localField": "_id",
    #         "foreignField": "client_id",
    #         "as": "product_document"
    #     }}, {
    #         '$unwind': {
    #             'path': '$product_document',
    #             'preserveNullAndEmptyArrays': True
    #         }
    # },
    #     {"$group": {
    #         "_id": None,
    #         "image_list": {"$addToSet": "$product_image.public_id"},
    #         "video_list": {"$addToSet": "$product_video.public_id"},
    #         "document_list": {"$addToSet": "$product_document.public_id"}
    #     }},{
    #         "$project":{
    #             '_id':0,
    #             "image_count":{'$size':"$image_list"},
    #             "video_count":{'$size':"$video_list"},
    #             "document_count":{'$size':"$document_list"}
    #         }
    #     }
    # ]

    # objects = list(client.objects.aggregate(pipeline))  
    data = dict()
    data['product_count'] = DatabaseModel.count_documents(product.objects,{'client_id':client_id})
    data['image_count'] = DatabaseModel.count_documents(ProductImage.objects,{'client_id':client_id})
    data['video_count'] = DatabaseModel.count_documents(ProductVideo.objects,{'client_id':client_id})
    data['document_count'] = DatabaseModel.count_documents(ProductDocument.objects,{'client_id':client_id})
    data['status'] = True
    return JsonResponse(data,safe=False)

@csrf_exempt
def importCategoryForChannel(request):
    data = {'status': False, 'error_list': [], 'is_error': False,'error_count':0,'added_count':0}
    # client_id = "679ca1bb6a0d7de1dd74ef87"
    client_id = ObjectId(get_current_client())
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    import_log_obj = DatabaseModel.save_documents(import_log,{'user_id':ObjectId(user_login_id),'client_id':client_id,'module_name':'channel-category'})
    channel_name = request.POST.get('channel_name')
    if 'file' not in request.FILES:
        return JsonResponse(data, safe=False)
    file = request.FILES['file']
    if file.name.endswith('.xlsx'):
        df = pd.read_excel(file)
    elif file.name.endswith('.csv') or file.name.endswith('.txt'):
        df = pd.read_csv(file)
    elif file.name.endswith('.ods'):
        df = pd.read_excel(file, engine='odf')
    else:
        return JsonResponse(data, safe=False)
    if df.empty:
        data['is_error'] = True
        data['error'] = "Excel file should not be empty"
        return JsonResponse(data, safe=False)
    count = 0
    data_list = list(df.iterrows()) 
    import_log_obj.total_count = len(data_list)
    import_log_obj.save()
    for index, row in df.iterrows():
        count +=1
        import_log_obj.inprogress_count = import_log_obj.total_count - import_log_obj.completed_count 
        import_log_obj.save()
        dict_error = dict()
        dict_error['error-row'] = index + 2
        dict_error['error_list'] = []
        dict_error['is_error'] = False
        flag = False
        row_dict = row.to_dict()
        categories = []
        for i in range(1, 9): 
            category_value = row_dict.get(f"Category {i}", "").strip() if pd.notna(row_dict.get(f"Category {i}")) else ""
            categories.append(category_value)
        if not categories[0]:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 1 is mandatory")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[2:8])  and not categories[1]:  
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 2 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[3:8]) and not categories[2]: 
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 3 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[4:8]) and not categories[3]:  
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 4 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[5:8]) and not categories[4]:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 5 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  any(categories[6:8]) and not categories[5]:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 6 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        if  categories[7] and not categories[6]:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category 7 Structure is missing")
            dict_error['is_error'] = True
            flag = True
        category_config_obj = None
        category_id_list = []
        if dict_error['is_error'] == False:
            for i in range(8):
                category_name = clean_value(categories[i], default="")
                if category_name:
                    category_name = str(category_name).title()
                    category_obj = DatabaseModel.get_document(category.objects,{'name__iexact':category_name,'client_id':client_id})
                    if category_obj == None:
                        category_obj = DatabaseModel.save_documents(category,{'name':category_name,'client_id':client_id})
                    category_config_obj_past = DatabaseModel.get_document(
                        category_config.objects, {'name__iexact': category_name, 'levels': category_id_list, 'client_id': client_id}
                    )
                    category_id_list.append(category_obj.id)
                    category_config_obj = DatabaseModel.get_document(category_config.objects,{'name__iexact':category_name,'levels':category_id_list})
                    if category_config_obj == None and category_name != "":
                        if category_config_obj_past:
                            category_config_obj_past.end_level = False
                            category_config_obj_past.save()
                        category_config_obj = DatabaseModel.save_documents(category_config,{'name':category_name,'levels':category_id_list,'end_level':True})
        if category_config_obj:
            category_config_id = category_config_obj.id
            channel_categories = []
            for i in range(1, 9): 
                category_value = row_dict.get(f"Channel Category {i}", "").strip() if pd.notna(row_dict.get(f"Channel Category {i}")) else ""
                channel_categories.append(category_value)
            if not channel_categories[0]:
                dict_error['error-row'] = index + 2
                dict_error['error_list'].append("Channel Category 1 is mandatory")
                dict_error['is_error'] = True
                flag = True
            if  any(channel_categories[2:8])  and not channel_categories[1]:  
                dict_error['error-row'] = index + 2
                dict_error['error_list'].append("Channel Category Structure is missing")
                dict_error['is_error'] = True
                flag = True
            if  any(channel_categories[3:8]) and not channel_categories[2]: 
                dict_error['error-row'] = index + 2
                dict_error['error_list'].append("Channel Category Structure is missing")
                dict_error['is_error'] = True
                flag = True
            if  any(channel_categories[4:8]) and not channel_categories[3]:  
                dict_error['error-row'] = index + 2
                dict_error['error_list'].append("Channel Category Structure is missing")
                dict_error['is_error'] = True
                flag = True
            if  any(channel_categories[5:8]) and not channel_categories[4]:
                dict_error['error-row'] = index + 2
                dict_error['error_list'].append("Channel Category Structure is missing")
                dict_error['is_error'] = True
                flag = True
            if  any(channel_categories[6:8]) and not channel_categories[5]:
                dict_error['error-row'] = index + 2
                dict_error['error_list'].append("Channel Category Structure is missing")
                dict_error['is_error'] = True
                flag = True
            if  channel_categories[7] and not channel_categories[6]:
                dict_error['error-row'] = index + 2
                dict_error['error_list'].append("Channel Category Structure is missing")
                dict_error['is_error'] = True
                flag = True
            channel_category_str_list = []
            for i in range(8):
                category_name = clean_value(channel_categories[i], default="")
                if category_name:
                    channel_category_str_list.append(category_name)
            channelCategory_obj = DatabaseModel.get_document(channelCategory.objects,{'channel_name':channel_name,'client_id':ObjectId(client_id),'category_config_id':category_config_id})
            if channelCategory_obj:
                DatabaseModel.update_documents(channelCategory.objects,{'id':channelCategory_obj.id},{'channel_name':channel_name,'client_id':ObjectId(client_id),'category_config_id':category_config_id,'taxonomy_level':channel_category_str_list})
            else:
                DatabaseModel.save_documents(channelCategory,{'channel_name':channel_name,'client_id':ObjectId(client_id),'category_config_id':category_config_id,'taxonomy_level':channel_category_str_list})
                import_log_obj.created_count +=1
                import_log_obj.created_id_list.append(str(category_obj.id)) 
                import_log_obj.save()
        else:
            dict_error['error-row'] = index + 2
            dict_error['error_list'].append("Category Level Is Not Available In The System")
            dict_error['is_error'] = True
            flag = True
        if flag:
            data['error_count'] += 1
            data['error_list'].append(dict_error)
            dict_error['is_error'] = True
        else:
            data['added_count'] += 1
        import_log_obj.completed_count +=1
        import_log_obj.inprogress_count -=1
        import_log_obj.save()
    if data['error_count'] == count:
        data['status'] = False
        data['is_error'] = True
    elif data['error_count'] >0:
        data['status'] = True
        data['is_error'] = True
    else:
        data['status'] = True
    import_log_obj.error_count = len(data['error_list'])
    import_log_obj.un_error_count = count - len(data['error_list'])
    import_log_obj.data = data
    import_log_obj.status = "completed"
    import_log_obj.save()
    data['total_category'] = count
    return JsonResponse(data, safe=False)


@csrf_exempt
def exportChannelCategory(request):
    client_id = get_current_client()
    channel_name = request.GET.get('channel_name')
    pipeline = [
         {
            '$match':{'client_id':ObjectId(client_id),'channel_name':channel_name}
        },
        {
            "$group": {
                "_id": "$_id",
                "taxonomy_level":{ "$first":"$taxonomy_level"},
                "category_config_id":{ "$first":"$category_config_id"},
        }
    },
    {
        '$sort':{'_id':-1}
    }
    ]
    result = list(channelCategory.objects.aggregate(*pipeline))
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Category"
    headers = ["Category 1","Category 2","Category 3","Category 4","Category 5","Category 6","Category 7","Category 8","Channel Category 1","Channel Category 2","Channel Category 3","Channel Category 4","Channel Category 5","Channel Category 6","Channel Category 7","Channel Category 8"]
    worksheet.append(headers)
    for i, item in enumerate(result):
        category_config_obj = DatabaseModel.get_document(category_config.objects,{'id':(item['category_config_id'])})
        category_config_name_count = category_config_obj.levels
        row = []
        for zii in range(8):
            if len(category_config_name_count) > zii:
                row.append(category_config_name_count[zii].name)
            else:
                row.append("")
        taxonomy_level_count = (item['taxonomy_level'])
        for zii in range(8):
            if len(taxonomy_level_count) > zii:
                row.append(taxonomy_level_count[zii])
            else:
                row.append("")
        worksheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0) 
    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="Channel Category export.xlsx"'
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def obtainChannelwiseTaxonomy(request):
    client_id = get_current_client()
    channel_name = request.GET.get('channel_name')
    data = dict()
    data['category_group_list'] = []
    channelCategory_obj = DatabaseModel.list_documents(channelCategory.objects,{'channel_name':channel_name,'client_id':client_id})
    for ihj in channelCategory_obj:
        channelCategoryConfig_obj = DatabaseModel.get_document(
            category_config.objects, {'id': ihj.category_config_id.id, 'client_id': client_id}
        )
        taxonomyconfig_level_str = []
        if channelCategoryConfig_obj:
            for xzz in channelCategoryConfig_obj.levels:
                taxonomyconfig_level_str.append(xzz.name)
        taxonomyconfig_level_str_str = " > ".join(taxonomyconfig_level_str).rstrip(">")
        taxonomy_level_str = " > ".join(ihj.taxonomy_level) if isinstance(ihj.taxonomy_level, list) else str(ihj.taxonomy_level)

        data['category_group_list'].append({
            # 'channel_name': ihj.channel_name,
            'category_taxonomy_level': taxonomy_level_str.rstrip(">"),
            'taxonomy_level': taxonomyconfig_level_str_str
        })
    return data

import logging
logger = logging.getLogger(__name__) 
def obtainDashboard(request):
    client_id = get_current_client()
    data = {
        'data': {
            'vendor_count': 0,
            'product_count': 0,
            'brand_count': 0,
            'category': {
                'total': 0,
                'level1': 0,
                'level2': 0,
                'end_level': 0,
            },
            'product_distribution_by_level1': [],
            'product_completeness': {
                'complete': 0,
                'incomplete': 0
            }
        }
    }
    data['data']['vendor_count'] = Vendor.objects.filter(client_id=client_id).count()
    data['data']['product_count'] = product.objects.filter(client_id=client_id).count()
    data['data']['brand_count'] = brand.objects.filter(client_id=client_id).count()
    categories = category_config.objects.filter(client_id=client_id)
    level1, level2, end_level = 0, 0, 0
    category_list = []
    for cat in categories:
        if len(cat.levels) == 1:
            if cat.name not in  category_list:
                level1 += 1
                category_list.append(cat.name)
        elif len(cat.levels) == 2:
            level2 += 1
        if cat.end_level == True:
            end_level += 1
    data['data']['category']['total'] = categories.count()
    data['data']['category']['level1'] = level1
    data['data']['category']['level2'] = level2
    data['data']['category']['end_level'] = end_level
    pipeline = [
        {"$match": {"client_id": ObjectId(client_id)}},
        {"$project": {
            "sku": 1,
            "filledFields": {
                "$add": [
                                        {
                        "$cond": [
                            {
                            "$or": [
                                { "$ne": [{ "$ifNull": ["$short_description", ""] }, ""] },
                                { "$ne": [{ "$ifNull": ["$personalized_short_description", ""] }, ""] },
                                { "$ne": [{ "$ifNull": ["$long_description", ""] }, ""] },
                                { "$ne": [{ "$ifNull": ["$personalized_long_description", ""] }, ""] }
                            ]
                            },
                            1,
                            0
                        ]
                        },
                    {"$cond": [{ "$gt": [{ "$size": { "$ifNull": ["$feature_list", []] } }, 0] }, 1, 0]},
                    {"$cond": [{ "$gt": [{ "$size": { "$ifNull": ["$attribute_list", []] } }, 0] }, 1, 0]},
                    {"$cond": [{ "$gt": [{ "$size": { "$ifNull": ["$image_list", []] } }, 0] }, 1, 0]},
                ]
            }
        }},
        {"$addFields": {
            "completeness": {
                "$round": [{ "$multiply": [ { "$divide": ["$filledFields", 4] }, 100 ] }, 2]
            }
        }},
        {"$project": {
            "product_id": "$sku",
            "completeness": 1,
            "_id": 0
        }}
    ]

    results = list(product.objects.aggregate(*pipeline))
    data['data']['product_completeness'] = results
    # logger.info(f"Dashboard loaded for client: {client_id} at {datetime.now()}")
    print(">>>>2")
    
    pipeline = [
    {
        "$match": {
            "$expr": {
                "$eq": [{ "$size": "$levels" }, 1]
            },
            'client_id':ObjectId(client_id)
        }
    }
]

    pipeline = [
        {
        "$match": {
            "$expr": {"$eq": [{"$size": "$levels"}, 1]}
        }
    },
    {
        "$lookup": {
            "from": "category",
            "localField": "name",
            "foreignField": "name",
            "as": "matched_category"
        }
    },
    {"$unwind": "$matched_category"},
    {
        "$match": {
            "matched_category.client_id": ObjectId(client_id)
        }
    },
    {
        "$lookup": {
            "from": "category_config",
            "let": {"parent_id": "$matched_category._id"},
            "pipeline": [
                {
                    "$match": {
                        "$expr": {
                            "$in": ["$$parent_id", "$levels"]
                        }
                    }
                }
            ],
            "as": "child_configs"
        }
    },
    {"$unwind": "$child_configs"},
    {
        "$lookup": {
            "from": "product",
            "localField": "child_configs._id",
            "foreignField": "category_id",
            "as": "products"
        }
    },
    {"$unwind": "$products"},
    {
        "$group": {
            "_id": "$name",
            "unique_products": {"$addToSet": "$products._id"}
        }
    },
    {
        "$project": {
            "name": "$_id",
            "product_count": {"$size": "$unique_products"},
            "_id": 0
        }
    }
]

    parent_category_list = list(category_config.objects.aggregate(pipeline))
    data["parent_category_list"] = parent_category_list
    print(">>>>4")

    return JsonResponse(data, safe=False)



def obtainVendorLog(request):

    client_id = get_current_client()
    pipeline = [
        {"$match":{'client_id':ObjectId(client_id)}},{
        '$lookup': {
            'from': 'user',
            'localField': 'user_id',
            'foreignField': '_id',
            'as': 'user'
        }
    }, 
    {
            '$unwind': {
                'path': '$user',
                # 'preserveNullAndEmptyArrays': True
            }
    },{
        '$lookup': {
            'from': 'vendor',
            'localField': 'vendor_id',
            'foreignField': '_id',
            'as': 'vendor'
        }
    }, 
    {
            '$unwind': {
                'path': '$vendor',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user.user_name"},
                "date_time":{'$first':"$logged_date"},
                "action":{'$first':"$action"},
                "vendor_name":{'$first':"$vendor.name"},
                "vendor_id":{'$first':"$vendor.id"}

        }
        },
        {
            '$sort':{'_id':-1}
        }
    ]
    vendor_log_list = list(vendor_log.objects.aggregate(*pipeline))
    data = dict()
    for i in vendor_log_list:
        del i['_id']
        i['vendor_id'] = str(i['vendor_id'])
        utc_time = datetime.strptime(str(i['date_time']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['date'] = us_time.strftime("%Y-%m-%d")
        i['time'] = us_time.strftime("%I:%M:%S %p") 
    data['vendor_log_list'] = vendor_log_list
    return data

def obtainBrandLog(request):
    client_id = get_current_client()
    pipeline = [
        {"$match":{'client_id':ObjectId(client_id)}},{
        '$lookup': {
            'from': 'user',
            'localField': 'user_id',
            'foreignField': '_id',
            'as': 'user'
        }
    }, 
    {
            '$unwind': {
                'path': '$user',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
        '$lookup': {
            'from': 'brand',
            'localField': 'brand_id',
            'foreignField': '_id',
            'as': 'brand'
        }
    }, 
    {
            '$unwind': {
                'path': '$brand',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user.user_name"},
                "date_time":{'$first':"$logged_date"},
                "action":{'$first':"$action"},
                "brand_id":{'$first':"$brand.id"},
                "brand_name":{'$first':"$brand.name"}

        }
        }, {
        '$sort': { '_id': -1 }
    },
    ]
    brand_log_list = list(brand_log.objects.aggregate(*pipeline))
    data = dict()
    for i in brand_log_list:
        del i['_id']
        i['brand_id'] = str(i['brand_id']) 
        utc_time = datetime.strptime(str(i['date_time']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['date'] = us_time.strftime("%Y-%m-%d")
        i['time'] = us_time.strftime("%I:%M:%S %p") 
    data['brand_log_list'] = brand_log_list
    return data


def obtainImportLog(request):
    action = request.GET.get('action')
    if action == 'supplier':
        action = 'vendor'
    if action == "other":
        action_ = ['brand','vendor','category','attribute','dam']
    elif action == "channel":
        action_ = ['channel-category']
    else:
        action_ = [action]
    client_id = get_current_client()
    pipeline = [
        {"$match":{'client_id':ObjectId(client_id),'module_name':{'$in':action_}}},{
        '$lookup': {
            'from': 'user',
            'localField': 'user_id',
            'foreignField': '_id',
            'as': 'user'
        }
    }, 
    {
            '$unwind': {
                'path': '$user',
                
            }
    },
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user.user_name"},
                "date_time":{'$first':"$logged_date"},
                "module_name":{'$first':"$module_name"},
                "status":{'$first':"$status"},
                "error_count":{'$first':"$error_count"},
                "un_error_count":{'$first':"$un_error_count"},
                "completed_count":{'$first':"$completed_count"},
                "inprogress_count":{'$first':"$inprogress_count"},
                "total_count":{'$first':"$total_count"},
                "created_count":{'$first':"$created_count"},
                "created_id_list":{'$first':"$created_id_list"},
                "updated_id_list":{'$first':"$updated_id_list"},
                "updated_count":{'$first':"$updated_count"},
                "image_count":{'$first':"$image_count"},
                "video_count":{'$first':"$video_count"},
                "document_count":{'$first':"$document_count"},
                
                "data":{'$first':"$data"}

        }
        }, {
        '$sort': { '_id': -1 }
    },
    ]
    import_log_list = list(import_log.objects.aggregate(*pipeline))
    data = dict()
    for i in import_log_list:
        del i['_id']
        utc_time = datetime.strptime(str(i['date_time']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['date'] = us_time.strftime("%Y-%m-%d")
        i['time'] = us_time.strftime("%I:%M:%S %p") 
        created_product_list = []
        updated_product_list = []
        if action == "product":
            created_product_list = DatabaseModel.list_documents(product.objects,{'id__in':i['created_id_list']})
            updated_product_list = DatabaseModel.list_documents(product.objects,{'id__in':i['updated_id_list']})
            i['created_id_list'] = []
            i['updated_id_list'] = []
            for ixx in created_product_list:
                i['created_id_list'].append(ixx.sku) 
            for ixx in updated_product_list:
                i['updated_id_list'].append(ixx.sku) 
        elif action == 'brand':
            created_product_list = DatabaseModel.list_documents(brand.objects,{'id__in':i['created_id_list']})
            updated_product_list = DatabaseModel.list_documents(brand.objects,{'id__in':i['updated_id_list']})
            i['created_id_list'] = []
            i['updated_id_list'] = []
            for ixx in created_product_list:
                i['created_id_list'].append(ixx.name) 
            for ixx in updated_product_list:
                i['updated_id_list'].append(ixx.name) 
        elif action == 'vendor':
            created_product_list = DatabaseModel.list_documents(Vendor.objects,{'id__in':i['created_id_list']})
            updated_product_list = DatabaseModel.list_documents(Vendor.objects,{'id__in':i['updated_id_list']}) 
            i['created_id_list'] = []
            i['updated_id_list'] = []
            for ixx in created_product_list:
                i['created_id_list'].append(ixx.name) 
            for ixx in updated_product_list:
                i['updated_id_list'].append(ixx.name) 
        # elif action == 'category':
        #     created_product_list = DatabaseModel.list_documents(category_config.objects,{'id__in':i['created_id_list']})
        #     updated_product_list = DatabaseModel.list_documents(category_config.objects,{'id__in':i['updated_id_list']}) 
        #     i['created_id_list'] = []
        #     i['updated_id_list'] = []
        #     for ixx in created_product_list:
        #         i['created_id_list'].append(ixx.name) 
        #     for ixx in updated_product_list:
        #         i['updated_id_list'].append(ixx.name) 
        elif action == 'attribute':
            created_product_list = DatabaseModel.list_documents(Attribute.objects,{'id__in':i['created_id_list']})
            updated_product_list = DatabaseModel.list_documents(Attribute.objects,{'id__in':i['updated_id_list']})
            i['created_id_list'] = []
            i['updated_id_list'] = []
            for ixx in created_product_list:
                i['created_id_list'].append(ixx.name) 
            for ixx in updated_product_list:
                i['updated_id_list'].append(ixx.name) 
        elif  action == 'dam':
            created_product_list = DatabaseModel.list_documents(product.objects,{'id__in':i['created_id_list']})
            updated_product_list = DatabaseModel.list_documents(product.objects,{'id__in':i['updated_id_list']})
            i['created_id_list'] = []
            i['updated_id_list'] = []
            for ixx in created_product_list:
                i['created_id_list'].append(ixx.sku) 
            for ixx in updated_product_list:
                i['updated_id_list'].append(ixx.sku) 
        elif  action == 'channel':
            created_product_list = DatabaseModel.list_documents(category_config.objects,{'id__in':i['created_id_list']})
            updated_product_list = DatabaseModel.list_documents(category_config.objects,{'id__in':i['updated_id_list']}) 
            i['created_id_list'] = []
            i['updated_id_list'] = []
            for ixx in created_product_list:
                i['created_id_list'].append(ixx.name) 
            for ixx in updated_product_list:
                i['updated_id_list'].append(ixx.name) 
        i['created_id_list'].extend(i['updated_id_list'])
    data['import_log_list'] = import_log_list
    return data

def obtainExportLog(request):
    client_id = get_current_client()
    action = request.GET.get('action')
    if action == 'supplier':
        action = 'vendor'
    if action == "other":
        action = ['brand','vendor','category','attribute','dam']
    elif action == "channel":
        action = ['product-amazon','product-shopify','product-bigcommerce',]
    else:
        action = [action]
    pipeline = [
        {"$match":{'client_id':ObjectId(client_id),'module_name':{'$in':action}}},{
        '$lookup': {
            'from': 'user',
            'localField': 'user_id',
            'foreignField': '_id',
            'as': 'user'
        }
    }, 
    {
            '$unwind': {
                'path': '$user',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user.user_name"},
                "date_time":{'$first':"$logged_date"},
                "module_name":{'$first':"$module_name"},
                "total_count":{'$first':"$total_count"}

        }
        }, {
        '$sort': { '_id': -1 }
    },
    ]
    export_log_list = list(export_log.objects.aggregate(*pipeline))
    data = dict()
    for i in export_log_list:
        del i['_id']
        utc_time = datetime.strptime(str(i['date_time']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['date'] = us_time.strftime("%Y-%m-%d")
        i['time'] = us_time.strftime("%I:%M:%S %p") 
    data['export_log_list'] = export_log_list
    return data

@csrf_exempt
def createClientUserApi(request):
    json_req = JSONParser().parse(request)
    email = json_req['email']
    user_name = json_req['user_name']
    # name = json_req['name']
    role = json_req['role']
    phone = json_req['phone']
    country_code = request.POST.get('country_code')
    if country_code and phone:
        phone = country_code + " "+phone
    client_id = get_current_client()
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    user_obj = DatabaseModel.get_document(user.objects,{'client_id':client_id,'user_name':user_name})
    if user_obj:
        data = dict()
        data['is_created'] = False
        data['error'] = "User Name Already Present"
        return data
    create_user_obj = DatabaseModel.get_document(user.objects,{'client_id':client_id,'id':user_login_id})
    DatabaseModel.save_documents(user,{'email':email,'user_name':user_name,'password':user_name,'role':role,'client_id':ObjectId(client_id),'added_by':create_user_obj.user_name,'phone':phone,'name':user_name})
    data = dict()
    data['is_created'] = True
    return data


def obtainClientUser(request):
    client_id = get_current_client()
    user_id = request.GET.get('id')
    search_term = request.GET.get('search')
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    sort_by = request.GET.get('sort_by')
    sort = request.GET.get('sort')
    if sort_by ==None:
        sort_by = '_id'
        sort = -1
    else:
        if sort == 'true':
            sort = -1
        else:
            sort = 1
    if user_id:
        match_obj = {'client_id':ObjectId(client_id),'_id':ObjectId(user_id)}
    else:
        match_obj = {
    "client_id": ObjectId(client_id),
    "_id": { "$ne": ObjectId(user_login_id) }
}
    pipeline = [
        {"$match":match_obj},
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user_name"},
                "email":{'$first':"$email"},
                "name":{'$first':"$name"},
                "role":{'$first':"$role"},
                "is_active":{'$first':"$is_active"},
                "phone":{'$first':"$phone"},
                "added_by":{'$first':"$added_by"},
                "last_updated":{'$first':"$last_updated"},

        }
        } ,{
    '$match': {
    '$or': [
        { 'name': { '$regex': search_term, '$options': 'i' } },
        { 'role': { '$regex': search_term, '$options': 'i' } },
        { 'user_name': { '$regex': search_term, '$options': 'i' } },
        { 'email': { '$regex': search_term, '$options': 'i' } },
        { 'phone': { '$regex': search_term, '$options': 'i' } },
]
    }
  },{
        '$sort': { sort_by: sort }
    },
    ]
    user_list = list(user.objects.aggregate(*pipeline))
    data = dict()
    for i in user_list:
        i['id'] = str(i['_id'])
        del i['_id']
        if i['last_updated'] == None:
            i['last_updated'] = datetime.now()
        utc_time = datetime.strptime(str(i['last_updated']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['last_updated']  = str(us_time.strftime("%Y-%m-%d") )+" : "+ str(us_time.strftime("%I:%M:%S %p"))
        if i['phone']:
            contact_info_phone = i['phone'].split(" ", 1) 
            if len(contact_info_phone) == 2: 
                i['country_code'] = contact_info_phone[0]
                i['cc_phone'] = i['phone']
                i['phone'] = contact_info_phone[1]
            else:
                i['country_code'] = ""
    data['user_list'] = user_list
    return data

def obtainCategoryLog(request):
    client_id = get_current_client()
    pipeline = [
        {"$match":{'client_id':ObjectId(client_id)}},{
        '$lookup': {
            'from': 'user',
            'localField': 'user_id',
            'foreignField': '_id',
            'as': 'user'
        }
    }, 
    {
            '$unwind': {
                'path': '$user',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
        '$lookup': {
            'from': 'category_config',
            'localField': 'category_config_id',
            'foreignField': '_id',
            'as': 'category_config'
        }
    }, 
    {
            '$unwind': {
                'path': '$category_config',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user.user_name"},
                "date_time":{'$first':"$logged_date"},
                "action":{'$first':"$action"},
                "category_id":{'$first':"$category_config.id"},
                "category_name":{'$first':"$category_config.name"}

        }
        }, {
        '$sort': { '_id': -1 }
    },
    ]
    category_log_list = list(category_log.objects.aggregate(*pipeline))
    data = dict()
    for i in category_log_list:
        del i['_id']
        i['category_id'] = str(i['category_id']) 
        utc_time = datetime.strptime(str(i['date_time']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['date'] = us_time.strftime("%Y-%m-%d")
        i['time'] = us_time.strftime("%I:%M:%S %p") 
    data['category_log_list'] = category_log_list
    return data



def obtainProductLog(request):
    client_id = get_current_client()
    pipeline = [
        {"$match":{'client_id':ObjectId(client_id)}},{
        '$lookup': {
            'from': 'user',
            'localField': 'user_id',
            'foreignField': '_id',
            'as': 'user'
        }
    }, 
    {
            '$unwind': {
                'path': '$user',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
        '$lookup': {
            'from': 'product',
            'localField': 'product_id',
            'foreignField': '_id',
            'as': 'product'
        }
    }, 
    {
            '$unwind': {
                'path': '$product',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user.user_name"},
                "date_time":{'$first':"$logged_date"},
                "action":{'$first':"$action"},
                "product_id":{'$first':"$product.id"},
                "product_name":{'$first':"$product.sku"}

        }
        }, {
        '$sort': { '_id': -1 }
    },
    ]
    product_log_list = list(product_log.objects.aggregate(*pipeline))
    data = dict()
    for i in product_log_list:
        del i['_id']
        i['product_id'] = str(i['product_id']) 
        utc_time = datetime.strptime(str(i['date_time']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['date'] = us_time.strftime("%Y-%m-%d")
        i['time'] = us_time.strftime("%I:%M:%S %p") 
    data['product_log_list'] = product_log_list
    return data


def obtainAttributeLog(request):
    client_id = get_current_client()
    pipeline = [
        {"$match":{'client_id':ObjectId(client_id)}},{
        '$lookup': {
            'from': 'user',
            'localField': 'user_id',
            'foreignField': '_id',
            'as': 'user'
        }
    }, 
    {
            '$unwind': {
                'path': '$user',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
        '$lookup': {
            'from': 'attribute',
            'localField': 'attribute_id',
            'foreignField': '_id',
            'as': 'attribute'
        }
    }, 
    {
            '$unwind': {
                'path': '$attribute',
                # 'preserveNullAndEmptyArrays': True
            }
    },
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user.user_name"},
                "date_time":{'$first':"$logged_date"},
                "action":{'$first':"$action"},
                "attribute_id":{'$first':"$attribute.id"},
                "attribute_name":{'$first':"$attribute.name"},
                "module_name":{'$first':"$module_name"}

        }
        }, {
        '$sort': { '_id': -1 }
    },
    ]
    attribute_log_list = list(attribute_log.objects.aggregate(*pipeline))
    data = dict()
    for i in attribute_log_list:
        del i['_id']
        i['attribute_id'] = str(i['attribute_id']) 
        utc_time = datetime.strptime(str(i['date_time']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['date'] = us_time.strftime("%Y-%m-%d")
        i['time'] = us_time.strftime("%I:%M:%S %p") 
    data['attribute_log_list'] = attribute_log_list
    return data



@csrf_exempt
def createClientForAdmin(request):
    logo = request.FILES.get('logo')
    name = request.POST.get('name')
    location = request.POST.get('location')
    website_url = request.POST.get('website_url')
    designation = request.POST.get('designation')
    status = request.POST.get('status')
    
    name = name.title()
    client_obj = DatabaseModel.get_document(client.objects,{'name':name})
    if client_obj:
        data = dict()
        data['is_created'] = False
        data['error'] = "Client Name Already Exists"
        return data
    if logo:
        upload_result = cloudinary.uploader.upload(logo,folder="KM-DIGI/image")
        logo = upload_result.get("secure_url")
    DatabaseModel.save_documents(client,{'logo':logo,'name':name,'location':location,'website_url':website_url,'designation':designation,'is_active':status})
    data = dict()
    data['is_created'] = True
    return data

def obtainClientuserForAdmin(request):
    client_id = request.GET.get('client_id')
    user_id = request.GET.get('id')
    search_term = request.GET.get('search')
    sort_by = request.GET.get('sort_by')
    sort = request.GET.get('sort')
    if sort_by ==None:
        sort_by = '_id'
        sort = -1
    else:
        if sort == 'true':
            sort = -1
        else:
            sort = 1
    if user_id:
        match_obj = {'client_id':ObjectId(client_id),'_id':ObjectId(user_id)}
    else:
        match_obj = {
    "client_id": ObjectId(client_id) }
    pipeline = [
        {"$match":match_obj},
    {
            '$group': {
                "_id":"$_id",
                "user_name":{'$first':"$user_name"},
                "email":{'$first':"$email"},
                "name":{'$first':"$name"},
                "role":{'$first':"$role"},
                "is_active":{'$first':"$is_active"},
                "phone":{'$first':"$phone"},
                "added_by":{'$first':"$added_by"},
                "last_updated":{'$first':"$last_updated"},

        }
        },
        {
    '$match': {
    '$or': [
        # { 'name': { '$regex': search_term, '$options': 'i' } },
        { 'role': { '$regex': search_term, '$options': 'i' } },
        { 'user_name': { '$regex': search_term, '$options': 'i' } },
        { 'email': { '$regex': search_term, '$options': 'i' } },
        { 'phone': { '$regex': search_term, '$options': 'i' } },
]
    }
  },
  {
        '$sort': { sort_by: sort }
    },
    ]
    user_list = list(user.objects.aggregate(*pipeline))
    data = dict()
    for i in user_list:
        i['id'] = str(i['_id'])
        del i['_id']
        if i['last_updated'] == None:
            i['last_updated'] = datetime.now()
        utc_time = datetime.strptime(str(i['last_updated']), "%Y-%m-%d %H:%M:%S.%f")
        utc_time = pytz.utc.localize(utc_time)
        us_time = utc_time.astimezone(us_timezone)
        i['last_updated']  = str(us_time.strftime("%Y-%m-%d") )+" : "+ str(us_time.strftime("%I:%M:%S %p"))
        if i['phone']:
            contact_info_phone = i['phone'].split(" ", 1) 
            if len(contact_info_phone) == 2: 
                i['country_code'] = contact_info_phone[0]
                i['cc_phone'] = i['phone']
                i['phone'] = contact_info_phone[1]
            else:
                i['country_code'] = ""
    data['user_list'] = user_list
    return data


def obtainClientForAdmin(request):
    client_id = request.GET.get('id')
    search_term = request.GET.get('search')
    sort_by = request.GET.get('sort_by')
    sort = request.GET.get('sort')
    if sort_by ==None:
        sort_by = '_id'
        sort = -1
    else:
        if sort == 'true':
            sort = -1
        else:
            sort = 1
    match_obj = {}
    if client_id:
        match_obj = {'_id':ObjectId(client_id)}
    pipeline = [
         {"$match":match_obj},
    {
            '$group': {
                "_id":"$_id",
                "name":{'$first':"$name"},
                "logo":{'$first':"$logo"},
                "location":{'$first':"$location"},
                "website_url":{'$first':"$website_url"},
                "is_active":{'$first':"$is_active"},
                "designation":{'$first':"$designation"},

        }
        },{
    '$match': {
    '$or': [
        { 'name': { '$regex': search_term, '$options': 'i' } },
        { 'designation': { '$regex': search_term, '$options': 'i' } },
        { 'location': { '$regex': search_term, '$options': 'i' } },
]
    }
  },  {
        '$sort': { sort_by:sort }
    },
    ]
    client_list = list(client.objects.aggregate(*pipeline))
    data = dict()
    for i in client_list:
        i['id'] = str(i['_id'])
        del i['_id']
    data['client_list'] = client_list
    return data


def obtainClientWiseDetailsForAdmin(request):
    client_id = request.GET.get('id')
    data = dict()
    data['data'] = dict()
    data['data']['vendor_count'] = DatabaseModel.count_documents(Vendor.objects,{'client_id':client_id})
    data['data']['product_count'] = DatabaseModel.count_documents(product.objects,{'client_id':client_id})
    data['data']['brand_count'] = DatabaseModel.count_documents(brand.objects,{'client_id':client_id})
    categories = DatabaseModel.list_documents(category_config.objects,{'client_id':client_id})
    level1, level2, end_level = 0, 0, 0
    for cat in categories:
        if len(cat.levels) == 1:
            level1 += 1
        elif len(cat.levels) == 2:
            level2 += 1
        if cat.end_level == True:
            end_level += 1
    data['data']['category'] = dict()
    data['data']['category']['total'] = len(categories)
    data['data']['category']['level1'] = level1
    data['data']['category']['level2'] = level2
    data['data']['category']['end_level'] = end_level
    return data


@csrf_exempt
def import_progress(request):
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    is_first = request.GET.get('is_first')
    client_id = get_current_client()
    if is_first == "true":
        return JsonResponse({
            'status': True,
            'completed_count': 0,
            'total_count': 0,
            'percentage': 0,
            'is_first':False
        })
    else:

        import_obj = import_log.objects.filter(
    user_id=ObjectId(user_login_id),
    client_id=ObjectId(client_id)
).order_by('-id').first()
        if not import_obj:
            return JsonResponse({'status': False, 'error': 'Import log not found'}, status=404)
        total = import_obj.total_count or 1  # avoid divide-by-zero
        completed = import_obj.completed_count or 0
        
        percentage = round((completed / total) * 100, 2)
        return JsonResponse({
            'status': True,
            'completed_count': completed,
            'total_count': total,
            'percentage': percentage,
            'is_first':False
        })
        
@csrf_exempt
def updateClientUser(request):
    json_req = JSONParser().parse(request)
    user_id = json_req['update_obj']['id']
    json_req['update_obj']['last_updated'] = datetime.now()
    del json_req['update_obj']['id']
    if 'country_code' in json_req['update_obj']:
        json_req['update_obj']['phone'] = json_req['update_obj']['country_code'] + " "+ json_req['update_obj']['phone']
        del json_req['update_obj']['country_code']
    user_obj = DatabaseModel.update_documents(user.objects,{'id':user_id},json_req['update_obj'])
    data = dict()
    data['is_updated'] = True
    return data


@csrf_exempt
def createClientUserApiSuperAdmin(request):
    json_req = JSONParser().parse(request)
    email = json_req['email']
    user_name = json_req['user_name']
    # name = json_req['name']
    role = json_req['role']
    phone = json_req['phone']
    country_code = request.POST.get('country_code')
    if country_code and phone:
        phone = country_code + " "+phone
    client_id =  json_req['client_id']
    user_login_id = request.META.get('HTTP_USER_LOGIN_ID')
    user_obj = DatabaseModel.get_document(user.objects,{'client_id':client_id,'user_name':user_name})
    if user_obj:
        data = dict()
        data['is_created'] = False
        data['error'] = "User Name Already Present"
        return data
    create_user_obj = DatabaseModel.get_document(user.objects,{'id':user_login_id})
    DatabaseModel.save_documents(user,{'email':email,'user_name':user_name,'password':user_name,'role':role,'client_id':ObjectId(client_id),'added_by':create_user_obj.user_name,'phone':phone,'name':user_name})
    data = dict()
    data['is_created'] = True
    return data

@csrf_exempt
def updateClient(request):
    json_req = JSONParser().parse(request)
    client_id = json_req['update_obj']['id']
    del json_req['update_obj']['id']
    DatabaseModel.update_documents(client.objects,{'id':client_id},json_req['update_obj'])
    data = dict()
    data['is_updated'] = True
    return data


